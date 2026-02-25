"""
Blueprint for POD (Registre des Pourboires) management.
Upload, view, edit, and generate bi-weekly tip distribution sheets.
Dual storage: Excel file (for compatibility) + SQLite DB (for persistence/ETL).
"""

from flask import Blueprint, request, jsonify, send_file, render_template
from routes.checklist import login_required
from database import db, PODPeriod, PODEntry
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, date as date_type
from copy import copy as copy_style
import json, io, os, re, ast, operator

pod_bp = Blueprint('pod', __name__)

# ── Safe arithmetic evaluator (replaces eval) ────────────────
_SAFE_OPS = {
    ast.Add: operator.add, ast.Sub: operator.sub,
    ast.Mult: operator.mul, ast.Div: operator.truediv,
    ast.USub: operator.neg, ast.UAdd: operator.pos,
}

def _safe_arithmetic(expr):
    """Evaluate simple arithmetic expressions without eval()."""
    tree = ast.parse(expr.strip(), mode='eval')
    def _eval(node):
        if isinstance(node, ast.Expression):
            return _eval(node.body)
        elif isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return node.value
        elif isinstance(node, ast.BinOp) and type(node.op) in _SAFE_OPS:
            return _SAFE_OPS[type(node.op)](_eval(node.left), _eval(node.right))
        elif isinstance(node, ast.UnaryOp) and type(node.op) in _SAFE_OPS:
            return _SAFE_OPS[type(node.op)](_eval(node.operand))
        raise ValueError(f'Unsafe node: {type(node).__name__}')
    return _eval(tree)

# ── Storage dir for uploaded POD files ──────────────────────────
POD_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'database', 'pod_files')
os.makedirs(POD_DIR, exist_ok=True)

# ── Constants ───────────────────────────────────────────────────
EMPLOYEE_START_ROW = 9      # First employee in Sommaire
SENTINEL_TEXT = 'Insérez avant cette ligne'
DAYS_PER_PERIOD = 14
COLS_PER_DAY = 4            # Ventes, Pourb, Dist, Reçus
DAY1_COL = 4                # Column D = first day Ventes
SUMMARY_COLS = {
    'ventes_total': 61,     # BI
    'pourb_bruts': 62,      # BJ
    'pourb_redist': 63,     # BK
    'pourb_recus': 64,      # BL
    'net_sem1': 65,          # BM
    'net_sem2': 66,          # BN
    'pourb_nets': 67,        # BO
    'montant_attribue': 68,  # BP
}


def _col_for_day(day_index, field_offset=0):
    """Return 1-based column number for a given day (0-13) and field (0=Ventes,1=Pourb,2=Dist,3=Reçus)."""
    return DAY1_COL + day_index * COLS_PER_DAY + field_offset


def _eval_cell(val):
    """Safely evaluate a cell value — handles numbers, simple formulas like =261.81+2321.48, or None."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return 0
    if s.startswith('='):
        expr = s[1:]
        # Safe arithmetic: only allow digits, operators, parens
        if re.match(r'^[\d\.\+\-\*\/\(\)\s]+$', expr):
            try:
                return float(_safe_arithmetic(expr))
            except Exception:
                return 0
        return 0
    try:
        return float(s)
    except:
        return 0


def _parse_pod(filepath):
    """Parse a POD xlsx and return structured data.
    Uses both data_only (for summary formulas with cell refs) and formula mode (for daily data)."""

    # Formula workbook — for daily cells which may have simple formulas like =261.81+2321.48
    wb_f = load_workbook(filepath)
    ws_f = wb_f['Sommaire']

    # Data-only workbook — for summary columns with complex cross-cell formulas
    wb_d = load_workbook(filepath, data_only=True)
    ws_d = wb_d['Sommaire']

    # Period start date
    period_start = ws_d.cell(5, 4).value  # D5
    if period_start is None:
        period_start = ws_f.cell(5, 4).value
    if isinstance(period_start, datetime):
        period_start_str = period_start.strftime('%Y-%m-%d')
    else:
        period_start_str = ''

    # Build date labels
    dates = []
    base_date = period_start if isinstance(period_start, datetime) else None
    if base_date:
        for i in range(DAYS_PER_PERIOD):
            d = base_date + timedelta(days=i)
            dates.append(d.strftime('%Y-%m-%d'))

    # Parse employees
    employees = []
    for r in range(EMPLOYEE_START_ROW, ws_f.max_row + 1):
        emp_id = ws_f.cell(r, 1).value  # A
        name = ws_f.cell(r, 2).value    # B

        if emp_id and str(emp_id).strip() == SENTINEL_TEXT:
            break
        if str(emp_id or '').strip() in ('', 'controle', 'Jour', 'mois', 'année'):
            if not name:
                continue
        if not (emp_id or name):
            continue

        # Daily data — use formula workbook + _eval_cell for simple formulas
        daily = []
        for d in range(DAYS_PER_PERIOD):
            ventes = _eval_cell(ws_f.cell(r, _col_for_day(d, 0)).value)
            pourb = _eval_cell(ws_f.cell(r, _col_for_day(d, 1)).value)
            dist = _eval_cell(ws_f.cell(r, _col_for_day(d, 2)).value)
            recus = _eval_cell(ws_f.cell(r, _col_for_day(d, 3)).value)
            daily.append({
                'ventes': round(ventes, 2),
                'pourb': round(pourb, 2),
                'dist': round(dist, 2),
                'recus': round(recus, 2),
            })

        # Summary — use data_only workbook (has cached values from Excel)
        summary = {}
        for key, col in SUMMARY_COLS.items():
            val = ws_d.cell(r, col).value
            summary[key] = round(float(val), 2) if val else 0

        employees.append({
            'row': r,
            'emp_id': str(emp_id or ''),
            'name': str(name or '').strip(),
            'daily': daily,
            'summary': summary,
        })

    wb_f.close()
    wb_d.close()

    return {
        'period_start': period_start_str,
        'dates': dates,
        'employees': employees,
    }


def _sync_pod_to_db(parsed_data, source_filename=None):
    """Sync parsed POD data into SQLite. Creates or updates PODPeriod + PODEntry rows."""
    period_start_str = parsed_data.get('period_start', '')
    if not period_start_str:
        return

    try:
        period_start = datetime.strptime(period_start_str, '%Y-%m-%d').date()
    except ValueError:
        return

    period_end = period_start + timedelta(days=DAYS_PER_PERIOD - 1)

    # Find or create period
    period = PODPeriod.query.filter_by(period_start=period_start).first()
    if not period:
        period = PODPeriod(
            period_start=period_start,
            period_end=period_end,
            source_filename=source_filename
        )
        db.session.add(period)
        db.session.flush()
    else:
        period.updated_at = datetime.utcnow()
        if source_filename:
            period.source_filename = source_filename

    # Delete existing entries for this period and re-insert
    PODEntry.query.filter_by(period_id=period.id).delete()

    dates = parsed_data.get('dates', [])
    for emp in parsed_data.get('employees', []):
        emp_name = emp.get('name', '').strip()
        if not emp_name:
            continue
        emp_id = emp.get('emp_id', '')
        summary = emp.get('summary', {})

        for d_idx, day_data in enumerate(emp.get('daily', [])):
            # Skip days with no data
            if all(day_data.get(k, 0) == 0 for k in ('ventes', 'pourb', 'dist', 'recus')):
                continue

            day_date = None
            if d_idx < len(dates):
                try:
                    day_date = datetime.strptime(dates[d_idx], '%Y-%m-%d').date()
                except ValueError:
                    pass

            entry = PODEntry(
                period_id=period.id,
                emp_id=emp_id,
                emp_name=emp_name,
                day_index=d_idx,
                day_date=day_date,
                ventes=day_data.get('ventes', 0),
                pourb=day_data.get('pourb', 0),
                dist=day_data.get('dist', 0),
                recus=day_data.get('recus', 0),
                ventes_total=summary.get('ventes_total', 0),
                pourb_bruts=summary.get('pourb_bruts', 0),
                pourb_redist=summary.get('pourb_redist', 0),
                pourb_recus=summary.get('pourb_recus', 0),
                pourb_nets=summary.get('pourb_nets', 0),
            )
            db.session.add(entry)

    db.session.commit()


def _get_active_pod_path():
    """Return path to the most recently uploaded POD file, or None."""
    meta_file = os.path.join(POD_DIR, 'active.json')
    if os.path.exists(meta_file):
        with open(meta_file) as f:
            meta = json.load(f)
        path = os.path.join(POD_DIR, meta.get('filename', ''))
        if os.path.exists(path):
            return path
    return None


def _set_active_pod(filename):
    meta_file = os.path.join(POD_DIR, 'active.json')
    with open(meta_file, 'w') as f:
        json.dump({'filename': filename, 'uploaded_at': datetime.now().isoformat()}, f)


# ── Routes ──────────────────────────────────────────────────────

@pod_bp.route('/pod')
@login_required
def pod_page():
    return render_template('pod.html')


@pod_bp.route('/api/pod/upload', methods=['POST'])
@login_required
def upload_pod():
    """Upload a POD xlsx file. Saves to disk + syncs to DB."""
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier fourni'}), 400

    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Format invalide. Utilisez un fichier .xlsx'}), 400

    # Size check (max 10 MB)
    f.seek(0, 2)
    size = f.tell()
    f.seek(0)
    if size > 10 * 1024 * 1024:
        return jsonify({'error': 'Fichier trop volumineux (max 10 Mo)'}), 400

    filename = f'POD_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(POD_DIR, filename)
    f.save(filepath)
    _set_active_pod(filename)

    data = _parse_pod(filepath)

    # ★ Auto-sync to database
    _sync_pod_to_db(data, source_filename=filename)

    return jsonify({'success': True, 'data': data, 'filename': filename})


@pod_bp.route('/api/pod/load')
@login_required
def load_pod():
    """Load the active POD file data."""
    path = _get_active_pod_path()
    if not path:
        return jsonify({'loaded': False, 'message': 'Aucun fichier POD actif'})

    data = _parse_pod(path)
    return jsonify({'loaded': True, 'data': data})


@pod_bp.route('/api/pod/employees')
@login_required
def get_employees():
    """Return employee list for autocomplete."""
    path = _get_active_pod_path()
    if not path:
        return jsonify({'employees': []})

    wb = load_workbook(path, data_only=True)
    ws = wb['Sommaire']
    employees = []
    for r in range(EMPLOYEE_START_ROW, ws.max_row + 1):
        emp_id = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        if emp_id and str(emp_id).strip() == SENTINEL_TEXT:
            break
        if emp_id and name and str(emp_id).strip() not in ('controle', 'Jour', 'mois', 'année', ''):
            employees.append({'emp_id': str(emp_id), 'name': str(name).strip(), 'row': r})
    wb.close()
    return jsonify({'employees': employees})


@pod_bp.route('/api/pod/save-day', methods=['POST'])
@login_required
def save_day():
    """
    Save tips data for a specific day.
    Expects JSON: { day_index: 0-13, entries: [{row, ventes, pourb, dist, recus}, ...] }
    Writes to Excel + syncs to DB.
    """
    path = _get_active_pod_path()
    if not path:
        return jsonify({'error': 'Aucun fichier POD actif'}), 400

    data = request.get_json()
    day_idx = data.get('day_index', 0)
    entries = data.get('entries', [])

    if not (0 <= day_idx < DAYS_PER_PERIOD):
        return jsonify({'error': 'Index de jour invalide'}), 400

    wb = load_workbook(path)
    ws = wb['Sommaire']

    for entry in entries:
        row = entry.get('row')
        if not row:
            continue

        ventes = entry.get('ventes', 0) or 0
        pourb = entry.get('pourb', 0) or 0
        dist = entry.get('dist', 0) or 0
        recus = entry.get('recus', 0) or 0

        v_col = _col_for_day(day_idx, 0)
        p_col = _col_for_day(day_idx, 1)
        d_col = _col_for_day(day_idx, 2)
        r_col = _col_for_day(day_idx, 3)

        ws.cell(row, v_col, float(ventes) if ventes else None)
        ws.cell(row, p_col, float(pourb) if pourb else None)
        ws.cell(row, d_col, float(dist) if dist else None)
        ws.cell(row, r_col, float(recus) if recus else None)

    wb.save(path)
    wb.close()

    # Re-read and sync to DB
    updated = _parse_pod(path)
    _sync_pod_to_db(updated)

    return jsonify({'success': True, 'data': updated})


@pod_bp.route('/api/pod/save-batch', methods=['POST'])
@login_required
def save_batch():
    """
    Batch save: apply same ventes amount to multiple employees for a day.
    JSON: { day_index, ventes, rows: [row1, row2, ...], auto_tips: true }
    """
    path = _get_active_pod_path()
    if not path:
        return jsonify({'error': 'Aucun fichier POD actif'}), 400

    data = request.get_json()
    day_idx = data.get('day_index', 0)
    ventes = float(data.get('ventes', 0) or 0)
    rows = data.get('rows', [])
    auto_tips = data.get('auto_tips', True)

    if not (0 <= day_idx < DAYS_PER_PERIOD):
        return jsonify({'error': 'Index de jour invalide'}), 400

    pourb = round(ventes * 0.10, 2) if auto_tips else 0

    wb = load_workbook(path)
    ws = wb['Sommaire']

    for row in rows:
        v_col = _col_for_day(day_idx, 0)
        p_col = _col_for_day(day_idx, 1)
        ws.cell(row, v_col, ventes)
        ws.cell(row, p_col, pourb)

    wb.save(path)
    wb.close()

    # Re-read and sync to DB
    updated = _parse_pod(path)
    _sync_pod_to_db(updated)

    return jsonify({'success': True, 'data': updated})


@pod_bp.route('/api/pod/generate-new', methods=['POST'])
@login_required
def generate_new_period():
    """
    Generate a new POD file for the next pay period.
    Copies the template structure but clears all daily data and sets new dates.
    """
    path = _get_active_pod_path()
    if not path:
        return jsonify({'error': 'Aucun fichier POD actif — uploadez d\'abord un fichier'}), 400

    data = request.get_json() or {}
    new_start_str = data.get('start_date')

    # If no date given, calculate next period (current + 14 days)
    wb_read = load_workbook(path, data_only=True)
    ws_read = wb_read['Sommaire']
    current_start = ws_read.cell(5, 4).value
    wb_read.close()

    if new_start_str:
        new_start = datetime.strptime(new_start_str, '%Y-%m-%d')
    elif isinstance(current_start, datetime):
        new_start = current_start + timedelta(days=14)
    else:
        return jsonify({'error': 'Impossible de déterminer la prochaine période'}), 400

    # Copy the existing file as template
    wb = load_workbook(path)
    ws = wb['Sommaire']

    # Update start date
    ws.cell(5, 4, new_start)

    # Clear all daily data (keep employee info)
    for r in range(EMPLOYEE_START_ROW, ws.max_row + 1):
        emp_id = ws.cell(r, 1).value
        if emp_id and str(emp_id).strip() == SENTINEL_TEXT:
            break
        if str(emp_id or '').strip() in ('controle', 'Jour', 'mois', 'année'):
            continue
        if not (emp_id or ws.cell(r, 2).value):
            continue

        for d in range(DAYS_PER_PERIOD):
            for offset in range(COLS_PER_DAY):
                col = _col_for_day(d, offset)
                ws.cell(r, col).value = None
                ws.cell(r, col).data_type = 'n'

    # Also update Feuil1 period date
    if 'Feuil1' in wb.sheetnames:
        ws2 = wb['Feuil1']
        for r in range(3, ws2.max_row + 1):
            if ws2.cell(r, 6).value:
                ws2.cell(r, 6, new_start)

    filename = f'POD{new_start.strftime("%m%d%Y")}.xlsx'
    filepath = os.path.join(POD_DIR, filename)
    wb.save(filepath)
    wb.close()

    _set_active_pod(filename)

    new_data = _parse_pod(filepath)

    # ★ Sync new (empty) period to DB
    _sync_pod_to_db(new_data, source_filename=filename)

    return jsonify({'success': True, 'data': new_data, 'filename': filename})


@pod_bp.route('/api/pod/download')
@login_required
def download_pod():
    """Download the active POD file."""
    path = _get_active_pod_path()
    if not path:
        return jsonify({'error': 'Aucun fichier POD actif'}), 404

    return send_file(
        path,
        as_attachment=True,
        download_name=os.path.basename(path),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ── ETL: Historique POD ────────────────────────────────────────
@pod_bp.route('/api/pod/history')
@login_required
def pod_history():
    """Return all POD periods and their entry counts for ETL/reporting."""
    periods = PODPeriod.query.order_by(PODPeriod.period_start.desc()).all()
    return jsonify({
        'periods': [p.to_dict() for p in periods],
        'total_entries': PODEntry.query.count(),
    })


@pod_bp.route('/api/pod/history/<int:period_id>')
@login_required
def pod_period_detail(period_id):
    """Return all entries for a specific POD period."""
    period = PODPeriod.query.get_or_404(period_id)
    entries = PODEntry.query.filter_by(period_id=period.id).order_by(PODEntry.emp_name, PODEntry.day_index).all()
    return jsonify({
        'period': period.to_dict(),
        'entries': [e.to_dict() for e in entries],
    })
