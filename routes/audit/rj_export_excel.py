"""
RJ Natif — Excel Export (XLSX)
Exports a NightAuditSession to a properly formatted Excel workbook (.xlsx)
matching the original Excel RJ format with 14 sheets covering all audit data.

Features:
  - 14 sheets matching RJ Natif tabs (Contrôle, DueBack, Recap, etc.)
  - Professional styling: Sheraton blue headers, auto-width columns, frozen panes
  - Currency/percentage formatting
  - Proper JSON parsing for variable-structure data
  - Batch export support (date range to ZIP)
  - Handles None/null values gracefully
"""

import io
import json
import zipfile
from datetime import datetime, date as date_type
from dateutil.parser import parse as parse_date

from flask import Blueprint, jsonify, send_file, request
from database.models import db, NightAuditSession

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

rj_excel_bp = Blueprint('rj_excel', __name__, url_prefix='/api/rj/export')

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CONSTANTS & STYLING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

SHERATON_BLUE = "003B71"  # Primary Sheraton blue
DARK_GRAY = "2C2C2C"
LIGHT_GRAY = "F5F5F5"
WHITE = "FFFFFF"
GREEN = "27AE60"
RED = "C0392B"
YELLOW = "F39C12"

# Excel format strings
CURRENCY_FORMAT = '$#,##0.00'
PERCENTAGE_FORMAT = '0.0%'
INTEGER_FORMAT = '#,##0'
DATE_FORMAT = 'YYYY-MM-DD'

# Styling helpers
def _header_fill():
    return PatternFill(start_color=SHERATON_BLUE, end_color=SHERATON_BLUE, fill_type='solid')

def _header_font():
    return Font(bold=True, color=WHITE, size=11)

def _subheader_fill():
    return PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')

def _subheader_font():
    return Font(bold=True, size=10)

def _total_fill():
    return PatternFill(start_color="E8F0F5", end_color="E8F0F5", fill_type='solid')

def _total_font():
    return Font(bold=True, size=10)

def _center_align():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def _right_align():
    return Alignment(horizontal='right', vertical='center')

def _left_align():
    return Alignment(horizontal='left', vertical='center')

def _border():
    thin = Side(style='thin', color='CCCCCC')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _apply_header(ws, row_num, headers):
    """Apply header styling to a row."""
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.alignment = _center_align()
        cell.border = _border()

def _apply_total_row(ws, row_num, value_col=None):
    """Apply total row styling to an entire row."""
    for col in ws.iter_cols(min_row=row_num, max_row=row_num, min_col=1, max_col=ws.max_column):
        if col[0].value is not None:
            col[0].fill = _total_fill()
            col[0].font = _total_font()
            col[0].border = _border()

def _safe_float(val):
    """Convert value to float safely, return 0 if None/invalid."""
    if val is None: return 0
    try: return float(val)
    except (ValueError, TypeError): return 0

def _safe_int(val):
    """Convert value to int safely, return 0 if None/invalid."""
    if val is None: return 0
    try: return int(float(val))
    except (ValueError, TypeError): return 0

def _safe_str(val):
    """Convert value to string safely, return empty if None."""
    if val is None: return ''
    return str(val)

def _set_column_width(ws, col_num, width=15):
    """Set column width."""
    ws.column_dimensions[get_column_letter(col_num)].width = width

def _freeze_panes(ws, row, col):
    """Freeze panes at specified row and column."""
    ws.freeze_panes = f"{get_column_letter(col)}{row}"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET BUILDERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _build_controle(wb, session):
    """Sheet 1: Contrôle — Date, auditor, weather, chambres à refaire"""
    ws = wb.create_sheet("Contrôle", 0)

    _apply_header(ws, 1, ["Contrôle"])
    ws.merge_cells('A1:B1')

    row = 3
    pairs = [
        ("Date d'audit", session.audit_date.isoformat() if session.audit_date else ''),
        ("Auditeur", _safe_str(session.auditor_name)),
        ("Température", _safe_str(session.temperature)),
        ("Condition météo", _safe_str(session.weather_condition)),
        ("Chambres à refaire", _safe_int(session.chambres_refaire)),
        ("Jours dans mois", _safe_int(session.jours_dans_mois)),
        ("Statut", _safe_str(session.status)),
        ("Notes", _safe_str(session.notes)),
    ]

    for label, value in pairs:
        ws.cell(row=row, column=1, value=label).font = _subheader_font()
        ws.cell(row=row, column=2, value=value)
        row += 1

    _set_column_width(ws, 1, 20)
    _set_column_width(ws, 2, 30)


def _build_dueback(wb, session):
    """Sheet 2: DueBack — Dynamic rows: receptionist + previous + nouveau"""
    ws = wb.create_sheet("DueBack", 1)

    _apply_header(ws, 1, ["Réceptionniste", "Précédent", "Nouveau"])
    _set_column_width(ws, 1, 20)
    _set_column_width(ws, 2, 15)
    _set_column_width(ws, 3, 15)

    row = 2
    entries = session.get_json('dueback_entries')
    if isinstance(entries, list):
        for entry in entries:
            ws.cell(row=row, column=1, value=_safe_str(entry.get('name', '')))
            val_prev = _safe_float(entry.get('previous', 0))
            val_new = _safe_float(entry.get('nouveau', 0))

            cell_prev = ws.cell(row=row, column=2, value=val_prev)
            cell_new = ws.cell(row=row, column=3, value=val_new)
            cell_prev.number_format = CURRENCY_FORMAT
            cell_new.number_format = CURRENCY_FORMAT

            row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL")
    total_cell = ws.cell(row=row, column=3, value=_safe_float(session.dueback_total))
    total_cell.number_format = CURRENCY_FORMAT
    _apply_total_row(ws, row)


def _build_recap(wb, session):
    """Sheet 3: Recap — 8 lines (lecture/correction/net) + deposits"""
    ws = wb.create_sheet("Recap", 2)

    _apply_header(ws, 1, ["Catégorie", "Lecture", "Correction", "Net"])
    _set_column_width(ws, 1, 25)
    for col in range(2, 5):
        _set_column_width(ws, col, 15)

    row = 2
    recap_items = [
        ("Cash LS", session.cash_ls_lecture, session.cash_ls_corr),
        ("Cash POS", session.cash_pos_lecture, session.cash_pos_corr),
        ("Chèque AR", session.cheque_ar_lecture, session.cheque_ar_corr),
        ("Chèque DR", session.cheque_dr_lecture, session.cheque_dr_corr),
        ("Remb. Gratuite", session.remb_gratuite_lecture, session.remb_gratuite_corr),
        ("Remb. Client", session.remb_client_lecture, session.remb_client_corr),
        ("DueBack Réception", session.dueback_reception_lecture, session.dueback_reception_corr),
        ("DueBack NB", session.dueback_nb_lecture, session.dueback_nb_corr),
    ]

    for label, lecture, correction in recap_items:
        ws.cell(row=row, column=1, value=label)

        lecture_val = _safe_float(lecture)
        corr_val = _safe_float(correction)
        net_val = lecture_val + corr_val

        c_lec = ws.cell(row=row, column=2, value=lecture_val)
        c_corr = ws.cell(row=row, column=3, value=corr_val)
        c_net = ws.cell(row=row, column=4, value=net_val)

        for c in [c_lec, c_corr, c_net]:
            c.number_format = CURRENCY_FORMAT

        row += 1

    # Deposits
    row += 1
    ws.cell(row=row, column=1, value="Dépôt CDN")
    c_cdn = ws.cell(row=row, column=2, value=_safe_float(session.deposit_cdn))
    c_cdn.number_format = CURRENCY_FORMAT
    row += 1

    ws.cell(row=row, column=1, value="Dépôt USD")
    c_usd = ws.cell(row=row, column=2, value=_safe_float(session.deposit_us))
    c_usd.number_format = CURRENCY_FORMAT
    row += 1

    # Balance
    row += 1
    ws.cell(row=row, column=1, value="SOLDE RECAP").font = _total_font()
    c_balance = ws.cell(row=row, column=2, value=_safe_float(session.recap_balance))
    c_balance.number_format = CURRENCY_FORMAT
    c_balance.fill = _total_fill()
    c_balance.font = _total_font()
    _apply_total_row(ws, row)


def _build_transelect(wb, session):
    """Sheet 4: Transelect — Restaurant matrix + Reception matrix"""
    ws = wb.create_sheet("Transelect", 3)

    # Restaurant section
    _apply_header(ws, 1, ["RESTAURANT — Terminal × Carte"])
    ws.merge_cells('A1:G1')

    row = 2
    rest = session.get_json('transelect_restaurant')
    card_types = ['debit', 'visa', 'mc', 'amex', 'discover']

    # Restaurant header
    headers = ["Terminal"] + [ct.upper() for ct in card_types] + ["Total"]
    _apply_header(ws, row, headers)
    row += 1

    rest_totals = {ct: 0 for ct in card_types}
    if isinstance(rest, dict):
        for term_name, term_data in sorted(rest.items()):
            ws.cell(row=row, column=1, value=term_name)
            for col_idx, card_type in enumerate(card_types, 2):
                val = _safe_float(term_data.get(card_type, 0) if isinstance(term_data, dict) else 0)
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.number_format = CURRENCY_FORMAT
                rest_totals[card_type] += val

            total_val = sum(_safe_float(term_data.get(ct, 0) if isinstance(term_data, dict) else 0)
                          for ct in card_types)
            cell_total = ws.cell(row=row, column=len(card_types)+2, value=total_val)
            cell_total.number_format = CURRENCY_FORMAT
            row += 1

    # Restaurant totals
    ws.cell(row=row, column=1, value="TOTAL").font = _total_font()
    for col_idx, card_type in enumerate(card_types, 2):
        cell = ws.cell(row=row, column=col_idx, value=rest_totals[card_type])
        cell.number_format = CURRENCY_FORMAT
        cell.fill = _total_fill()
        cell.font = _total_font()

    grand_total = sum(rest_totals.values())
    cell_grand = ws.cell(row=row, column=len(card_types)+2, value=grand_total)
    cell_grand.number_format = CURRENCY_FORMAT
    cell_grand.fill = _total_fill()
    cell_grand.font = _total_font()

    # Reception section
    row += 3
    _apply_header(ws, row, ["RÉCEPTION — Carte × Terminal"])
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    recep = session.get_json('transelect_reception')

    # Reception header
    headers = ["Carte", "Fusebox", "Term8", "K053", "Daily Rev", "Total"]
    _apply_header(ws, row, headers)
    row += 1

    rec_totals = {ct: 0 for ct in card_types}
    if isinstance(recep, dict):
        for card_type in card_types:
            ct_data = recep.get(card_type, {})
            if not isinstance(ct_data, dict):
                ct_data = {}

            ws.cell(row=row, column=1, value=card_type.upper())

            fusebox = _safe_float(ct_data.get('fusebox', 0))
            term8 = _safe_float(ct_data.get('term8', 0))
            k053 = _safe_float(ct_data.get('k053', 0))
            daily_rev = _safe_float(ct_data.get('daily_rev', 0))
            total_rec = fusebox + term8 + k053

            for col_idx, val in enumerate([fusebox, term8, k053, daily_rev, total_rec], 2):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.number_format = CURRENCY_FORMAT

            rec_totals[card_type] = total_rec
            row += 1

    # Column widths
    _set_column_width(ws, 1, 15)
    for col in range(2, 8):
        _set_column_width(ws, col, 15)


def _build_geac(wb, session):
    """Sheet 5: GEAC — Cashout vs Daily Revenue by card + AR balance"""
    ws = wb.create_sheet("GEAC", 4)

    _apply_header(ws, 1, ["GEAC — Cashout vs Daily Revenue"])
    ws.merge_cells('A1:C1')

    row = 2
    geac_co = session.get_json('geac_cashout')
    geac_dr = session.get_json('geac_daily_rev')
    card_types = ['debit', 'visa', 'mc', 'amex', 'discover']

    _apply_header(ws, row, ["Carte", "Cashout", "Daily Rev", "Variance"])
    _set_column_width(ws, 1, 15)
    for col in range(2, 5):
        _set_column_width(ws, col, 15)
    row += 1

    total_co = 0
    total_dr = 0
    for card_type in card_types:
        ws.cell(row=row, column=1, value=card_type.upper())

        co_val = _safe_float(geac_co.get(card_type, 0) if isinstance(geac_co, dict) else 0)
        dr_val = _safe_float(geac_dr.get(card_type, 0) if isinstance(geac_dr, dict) else 0)
        var_val = co_val - dr_val

        c_co = ws.cell(row=row, column=2, value=co_val)
        c_dr = ws.cell(row=row, column=3, value=dr_val)
        c_var = ws.cell(row=row, column=4, value=var_val)

        for c in [c_co, c_dr, c_var]:
            c.number_format = CURRENCY_FORMAT

        total_co += co_val
        total_dr += dr_val
        row += 1

    # Totals
    ws.cell(row=row, column=1, value="TOTAL").font = _total_font()
    c_tot_co = ws.cell(row=row, column=2, value=total_co)
    c_tot_dr = ws.cell(row=row, column=3, value=total_dr)
    c_tot_var = ws.cell(row=row, column=4, value=total_co - total_dr)
    for c in [c_tot_co, c_tot_dr, c_tot_var]:
        c.number_format = CURRENCY_FORMAT
        c.fill = _total_fill()
        c.font = _total_font()
    _apply_total_row(ws, row)

    # AR Balance
    row += 3
    _apply_header(ws, row, ["Analyse AR"])
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    ar_items = [
        ("Solde Précédent", session.geac_ar_previous),
        ("Frais", session.geac_ar_charges),
        ("Paiements", session.geac_ar_payments),
        ("Nouveau Solde", session.geac_ar_new_balance),
        ("Variance", session.geac_ar_variance),
    ]

    for label, value in ar_items:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=_safe_float(value))
        c.number_format = CURRENCY_FORMAT
        row += 1


def _build_sd(wb, session):
    """Sheet 6: SD/Dépôt — SD entries + Depot envelopes"""
    ws = wb.create_sheet("SD-Dépôt", 5)

    # SD section
    _apply_header(ws, 1, ["SD — Sommaire Journalier des Dépôts"])
    ws.merge_cells('A1:F1')

    row = 2
    _apply_header(ws, row, ["Département", "Nom", "Devise", "Montant", "Vérifié", "Remboursement"])
    _set_column_width(ws, 1, 15)
    _set_column_width(ws, 2, 15)
    _set_column_width(ws, 3, 10)
    for col in range(4, 7):
        _set_column_width(ws, col, 15)
    row += 1

    sd_entries = session.get_json('sd_entries')
    sd_total = 0
    if isinstance(sd_entries, list):
        for entry in sd_entries:
            ws.cell(row=row, column=1, value=_safe_str(entry.get('department', '')))
            ws.cell(row=row, column=2, value=_safe_str(entry.get('name', '')))
            ws.cell(row=row, column=3, value=_safe_str(entry.get('currency', '')))

            verified = _safe_float(entry.get('verified', 0))
            reimb = _safe_float(entry.get('reimbursement', 0))

            c_amt = ws.cell(row=row, column=4, value=verified)
            c_reimb = ws.cell(row=row, column=5, value=reimb)
            c_amt.number_format = CURRENCY_FORMAT
            c_reimb.number_format = CURRENCY_FORMAT

            sd_total += verified
            row += 1

    # SD Total
    ws.cell(row=row, column=1, value="TOTAL SD")
    c_total = ws.cell(row=row, column=4, value=sd_total)
    c_total.number_format = CURRENCY_FORMAT
    c_total.fill = _total_fill()
    c_total.font = _total_font()
    _apply_total_row(ws, row)

    # Depot section
    row += 3
    _apply_header(ws, row, ["Enveloppes de Dépôt"])
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    depot = session.get_json('depot_data')
    for client_id in ['client6', 'client8']:
        client_data = depot.get(client_id, {}) if isinstance(depot, dict) else {}

        _apply_header(ws, row, [f"Client {client_id[-1]} — {client_data.get('date', '')}"])
        ws.merge_cells(f'A{row}:B{row}')
        row += 1

        amounts = client_data.get('amounts', [])
        depot_subtotal = 0
        if isinstance(amounts, list):
            for idx, amt in enumerate(amounts, 1):
                ws.cell(row=row, column=1, value=f"Enveloppe {idx}")
                c = ws.cell(row=row, column=2, value=_safe_float(amt))
                c.number_format = CURRENCY_FORMAT
                depot_subtotal += _safe_float(amt)
                row += 1

        ws.cell(row=row, column=1, value=f"Sous-total Client {client_id[-1]}")
        c_sub = ws.cell(row=row, column=2, value=depot_subtotal)
        c_sub.number_format = CURRENCY_FORMAT
        c_sub.fill = _total_fill()
        c_sub.font = _total_font()
        _apply_total_row(ws, row)
        row += 1


def _build_setd(wb, session):
    """Sheet 7: SetD — Personnel set-déjeuner entries"""
    ws = wb.create_sheet("SetD", 6)

    _apply_header(ws, 1, ["SetD — Personnel"])
    ws.merge_cells('A1:C1')

    row = 2
    _apply_header(ws, row, ["Nom", "Colonne", "Montant"])
    _set_column_width(ws, 1, 20)
    _set_column_width(ws, 2, 10)
    _set_column_width(ws, 3, 15)
    row += 1

    personnel = session.get_json('setd_personnel')
    setd_total = 0
    if isinstance(personnel, list):
        for entry in personnel:
            ws.cell(row=row, column=1, value=_safe_str(entry.get('name', '')))
            ws.cell(row=row, column=2, value=_safe_str(entry.get('column_letter', '')))

            amt = _safe_float(entry.get('amount', 0))
            c = ws.cell(row=row, column=3, value=amt)
            c.number_format = CURRENCY_FORMAT

            setd_total += amt
            row += 1

    # Total
    ws.cell(row=row, column=1, value="TOTAL")
    c_total = ws.cell(row=row, column=3, value=setd_total)
    c_total.number_format = CURRENCY_FORMAT
    c_total.fill = _total_fill()
    c_total.font = _total_font()
    _apply_total_row(ws, row)

    # RJ Balance
    row += 2
    ws.cell(row=row, column=1, value="Solde RJ")
    c_balance = ws.cell(row=row, column=3, value=_safe_float(session.setd_rj_balance))
    c_balance.number_format = CURRENCY_FORMAT


def _build_hp_admin(wb, session):
    """Sheet 8: HP/Admin — HP invoice entries with all F&B columns"""
    ws = wb.create_sheet("HP-Admin", 7)

    _apply_header(ws, 1, ["HP/Admin — Factures"])
    ws.merge_cells('A1:J1')

    row = 2
    headers = ["Secteur", "Nourriture", "Boisson", "Bière", "Vin", "Minéraux",
               "Autre", "Pourboire", "Raison", "Autorisé par"]
    _apply_header(ws, row, headers)
    _set_column_width(ws, 1, 12)
    for col in range(2, 11):
        _set_column_width(ws, col, 12)
    row += 1

    hp_entries = session.get_json('hp_admin_entries')
    hp_total = 0
    if isinstance(hp_entries, list):
        for entry in hp_entries:
            ws.cell(row=row, column=1, value=_safe_str(entry.get('area', '')))

            col = 2
            for field in ['nourriture', 'boisson', 'biere', 'vin', 'mineraux', 'autre', 'pourboire']:
                val = _safe_float(entry.get(field, 0))
                c = ws.cell(row=row, column=col, value=val)
                c.number_format = CURRENCY_FORMAT
                hp_total += val
                col += 1

            ws.cell(row=row, column=col, value=_safe_str(entry.get('raison', '')))
            col += 1
            ws.cell(row=row, column=col, value=_safe_str(entry.get('autorise_par', '')))
            row += 1

    # Total
    ws.cell(row=row, column=1, value="TOTAL")
    c_total = ws.cell(row=row, column=2, value=hp_total)
    c_total.number_format = CURRENCY_FORMAT
    c_total.fill = _total_fill()
    c_total.font = _total_font()
    _apply_total_row(ws, row)


def _build_internet(wb, session):
    """Sheet 9: Internet — CD 36.1 vs CD 36.5 variance"""
    ws = wb.create_sheet("Internet", 8)

    _apply_header(ws, 1, ["Internet — Détail"])
    ws.merge_cells('A1:B1')

    _set_column_width(ws, 1, 20)
    _set_column_width(ws, 2, 15)

    row = 2
    items = [
        ("CD 36.1", session.internet_ls_361),
        ("CD 36.5", session.internet_ls_365),
        ("Variance", session.internet_variance),
    ]

    for label, value in items:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=_safe_float(value))
        c.number_format = CURRENCY_FORMAT
        row += 1


def _build_sonifi(wb, session):
    """Sheet 10: Sonifi — CD 35.2 vs email PDF variance"""
    ws = wb.create_sheet("Sonifi", 9)

    _apply_header(ws, 1, ["Sonifi — Détail"])
    ws.merge_cells('A1:B1')

    _set_column_width(ws, 1, 20)
    _set_column_width(ws, 2, 15)

    row = 2
    items = [
        ("CD 35.2 (Courriel 03h00)", session.sonifi_cd_352),
        ("Montant Courriel Sonifi", session.sonifi_email),
        ("Variance", session.sonifi_variance),
    ]

    for label, value in items:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=_safe_float(value))
        c.number_format = CURRENCY_FORMAT
        row += 1


def _build_jour(wb, session):
    """Sheet 11: Jour — Full F&B breakdown, hébergement, taxes, occupation, KPIs"""
    ws = wb.create_sheet("Jour", 10)

    _set_column_width(ws, 1, 25)
    _set_column_width(ws, 2, 15)

    row = 1

    # F&B Revenue Section
    _apply_header(ws, row, ["F&B — RESTAURATION"])
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    _apply_header(ws, row, ["Département", "Nourriture", "Boisson", "Bière", "Minéraux", "Vin"])
    for col in range(3, 8):
        _set_column_width(ws, col, 12)
    row += 1

    fb_depts = [
        ("Café Link", 'cafe'),
        ("Piazza/Cupola", 'piazza'),
        ("Marché Spesa", 'spesa'),
        ("Service Chambres", 'chambres_svc'),
        ("Banquet", 'banquet'),
    ]

    fb_cats = ['nourriture', 'boisson', 'bieres', 'mineraux', 'vins']
    dept_totals = {}

    for dept_label, dept_key in fb_depts:
        ws.cell(row=row, column=1, value=dept_label)
        dept_total = 0

        for col_idx, cat in enumerate(fb_cats, 2):
            val = _safe_float(getattr(session, f'jour_{dept_key}_{cat}', 0))
            c = ws.cell(row=row, column=col_idx, value=val)
            c.number_format = CURRENCY_FORMAT
            dept_total += val

        dept_totals[dept_key] = dept_total
        row += 1

    # F&B Extras
    row += 1
    ws.cell(row=row, column=1, value="Pourboires")
    c = ws.cell(row=row, column=2, value=_safe_float(session.jour_pourboires))
    c.number_format = CURRENCY_FORMAT
    row += 1

    ws.cell(row=row, column=1, value="Tabagie")
    c = ws.cell(row=row, column=2, value=_safe_float(session.jour_tabagie))
    c.number_format = CURRENCY_FORMAT
    row += 1

    ws.cell(row=row, column=1, value="Location Salle")
    c = ws.cell(row=row, column=2, value=_safe_float(session.jour_location_salle))
    c.number_format = CURRENCY_FORMAT
    row += 1

    # Hébergement Section
    row += 1
    _apply_header(ws, row, ["HÉBERGEMENT"])
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    hebergement_items = [
        ("Revenue Chambres", session.jour_room_revenue),
        ("Téléphone Local", session.jour_tel_local),
        ("Téléphone Interurbain", session.jour_tel_interurbain),
        ("Téléphone Publics", session.jour_tel_publics),
    ]

    for label, value in hebergement_items:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=_safe_float(value))
        c.number_format = CURRENCY_FORMAT
        row += 1

    # Autres Revenus Section
    row += 1
    _apply_header(ws, row, ["AUTRES REVENUS"])
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    autres_items = [
        ("Nettoyeur", session.jour_nettoyeur),
        ("Machine Distributrice", session.jour_machine_distrib),
        ("Autres GL", session.jour_autres_gl),
        ("Sonifi", session.jour_sonifi),
        ("Lit Pliant", session.jour_lit_pliant),
        ("Boutique", session.jour_boutique),
        ("Internet", session.jour_internet),
        ("Massage", session.jour_massage),
        ("Forfait SJ", session.jour_forfait_sj),
        ("Différence Forfait", session.jour_diff_forfait),
        ("G4 Montant", session.g4_montant),
    ]

    for label, value in autres_items:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=_safe_float(value))
        c.number_format = CURRENCY_FORMAT
        row += 1

    # Taxes Section
    row += 1
    _apply_header(ws, row, ["TAXES"])
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    tax_items = [
        ("TVQ", session.jour_tvq),
        ("TPS", session.jour_tps),
        ("Taxe Hébergement", session.jour_taxe_hebergement),
    ]

    for label, value in tax_items:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=_safe_float(value))
        c.number_format = CURRENCY_FORMAT
        row += 1

    # Occupation Section
    row += 1
    _apply_header(ws, row, ["OCCUPATION"])
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    occ_items = [
        ("Chambres Simples", session.jour_rooms_simple),
        ("Chambres Doubles", session.jour_rooms_double),
        ("Suites", session.jour_rooms_suite),
        ("Chambres Comp", session.jour_rooms_comp),
        ("Nombre Clients", session.jour_nb_clients),
        ("Chambres Hors Usage", session.jour_rooms_hors_usage),
    ]

    for label, value in occ_items:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=_safe_int(value))
        c.number_format = INTEGER_FORMAT
        row += 1

    # KPIs Section
    row += 1
    _apply_header(ws, row, ["INDICATEURS"])
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    kpi_items = [
        ("Total F&B", session.jour_total_fb),
        ("Total Revenue", session.jour_total_revenue),
        ("ADR", session.jour_adr),
        ("RevPAR", session.jour_revpar),
        ("Taux Occupation", session.jour_occupancy_rate),
    ]

    for label, value in kpi_items:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=_safe_float(value))
        if 'Taux' in label:
            c.number_format = PERCENTAGE_FORMAT
        else:
            c.number_format = CURRENCY_FORMAT
        row += 1


def _build_quasimodo(wb, session):
    """Sheet 12: Quasimodo — Card reconciliation + cash + variance"""
    ws = wb.create_sheet("Quasimodo", 11)

    _apply_header(ws, 1, ["Quasimodo — Réconciliation Globale"])
    ws.merge_cells('A1:C1')

    _set_column_width(ws, 1, 20)
    _set_column_width(ws, 2, 15)
    _set_column_width(ws, 3, 15)

    row = 2
    _apply_header(ws, row, ["Carte", "F&B", "Réception"])
    row += 1

    card_types = ['debit', 'visa', 'mc', 'amex', 'discover']
    total_fb = 0
    total_rec = 0

    for card_type in card_types:
        ws.cell(row=row, column=1, value=card_type.upper())

        fb_val = _safe_float(getattr(session, f'quasi_fb_{card_type}', 0))
        rec_val = _safe_float(getattr(session, f'quasi_rec_{card_type}', 0))

        c_fb = ws.cell(row=row, column=2, value=fb_val)
        c_rec = ws.cell(row=row, column=3, value=rec_val)

        c_fb.number_format = CURRENCY_FORMAT
        c_rec.number_format = CURRENCY_FORMAT

        total_fb += fb_val
        total_rec += rec_val
        row += 1

    # Totals
    ws.cell(row=row, column=1, value="TOTAL CARTES").font = _total_font()
    c_tot_fb = ws.cell(row=row, column=2, value=total_fb)
    c_tot_rec = ws.cell(row=row, column=3, value=total_rec)
    c_tot_fb.number_format = CURRENCY_FORMAT
    c_tot_rec.number_format = CURRENCY_FORMAT
    c_tot_fb.fill = _total_fill()
    c_tot_rec.fill = _total_fill()
    c_tot_fb.font = _total_font()
    c_tot_rec.font = _total_font()
    _apply_total_row(ws, row)

    # Cash and Summary
    row += 2
    ws.cell(row=row, column=1, value="Caisse CDN")
    c = ws.cell(row=row, column=2, value=_safe_float(session.quasi_cash_cdn))
    c.number_format = CURRENCY_FORMAT
    row += 1

    ws.cell(row=row, column=1, value="Caisse USD")
    c = ws.cell(row=row, column=2, value=_safe_float(session.quasi_cash_usd))
    c.number_format = CURRENCY_FORMAT
    row += 1

    ws.cell(row=row, column=1, value="Facteur AMEX")
    c = ws.cell(row=row, column=2, value=_safe_float(session.quasi_amex_factor))
    c.number_format = '0.0000'
    row += 2

    ws.cell(row=row, column=1, value="Total Quasimodo").font = _total_font()
    c = ws.cell(row=row, column=2, value=_safe_float(session.quasi_total))
    c.number_format = CURRENCY_FORMAT
    c.fill = _total_fill()
    c.font = _total_font()
    row += 1

    ws.cell(row=row, column=1, value="Total RJ").font = _total_font()
    c = ws.cell(row=row, column=2, value=_safe_float(session.quasi_rj_total))
    c.number_format = CURRENCY_FORMAT
    c.fill = _total_fill()
    c.font = _total_font()
    row += 1

    ws.cell(row=row, column=1, value="VARIANCE").font = _total_font()
    c = ws.cell(row=row, column=2, value=_safe_float(session.quasi_variance))
    c.number_format = CURRENCY_FORMAT
    c.fill = _total_fill()
    c.font = _total_font()


def _build_dbrs(wb, session):
    """Sheet 13: DBRS — Daily Business Review Summary"""
    ws = wb.create_sheet("DBRS", 12)

    _apply_header(ws, 1, ["DBRS — Daily Business Review Summary"])
    ws.merge_cells('A1:B1')

    _set_column_width(ws, 1, 25)
    _set_column_width(ws, 2, 15)

    row = 2

    # Market Segments
    _apply_header(ws, row, ["Segments de Marché"])
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    market_segs = session.get_json('dbrs_market_segments')
    if isinstance(market_segs, dict):
        for seg_name, seg_val in market_segs.items():
            ws.cell(row=row, column=1, value=seg_name.replace('_', ' ').title())
            c = ws.cell(row=row, column=2, value=_safe_float(seg_val))
            c.number_format = CURRENCY_FORMAT
            row += 1

    row += 1

    # KPIs
    _apply_header(ws, row, ["Indicateurs"])
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    kpi_items = [
        ("Revenue du Jour", session.dbrs_daily_rev_today),
        ("ADR (Average Daily Rate)", session.dbrs_adr),
        ("Nombre de Chambres", session.dbrs_house_count),
        ("Noshows", session.dbrs_noshow_count),
        ("Revenue Noshows", session.dbrs_noshow_revenue),
    ]

    for label, value in kpi_items:
        ws.cell(row=row, column=1, value=label)
        if 'Chambres' in label or 'No' in label:
            c = ws.cell(row=row, column=2, value=_safe_int(value))
            c.number_format = INTEGER_FORMAT
        else:
            c = ws.cell(row=row, column=2, value=_safe_float(value))
            c.number_format = CURRENCY_FORMAT
        row += 1


def _build_sommaire(wb, session):
    """Sheet 14: Sommaire — Validation status + check results"""
    ws = wb.create_sheet("Sommaire", 13)

    _apply_header(ws, 1, ["SOMMAIRE — Validation"])
    ws.merge_cells('A1:D1')

    _set_column_width(ws, 1, 30)
    _set_column_width(ws, 2, 15)
    _set_column_width(ws, 3, 25)
    _set_column_width(ws, 4, 15)

    row = 2
    _apply_header(ws, row, ["Vérification", "Statut", "Détail", "Variance"])
    row += 1

    checks = [
        ("Recap Balanced", session.is_recap_balanced, f"Balance: {_safe_float(session.recap_balance):.2f}", session.recap_balance),
        ("Transelect Balanced", session.is_transelect_balanced, f"Variance: {_safe_float(session.transelect_variance):.2f}", session.transelect_variance),
        ("AR Balanced", session.is_ar_balanced, f"Variance: {_safe_float(session.geac_ar_variance):.2f}", session.geac_ar_variance),
        ("Fully Balanced", session.is_fully_balanced, "Tous les soldes OK", None),
    ]

    for check_name, check_status, detail, variance in checks:
        ws.cell(row=row, column=1, value=check_name)

        status_val = "✓ OK" if check_status else "✗ ERREUR"
        status_cell = ws.cell(row=row, column=2, value=status_val)
        if check_status:
            status_cell.fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type='solid')
            status_cell.font = Font(bold=True, color=WHITE)
        else:
            status_cell.fill = PatternFill(start_color=RED, end_color=RED, fill_type='solid')
            status_cell.font = Font(bold=True, color=WHITE)

        ws.cell(row=row, column=3, value=detail)

        if variance is not None:
            c = ws.cell(row=row, column=4, value=_safe_float(variance))
            c.number_format = CURRENCY_FORMAT

        row += 1

    # Session Info
    row += 2
    _apply_header(ws, row, ["Informations de Session"])
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    info_items = [
        ("Date d'Audit", session.audit_date.isoformat() if session.audit_date else ''),
        ("Auditeur", _safe_str(session.auditor_name)),
        ("Statut", _safe_str(session.status)),
        ("Créé à", session.started_at.isoformat() if session.started_at else ''),
        ("Complété à", session.completed_at.isoformat() if session.completed_at else ''),
    ]

    for label, value in info_items:
        ws.cell(row=row, column=1, value=label).font = _subheader_font()
        ws.cell(row=row, column=2, value=value)
        row += 1


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# MAIN EXPORT FUNCTIONS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _create_excel_workbook(session):
    """Create complete Excel workbook from NightAuditSession."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Build all 14 sheets
    _build_controle(wb, session)
    _build_dueback(wb, session)
    _build_recap(wb, session)
    _build_transelect(wb, session)
    _build_geac(wb, session)
    _build_sd(wb, session)
    _build_setd(wb, session)
    _build_hp_admin(wb, session)
    _build_internet(wb, session)
    _build_sonifi(wb, session)
    _build_jour(wb, session)
    _build_quasimodo(wb, session)
    _build_dbrs(wb, session)
    _build_sommaire(wb, session)

    return wb


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# FLASK ROUTES
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@rj_excel_bp.route('/excel/<date_str>', methods=['GET'])
def export_excel(date_str):
    """
    GET /api/rj/export/excel/<date>
    Generate and download .xlsx for a single session.

    Args:
        date_str: Date in format YYYY-MM-DD

    Returns:
        XLSX file download
    """
    try:
        audit_date = parse_date(date_str).date()
    except (ValueError, TypeError, AttributeError):
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

    session = NightAuditSession.query.filter_by(audit_date=audit_date).first()
    if not session:
        return jsonify({'error': f'No session found for {audit_date}'}), 404

    wb = _create_excel_workbook(session)

    # Save to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"RJ_Natif_{audit_date.isoformat()}.xlsx"

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@rj_excel_bp.route('/excel/batch', methods=['GET'])
def export_batch():
    """
    GET /api/rj/export/excel/batch?start=<date>&end=<date>
    Batch export date range to ZIP file containing one XLSX per session.

    Args:
        start: Start date (YYYY-MM-DD)
        end: End date (YYYY-MM-DD)

    Returns:
        ZIP file containing multiple XLSX files
    """
    try:
        start_date = parse_date(request.args.get('start', '')).date()
        end_date = parse_date(request.args.get('end', '')).date()
    except (ValueError, TypeError, AttributeError):
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

    if start_date > end_date:
        return jsonify({'error': 'start_date must be before end_date'}), 400

    sessions = NightAuditSession.query.filter(
        NightAuditSession.audit_date >= start_date,
        NightAuditSession.audit_date <= end_date
    ).order_by(NightAuditSession.audit_date).all()

    if not sessions:
        return jsonify({'error': f'No sessions found between {start_date} and {end_date}'}), 404

    # Create ZIP buffer
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for session in sessions:
            wb = _create_excel_workbook(session)

            # Save workbook to bytes
            xlsx_buffer = io.BytesIO()
            wb.save(xlsx_buffer)
            xlsx_buffer.seek(0)

            filename = f"RJ_Natif_{session.audit_date.isoformat()}.xlsx"
            zf.writestr(filename, xlsx_buffer.getvalue())

    zip_buffer.seek(0)

    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f"RJ_Export_{start_date}_{end_date}.zip"
    )
