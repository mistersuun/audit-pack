"""
RJ Native — Web-based night audit form (no Excel required).

Replaces the Excel-based RJ workflow with a native database-backed form.
Auditors enter data directly in the browser; calculations happen automatically;
everything is saved to the database. Export to Excel/PDF on demand.
"""

from flask import Blueprint, request, jsonify, render_template, session, send_file
from functools import wraps
from datetime import datetime, date, timedelta
from database.models import db, NightAuditSession, DailyReconciliation, DueBack, DailyJourMetrics, RJArchive, RJSheetData
import json
import logging
import io
import xlrd
from xlutils.copy import copy as xlutils_copy
import xlwt
import os

# Import Excel utilities for import/export
from utils.rj_reader import RJReader
from utils.rj_filler import RJFiller
from utils.jour_mapping import (
    JOUR_NAS_TO_COL, JOUR_COL_TO_NAS, JOUR_MACRO_COLS,
    nas_jour_to_excel_dict, excel_jour_to_nas_dict,
)
from utils.ole_builder import rebuild_xls_with_vba

logger = logging.getLogger(__name__)

rj_native_bp = Blueprint('rj_native', __name__)


def auth_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('authenticated'):
            from flask import redirect, url_for
            return redirect(url_for('auth_v2.login'))
        return f(*args, **kwargs)
    return decorated


# ═══════════════════════════════════════
# PAGES
# ═══════════════════════════════════════

@rj_native_bp.route('/rj/native')
@auth_required
def native_page():
    """Main native RJ form page."""
    today = date.today().isoformat()
    return render_template('audit/rj/rj_native.html', today=today)


@rj_native_bp.route('/rj/native/<audit_date>')
@auth_required
def native_page_date(audit_date):
    """Load native form for a specific date."""
    return render_template('audit/rj/rj_native.html', today=audit_date)


# ═══════════════════════════════════════
# HELPERS — CARRY-FORWARD & CLEAR MACROS
# ═══════════════════════════════════════

# Fields to NEVER copy (session identity — everything else is copied verbatim)
_SKIP_FIELDS = {
    'id', 'audit_date', 'auditor_name', 'status', 'started_at', 'completed_at',
    'created_at', 'notes',
}

# Fields cleared by the "Effacer Recap" macro button
_RECAP_FIELDS = [
    'cash_ls_lecture', 'cash_ls_corr', 'cash_pos_lecture', 'cash_pos_corr',
    'cheque_ar_lecture', 'cheque_ar_corr', 'cheque_dr_lecture', 'cheque_dr_corr',
    'remb_gratuite_lecture', 'remb_gratuite_corr', 'remb_client_lecture', 'remb_client_corr',
    'dueback_reception_lecture', 'dueback_reception_corr', 'dueback_nb_lecture', 'dueback_nb_corr',
    'deposit_cdn', 'deposit_us', 'recap_balance',
]

# Fields cleared by the "Effacer Transelect" macro button
_TRANSELECT_FIELDS = [
    'transelect_restaurant', 'transelect_reception', 'transelect_quasimodo', 'transelect_variance',
]

# Fields cleared by the "Effacer GEAC" macro button
_GEAC_FIELDS = [
    'geac_cashout', 'geac_daily_rev',
    'geac_ar_previous', 'geac_ar_charges', 'geac_ar_payments',
    'geac_ar_new_balance', 'geac_ar_variance',
]


# Réceptionnistes qui ne gardent JAMAIS de solde "previous" reporté.
# Leur "nouveau" d'hier ne devient PAS le "previous" d'aujourd'hui.
_DUEBACK_NO_CARRY = {'emery', 'nikoleta', 'thaneekan', 'manolo'}


def _dueback_skip_previous(name):
    """Return True if this receptionist should never have a previous balance."""
    name_lower = (name or '').lower()
    return any(skip in name_lower for skip in _DUEBACK_NO_CARRY)


def _carry_forward(prev, nas):
    """Copy ALL data from previous session — like copying yesterday's RJ Excel file.

    Mimics the exact Excel workflow:
    1. Copy yesterday's RJ file (EVERYTHING is preserved)
    2. Change the date and auditor name
    3. Auditor clicks "Effacer Recap/Transelect/GEAC" macro buttons when ready
    4. Auditor fills in today's new data

    ALL fields carry forward including Recap, Transelect, GEAC (yesterday's values
    stay visible until the auditor explicitly clears them with macro buttons).
    DueBack gets special treatment: names rotate (yesterday's 'nouveau' → today's 'previous').
    """
    # 1. Copy ALL columns except session identity fields
    for col in NightAuditSession.__table__.columns:
        name = col.name
        if name in _SKIP_FIELDS:
            continue
        val = getattr(prev, name, None)
        if val is not None:
            setattr(nas, name, val)

    # 2. DueBack special rotation: yesterday's 'nouveau' becomes today's 'previous'
    #    Exception: certain receptionists NEVER carry a previous balance
    prev_entries = prev.get_json('dueback_entries')
    if isinstance(prev_entries, list) and prev_entries:
        carried = []
        for e in prev_entries:
            ename = e.get('name', '')
            if not ename:
                continue
            nouveau = e.get('nouveau', 0) or 0
            # Previous is always NEGATIVE (debit carried forward)
            prev_val = 0 if _dueback_skip_previous(ename) else -abs(float(nouveau))
            carried.append({'name': ename, 'previous': prev_val, 'nouveau': 0})
        nas.set_json('dueback_entries', carried)


# ═══════════════════════════════════════
# API — SESSION MANAGEMENT
# ═══════════════════════════════════════

@rj_native_bp.route('/api/rj/native/new', methods=['POST'])
@auth_required
def create_session():
    """Create a new night audit session for a date."""
    data = request.get_json(force=True)
    audit_date_str = data.get('date')
    auditor = data.get('auditor', session.get('user_name', ''))

    if not audit_date_str:
        return jsonify({'error': 'Date requise'}), 400

    try:
        audit_date = datetime.strptime(audit_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide (YYYY-MM-DD)'}), 400

    # Check if session already exists
    existing = NightAuditSession.query.filter_by(audit_date=audit_date).first()
    if existing:
        return jsonify({
            'success': True,
            'session': existing.to_dict(),
            'message': 'Session existante chargée',
            'is_new': False
        })

    # Load previous day data for carry-forward (like copying yesterday's RJ in Excel)
    prev_session = NightAuditSession.query.filter(
        NightAuditSession.audit_date < audit_date
    ).order_by(NightAuditSession.audit_date.desc()).first()

    nas = NightAuditSession(
        audit_date=audit_date,
        auditor_name=auditor,
        status='draft'
    )

    # Carry forward ALL cumulative data from previous session
    if prev_session:
        _carry_forward(prev_session, nas)

    db.session.add(nas)
    db.session.commit()

    return jsonify({
        'success': True,
        'session': nas.to_dict(),
        'message': 'Nouvelle session créée (données de la veille copiées)' if prev_session else 'Nouvelle session créée',
        'is_new': True,
        'carried_from': prev_session.audit_date.isoformat() if prev_session else None
    })


@rj_native_bp.route('/api/rj/native/load/<audit_date>')
@auth_required
def load_session(audit_date):
    """Load all data for a specific date."""
    try:
        d = datetime.strptime(audit_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400

    nas = NightAuditSession.query.filter_by(audit_date=d).first()
    if not nas:
        return jsonify({'exists': False})

    return jsonify({
        'exists': True,
        'session': nas.to_dict()
    })



@rj_native_bp.route('/api/rj/native/session/<audit_date>')
@auth_required
def session_endpoint(audit_date):
    """Get or create a session for a specific date."""
    try:
        d = datetime.strptime(audit_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400

    # Try to load existing session
    nas = NightAuditSession.query.filter_by(audit_date=d).first()
    
    # If not found, create a new one
    if not nas:
        auditor = session.get('user_name', '')
        # Load previous day data for carry-forward
        prev_session = NightAuditSession.query.filter(
            NightAuditSession.audit_date < d
        ).order_by(NightAuditSession.audit_date.desc()).first()

        nas = NightAuditSession(
            audit_date=d,
            auditor_name=auditor,
            status='draft'
        )

        # Carry forward ALL cumulative data from previous session
        if prev_session:
            _carry_forward(prev_session, nas)

        db.session.add(nas)
        db.session.commit()

    return jsonify(nas.to_dict())


@rj_native_bp.route('/api/rj/native/list')
@auth_required
def list_sessions():
    """List recent sessions."""
    sessions = NightAuditSession.query.order_by(
        NightAuditSession.audit_date.desc()
    ).limit(30).all()
    return jsonify({
        'sessions': [{
            'date': s.audit_date.isoformat(),
            'auditor': s.auditor_name,
            'status': s.status,
            'is_balanced': s.is_fully_balanced
        } for s in sessions]
    })


# ═══════════════════════════════════════
# API — CLEAR MACROS (emulates Excel VBA)
# ═══════════════════════════════════════

@rj_native_bp.route('/api/rj/native/clear/recap', methods=['POST'])
@auth_required
def clear_recap():
    """Efface Recap — like efface_recap() VBA macro."""
    data = request.get_json(force=True)
    audit_date_str = data.get('date')
    if not audit_date_str:
        return jsonify({'error': 'Date requise'}), 400
    try:
        d = datetime.strptime(audit_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400
    nas = NightAuditSession.query.filter_by(audit_date=d).first()
    if not nas:
        return jsonify({'error': 'Session introuvable'}), 404

    for field in _RECAP_FIELDS:
        setattr(nas, field, 0)
    nas.is_recap_balanced = False
    db.session.commit()
    return jsonify({'success': True, 'message': 'Recap effacé', 'session': nas.to_dict()})


@rj_native_bp.route('/api/rj/native/clear/transelect', methods=['POST'])
@auth_required
def clear_transelect():
    """Efface Transelect — like eff_trans() VBA macro."""
    data = request.get_json(force=True)
    audit_date_str = data.get('date')
    if not audit_date_str:
        return jsonify({'error': 'Date requise'}), 400
    try:
        d = datetime.strptime(audit_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400
    nas = NightAuditSession.query.filter_by(audit_date=d).first()
    if not nas:
        return jsonify({'error': 'Session introuvable'}), 404

    nas.transelect_restaurant = '{}'
    nas.transelect_reception = '{}'
    nas.transelect_quasimodo = '{}'
    nas.transelect_variance = 0
    nas.is_transelect_balanced = False
    db.session.commit()
    return jsonify({'success': True, 'message': 'Transelect effacé', 'session': nas.to_dict()})


@rj_native_bp.route('/api/rj/native/clear/geac', methods=['POST'])
@auth_required
def clear_geac():
    """Efface GEAC — like efface_rapport_geac() VBA macro."""
    data = request.get_json(force=True)
    audit_date_str = data.get('date')
    if not audit_date_str:
        return jsonify({'error': 'Date requise'}), 400
    try:
        d = datetime.strptime(audit_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400
    nas = NightAuditSession.query.filter_by(audit_date=d).first()
    if not nas:
        return jsonify({'error': 'Session introuvable'}), 404

    nas.geac_cashout = '{}'
    nas.geac_daily_rev = '{}'
    nas.geac_ar_previous = 0
    nas.geac_ar_charges = 0
    nas.geac_ar_payments = 0
    nas.geac_ar_new_balance = 0
    nas.geac_ar_variance = 0
    nas.is_ar_balanced = False
    db.session.commit()
    return jsonify({'success': True, 'message': 'GEAC effacé', 'session': nas.to_dict()})


@rj_native_bp.route('/api/rj/native/clear/all-daily', methods=['POST'])
@auth_required
def clear_all_daily():
    """Efface Recap + Transelect + GEAC en un clic (Phase 1 back audit)."""
    data = request.get_json(force=True)
    audit_date_str = data.get('date')
    if not audit_date_str:
        return jsonify({'error': 'Date requise'}), 400
    try:
        d = datetime.strptime(audit_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400
    nas = NightAuditSession.query.filter_by(audit_date=d).first()
    if not nas:
        return jsonify({'error': 'Session introuvable'}), 404

    # Clear Recap
    for field in _RECAP_FIELDS:
        setattr(nas, field, 0)
    nas.is_recap_balanced = False
    # Clear Transelect
    nas.transelect_restaurant = '{}'
    nas.transelect_reception = '{}'
    nas.transelect_quasimodo = '{}'
    nas.transelect_variance = 0
    nas.is_transelect_balanced = False
    # Clear GEAC
    nas.geac_cashout = '{}'
    nas.geac_daily_rev = '{}'
    nas.geac_ar_previous = 0
    nas.geac_ar_charges = 0
    nas.geac_ar_payments = 0
    nas.geac_ar_new_balance = 0
    nas.geac_ar_variance = 0
    nas.is_ar_balanced = False

    db.session.commit()
    return jsonify({'success': True, 'message': 'Recap + Transelect + GEAC effacés', 'session': nas.to_dict()})


# ═══════════════════════════════════════════════════════════════════════════════
# IMPORT / EXPORT — Upload yesterday's RJ → Web edit → Download updated RJ
# ═══════════════════════════════════════════════════════════════════════════════

# Storage for uploaded RJ files (keyed by audit_date ISO string)
# Used as base for export: we inject the new day's data into the uploaded file
# Files are also persisted to disk so they survive server restarts.
_UPLOADED_RJ = {}  # 'YYYY-MM-DD' → BytesIO

_RJ_CACHE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                             'database', 'rj_cache')
os.makedirs(_RJ_CACHE_DIR, exist_ok=True)


def _persist_rj(date_iso, file_bytes_io):
    """Save uploaded RJ to memory AND disk."""
    file_bytes_io.seek(0)
    _UPLOADED_RJ[date_iso] = io.BytesIO(file_bytes_io.getvalue())
    cache_path = os.path.join(_RJ_CACHE_DIR, f'{date_iso}.xls')
    with open(cache_path, 'wb') as f:
        f.write(file_bytes_io.getvalue())


def _load_rj(date_iso):
    """Load RJ from memory or disk cache. Returns BytesIO or None."""
    if date_iso in _UPLOADED_RJ:
        _UPLOADED_RJ[date_iso].seek(0)
        return _UPLOADED_RJ[date_iso]
    cache_path = os.path.join(_RJ_CACHE_DIR, f'{date_iso}.xls')
    if os.path.exists(cache_path):
        with open(cache_path, 'rb') as f:
            _UPLOADED_RJ[date_iso] = io.BytesIO(f.read())
        return _UPLOADED_RJ[date_iso]
    return None


def _archive_rj_to_db(file_bytes, audit_date, source_filename=None, uploaded_by=None):
    """Archive ALL sheets from an RJ Excel file into the database.
    Stores both raw binary and parsed cell data for every sheet.
    This ensures zero data loss even if the original Excel file is deleted."""
    try:
        # Check if already archived for this date
        existing = RJArchive.query.filter_by(audit_date=audit_date).first()
        if existing:
            # Update existing archive
            existing.file_binary = file_bytes
            existing.uploaded_at = datetime.utcnow()
            if source_filename:
                existing.source_filename = source_filename
            if uploaded_by:
                existing.uploaded_by = uploaded_by
            # Delete old sheet data
            RJSheetData.query.filter_by(archive_id=existing.id).delete()
            archive = existing
        else:
            archive = RJArchive(
                audit_date=audit_date,
                source_filename=source_filename,
                uploaded_by=uploaded_by,
                file_binary=file_bytes,
            )
            db.session.add(archive)
            db.session.flush()

        # Parse all sheets using xlrd (for .xls) or openpyxl (for .xlsx)
        sheet_names_list = []
        total_rows = 0

        try:
            # Try xlrd first (.xls format)
            wb = xlrd.open_workbook(file_contents=file_bytes, formatting_info=False)
            sheet_names_list = wb.sheet_names()

            for idx, sname in enumerate(sheet_names_list):
                ws = wb.sheet_by_index(idx)
                rows_data = []
                for r in range(min(ws.nrows, 500)):  # Cap at 500 rows per sheet
                    row = []
                    for c in range(ws.ncols):
                        cell = ws.cell(r, c)
                        if cell.ctype == xlrd.XL_CELL_EMPTY:
                            row.append(None)
                        elif cell.ctype == xlrd.XL_CELL_NUMBER:
                            v = cell.value
                            row.append(int(v) if v == int(v) else round(v, 4))
                        elif cell.ctype == xlrd.XL_CELL_DATE:
                            try:
                                dt = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                                row.append(f'{dt[0]:04d}-{dt[1]:02d}-{dt[2]:02d}')
                            except:
                                row.append(str(cell.value))
                        elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                            row.append(bool(cell.value))
                        else:
                            row.append(str(cell.value) if cell.value else None)
                    rows_data.append(row)

                headers = rows_data[0] if rows_data else []
                total_rows += len(rows_data)

                sheet_data = RJSheetData(
                    archive_id=archive.id,
                    audit_date=audit_date,
                    sheet_name=sname,
                    sheet_index=idx,
                    row_count=len(rows_data),
                    col_count=ws.ncols,
                    data_json=json.dumps(rows_data, ensure_ascii=False, default=str),
                    headers_json=json.dumps(headers, ensure_ascii=False, default=str),
                )
                db.session.add(sheet_data)

        except xlrd.XLRDError:
            # Try openpyxl (.xlsx format)
            try:
                from openpyxl import load_workbook
                wb_x = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
                sheet_names_list = wb_x.sheetnames

                for idx, sname in enumerate(sheet_names_list):
                    ws = wb_x[sname]
                    rows_data = []
                    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                        if r_idx >= 500:
                            break
                        row_list = []
                        for cell in row:
                            if cell is None:
                                row_list.append(None)
                            elif isinstance(cell, (int, float)):
                                row_list.append(round(cell, 4) if isinstance(cell, float) else cell)
                            elif isinstance(cell, datetime):
                                row_list.append(cell.strftime('%Y-%m-%d'))
                            else:
                                row_list.append(str(cell))
                        rows_data.append(row_list)

                    headers = rows_data[0] if rows_data else []
                    ncols = max(len(r) for r in rows_data) if rows_data else 0
                    total_rows += len(rows_data)

                    sheet_data = RJSheetData(
                        archive_id=archive.id,
                        audit_date=audit_date,
                        sheet_name=sname,
                        sheet_index=idx,
                        row_count=len(rows_data),
                        col_count=ncols,
                        data_json=json.dumps(rows_data, ensure_ascii=False, default=str),
                        headers_json=json.dumps(headers, ensure_ascii=False, default=str),
                    )
                    db.session.add(sheet_data)

                wb_x.close()
            except Exception as e:
                logger.warning(f"Could not parse xlsx for archive: {e}")

        archive.sheet_names = json.dumps(sheet_names_list, ensure_ascii=False)
        archive.total_sheets = len(sheet_names_list)
        archive.total_rows = total_rows
        db.session.commit()

        logger.info(f"RJ archived: {audit_date} — {len(sheet_names_list)} sheets, {total_rows} rows")
        return {'sheets': len(sheet_names_list), 'rows': total_rows}

    except Exception as e:
        logger.error(f"RJ archive error: {e}", exc_info=True)
        return {'error': str(e)}


@rj_native_bp.route('/api/rj/native/import/excel', methods=['POST'])
@auth_required
def import_excel():
    """Import yesterday's RJ Excel file and create a session for the NEXT day.

    The auditor uploads yesterday's completed RJ (e.g. Rj 07-02-2026.xls).
    We read all data from it, create a session for the NEXT day (day 8),
    and store the original file in memory so the export can inject new data into it.

    Returns:
        JSON with the new session data and import summary.
    """
    try:
        # 1. Get the uploaded file
        if 'rj_file' not in request.files:
            return jsonify({'error': 'Aucun fichier sélectionné'}), 400

        f = request.files['rj_file']
        if not f.filename:
            return jsonify({'error': 'Aucun fichier sélectionné'}), 400

        # Read into BytesIO
        file_bytes = io.BytesIO(f.read())
        file_bytes.seek(0)

        # 2. Parse the uploaded RJ with RJReader
        reader = RJReader(io.BytesIO(file_bytes.getvalue()))

        # 3. Extract the date from the controle sheet
        controle = reader.read_controle()
        vjour = int(controle.get('jour', 0))
        vmois = int(controle.get('mois', 0))
        vannee = int(controle.get('annee', 0))

        if not vjour or not vmois or not vannee:
            return jsonify({'error': 'Impossible de lire la date dans le fichier RJ (controle)'}), 400

        # The uploaded RJ is for day X; the auditor works on day X+1
        try:
            uploaded_date = date(vannee, vmois, vjour)
            new_date = uploaded_date + timedelta(days=1)
        except ValueError as e:
            return jsonify({'error': f'Date invalide dans le fichier: {e}'}), 400

        # 4. Create or load session for the new day
        nas = NightAuditSession.query.filter_by(audit_date=new_date).first()
        is_new = False

        if not nas:
            nas = NightAuditSession(
                audit_date=new_date,
                auditor_name=session.get('user_name', ''),
                status='draft'
            )
            db.session.add(nas)
            is_new = True

        # 5. Map Excel data → NAS fields
        import_summary = _import_excel_to_nas(reader, nas, vjour)

        # 6. Store the original file in memory for later export
        file_bytes.seek(0)
        file_bytes.seek(0)
        _persist_rj(new_date.isoformat(), file_bytes)

        # 7. ★ Archive source (yesterday) to DB
        file_bytes.seek(0)
        raw_bytes = file_bytes.getvalue()
        archive_result = _archive_rj_to_db(
            raw_bytes, uploaded_date,
            source_filename=f.filename,
            uploaded_by=session.get('user_name', '')
        )

        # 8. ★ Also archive as LIVE working copy for today's date
        #    This is the base file the auditor will fill tonight.
        #    Each export will re-archive the updated version.
        _archive_rj_to_db(
            raw_bytes, new_date,
            source_filename=f"Rj {new_date.strftime('%d-%m-%Y')}.xls",
            uploaded_by=session.get('user_name', '')
        )

        # 9. Recalculate and save
        nas.calculate_all()
        db.session.commit()

        return jsonify({
            'success': True,
            'session': nas.to_dict(),
            'is_new': is_new,
            'imported_from': uploaded_date.isoformat(),
            'new_date': new_date.isoformat(),
            'summary': import_summary,
            'archive': archive_result,
            'message': f'RJ du {uploaded_date.strftime("%d/%m/%Y")} importé et archivé. '
                       f'Session créée pour le {new_date.strftime("%d/%m/%Y")}.'
        })

    except Exception as e:
        logger.error(f"Import Excel error: {e}", exc_info=True)
        return jsonify({'error': f"Erreur lors de l'import: {str(e)}"}), 500


# ── ETL: RJ Archives ──────────────────────────────────────────
@rj_native_bp.route('/api/rj/archives')
@auth_required
def list_archives():
    """List all archived RJ workbooks."""
    from sqlalchemy import desc
    archives = RJArchive.query.order_by(desc(RJArchive.audit_date)).all()
    return jsonify({
        'total': len(archives),
        'archives': [a.to_dict() for a in archives],
    })


@rj_native_bp.route('/api/rj/archives/<int:archive_id>')
@auth_required
def get_archive(archive_id):
    """Get archive metadata + list of sheets."""
    archive = RJArchive.query.get_or_404(archive_id)
    sheets = RJSheetData.query.filter_by(archive_id=archive.id).order_by(RJSheetData.sheet_index).all()
    return jsonify({
        'archive': archive.to_dict(),
        'sheets': [s.to_dict() for s in sheets],
    })


@rj_native_bp.route('/api/rj/archives/<int:archive_id>/sheet/<sheet_name>')
@auth_required
def get_archive_sheet(archive_id, sheet_name):
    """Get raw data from a specific sheet in an archive."""
    sheet = RJSheetData.query.filter_by(archive_id=archive_id, sheet_name=sheet_name).first_or_404()
    return jsonify({
        'sheet': sheet.to_dict(),
        'data': json.loads(sheet.data_json) if sheet.data_json else [],
    })


@rj_native_bp.route('/api/rj/archives/<int:archive_id>/download')
@auth_required
def download_archive(archive_id):
    """Download the original Excel binary from the archive."""
    archive = RJArchive.query.get_or_404(archive_id)
    if not archive.file_binary:
        return jsonify({'error': 'Fichier binaire non disponible'}), 404

    ext = '.xlsx' if archive.source_filename and archive.source_filename.endswith('.xlsx') else '.xls'
    fname = archive.source_filename or f'RJ_{archive.audit_date.isoformat()}{ext}'

    return send_file(
        io.BytesIO(archive.file_binary),
        as_attachment=True,
        download_name=fname,
        mimetype='application/vnd.ms-excel'
    )


def _import_excel_to_nas(reader, nas, source_day):
    """Map data read from an RJ Excel file into a NightAuditSession.

    Reads cumulative data (JOUR history, DueBack, SD, Depot, SetD) and
    snapshot data (Recap, Transelect, GEAC) from the uploaded RJ.

    Args:
        reader: RJReader instance
        nas: NightAuditSession to populate
        source_day: Day number in the uploaded RJ (the day the file was created for)

    Returns:
        dict with import summary counts
    """
    summary = {'sections': {}, 'total_fields': 0}

    # ─── RECAP ───
    try:
        recap = reader.read_recap()
        field_map = {
            'comptant_lightspeed_lecture': 'cash_ls_lecture',
            'comptant_lightspeed_corr': 'cash_ls_corr',
            'comptant_positouch_lecture': 'cash_pos_lecture',
            'comptant_positouch_corr': 'cash_pos_corr',
            'cheque_payment_lecture': 'cheque_ar_lecture',
            'cheque_payment_corr': 'cheque_ar_corr',
            'remb_gratuite_lecture': 'remb_gratuite_lecture',
            'remb_gratuite_corr': 'remb_gratuite_corr',
            'remb_client_lecture': 'remb_client_lecture',
            'remb_client_corr': 'remb_client_corr',
            'due_back_reception_lecture': 'dueback_reception_lecture',
            'due_back_reception_corr': 'dueback_reception_corr',
            'due_back_nb_lecture': 'dueback_nb_lecture',
            'due_back_nb_corr': 'dueback_nb_corr',
            'surplus_deficit_lecture': 'recap_surplus_lecture',
            'surplus_deficit_corr': 'recap_surplus_corr',
            'depot_canadien_lecture': 'deposit_cdn',
        }
        count = 0
        for excel_field, nas_field in field_map.items():
            val = recap.get(excel_field)
            if val is not None and hasattr(nas, nas_field):
                try:
                    setattr(nas, nas_field, float(val))
                    count += 1
                except (ValueError, TypeError):
                    pass
        summary['sections']['recap'] = count
        summary['total_fields'] += count
    except Exception as e:
        logger.warning(f"Import Recap failed: {e}")
        summary['sections']['recap'] = f'error: {e}'

    # ─── TRANSELECT ───
    try:
        trans = reader.read_transelect()
        # Build restaurant JSON dict
        restaurant = {}
        terminal_map = {
            'bar_701': 'Bar 701', 'bar_702': 'Bar 702', 'bar_703': 'Bar 703',
            'spesa_704': 'Spesa 704', 'room_705': 'Room 705'
        }
        for prefix, term_name in terminal_map.items():
            term_data = {}
            for card in ['debit', 'visa', 'master', 'amex']:
                key = f'{prefix}_{card}'
                val = trans.get(key)
                if val is not None:
                    try:
                        term_data[card if card != 'master' else 'mc'] = float(val)
                    except (ValueError, TypeError):
                        pass
            if term_data:
                restaurant[term_name] = term_data

        # Build reception JSON dict
        reception = {}
        for card in ['debit', 'visa', 'mc', 'amex']:
            card_data = {}
            excel_card = card if card != 'mc' else 'master'
            # fusebox
            val = trans.get(f'fusebox_{excel_card}')
            if val: card_data['fusebox'] = float(val) if val else 0
            # k053 terminal
            val = trans.get(f'reception_{excel_card}_term')
            if val: card_data['k053'] = float(val) if val else 0
            # debit special handling
            if card == 'debit':
                val = trans.get('reception_debit')
                if val: card_data['k053'] = float(val) if val else 0
                val = trans.get('reception_debit_term8')
                if val: card_data['term8'] = float(val) if val else 0
            if card_data:
                reception[card] = card_data

        nas.set_json('transelect_restaurant', restaurant)
        nas.set_json('transelect_reception', reception)
        summary['sections']['transelect'] = len(restaurant) + len(reception)
        summary['total_fields'] += len(restaurant) + len(reception)
    except Exception as e:
        logger.warning(f"Import Transelect failed: {e}")
        summary['sections']['transelect'] = f'error: {e}'

    # ─── GEAC ───
    try:
        geac = reader.read_geac_ux()
        cashout = {}
        daily_rev = {}
        for card, prefix in [('amex', 'amex'), ('master', 'master'), ('visa', 'visa'),
                              ('diners', 'diners'), ('discover', 'discover')]:
            co = geac.get(f'{prefix}_cash_out')
            if co is not None:
                try: cashout[card] = float(co)
                except: pass
            dr = geac.get(f'{prefix}_daily_revenue')
            if dr is not None:
                try: daily_rev[card] = float(dr)
                except: pass

        nas.set_json('geac_cashout', cashout)
        nas.set_json('geac_daily_rev', daily_rev)

        # AR balance fields
        for excel_f, nas_f in [
            ('balance_previous', 'geac_ar_previous'),
            ('balance_today', 'geac_ar_charges'),
            ('new_balance', 'geac_ar_new_balance'),
        ]:
            val = geac.get(excel_f)
            if val is not None and hasattr(nas, nas_f):
                try: setattr(nas, nas_f, float(val))
                except: pass

        summary['sections']['geac'] = len(cashout) + len(daily_rev)
        summary['total_fields'] += len(cashout) + len(daily_rev)
    except Exception as e:
        logger.warning(f"Import GEAC failed: {e}")
        summary['sections']['geac'] = f'error: {e}'

    # ─── JOUR (read source day, or fall back to most recent day with data) ───
    try:
        jour_data = reader.read_jour_day(source_day)
        nas_fields = excel_jour_to_nas_dict(jour_data) if jour_data else {}

        # If the source day has no mapped data, try previous days
        # (the uploaded RJ might not have been filled for this day yet)
        if not nas_fields and source_day > 1:
            for try_day in range(source_day - 1, 0, -1):
                fallback = reader.read_jour_day(try_day)
                if fallback:
                    nas_fields = excel_jour_to_nas_dict(fallback)
                    if nas_fields:
                        logger.info(f"Jour: day {source_day} empty, using day {try_day} data")
                        break

        count = 0
        for field, val in nas_fields.items():
            if hasattr(nas, field):
                try:
                    if field.startswith('jour_rooms_') or field == 'jour_nb_clients' or field == 'jour_rooms_hors_usage':
                        setattr(nas, field, int(val))
                    else:
                        setattr(nas, field, float(val))
                    count += 1
                except (ValueError, TypeError):
                    pass
        summary['sections']['jour'] = count
        summary['total_fields'] += count
    except Exception as e:
        logger.warning(f"Import Jour failed: {e}")
        summary['sections']['jour'] = f'error: {e}'

    # ─── DUEBACK ───
    try:
        dueback = reader.read_dueback(day=source_day)
        if dueback and source_day in dueback.get('days', {}):
            day_data = dueback['days'][source_day]
            receptionists = dueback.get('receptionists', [])

            # Build entries list: yesterday's 'nouveau' becomes today's 'previous'
            # Exception: certain receptionists NEVER carry a previous balance
            entries = []
            for recep in receptionists:
                col = recep['col_letter']
                name = recep.get('full_name', recep.get('last_name', ''))
                # nouveau from the source day becomes previous for the new day
                nouveau_val = day_data.get('nouveau', {}).get(col, {}).get('amount', 0) or 0
                # Previous is always NEGATIVE (debit carried forward)
                prev_val = 0 if _dueback_skip_previous(name) else -abs(float(nouveau_val))
                entries.append({
                    'name': name,
                    'previous': prev_val,
                    'nouveau': 0
                })

            if entries:
                nas.set_json('dueback_entries', entries)
                summary['sections']['dueback'] = len(entries)
                summary['total_fields'] += len(entries)
    except Exception as e:
        logger.warning(f"Import DueBack failed: {e}")
        summary['sections']['dueback'] = f'error: {e}'

    # ─── SETD ───
    try:
        setd_data = reader.read_setd_day(source_day)
        if setd_data:
            # SetD personnel: build list of {name, amount}
            personnel = []
            for col_letter, val in setd_data.items():
                if col_letter in ('A', 'B'):
                    continue  # Skip day number and RJ column
                if val is not None and val != 0:
                    try:
                        personnel.append({
                            'column': col_letter,
                            'amount': float(val)
                        })
                    except (ValueError, TypeError):
                        pass
            if personnel:
                nas.set_json('setd_personnel', personnel)
                summary['sections']['setd'] = len(personnel)
                summary['total_fields'] += len(personnel)
    except Exception as e:
        logger.warning(f"Import SetD failed: {e}")
        summary['sections']['setd'] = f'error: {e}'

    # ─── CONTROLE metadata ───
    try:
        controle = reader.read_controle()
        if controle.get('prepare_par'):
            nas.auditor_name = str(controle['prepare_par'])
        summary['sections']['controle'] = 'ok'
    except Exception as e:
        logger.warning(f"Import Controle failed: {e}")

    return summary


@rj_native_bp.route('/api/rj/native/export/rj-filled/<audit_date>', methods=['GET'])
@auth_required
def export_rj_filled(audit_date):
    """Export RJ Excel file using the uploaded original as base.

    Takes the originally-uploaded RJ file (yesterday's), injects the new day's
    data using RJFiller, and returns the updated .xls file for download.

    If no uploaded file exists, falls back to a blank template.
    """
    try:
        # Parse date
        try:
            d = datetime.strptime(audit_date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Format de date invalide (YYYY-MM-DD)'}), 400

        # Load session
        nas = NightAuditSession.query.filter_by(audit_date=d).first()
        if not nas:
            return jsonify({'error': 'Session non trouvée'}), 404

        # Get the base file: DB archive (source of truth) → cache → template
        base_bytes = None

        # 1) DB archive — primary source (today's live copy, or yesterday's)
        archive = RJArchive.query.filter_by(audit_date=d).first()
        if not archive:
            yesterday = d - timedelta(days=1)
            archive = RJArchive.query.filter_by(audit_date=yesterday).first()
        if archive and archive.file_binary:
            base_bytes = io.BytesIO(archive.file_binary)
            logger.info(f"Export: base RJ from DB archive ({archive.audit_date})")

        # 2) Fallback: memory/disk cache
        if not base_bytes:
            rj_cached = _load_rj(d.isoformat())
            if rj_cached:
                rj_cached.seek(0)
                base_bytes = io.BytesIO(rj_cached.getvalue())

        # 3) Fallback: blank template
        if not base_bytes:
            template_path = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                'RJ 2024-2025', 'RJ 2025-2026', '12-Février 2026', 'Rj Vierge.xls'
            )
            if os.path.exists(template_path):
                with open(template_path, 'rb') as tf:
                    base_bytes = io.BytesIO(tf.read())
            else:
                return jsonify({'error': 'Aucun fichier RJ uploadé et template introuvable'}), 404

        # Create RJFiller from the base file
        filler = RJFiller(base_bytes)

        # Determine the day number for this audit date
        vjour = d.day

        # ─── 1. Update CONTROLE with new date ───
        filler.update_controle(
            vjour=vjour,
            mois=d.month,
            annee=d.year
        )

        # ─── 2. Fill RECAP ───
        recap_data = {
            'comptant_lightspeed_lecture': nas.cash_ls_lecture or 0,
            'comptant_lightspeed_corr': nas.cash_ls_corr or 0,
            'comptant_positouch_lecture': nas.cash_pos_lecture or 0,
            'comptant_positouch_corr': nas.cash_pos_corr or 0,
            'cheque_payment_register_lecture': nas.cheque_ar_lecture or 0,
            'cheque_payment_register_corr': nas.cheque_ar_corr or 0,
            'cheque_daily_revenu_lecture': nas.cheque_dr_lecture or 0,
            'cheque_daily_revenu_corr': nas.cheque_dr_corr or 0,
            'remb_gratuite_lecture': nas.remb_gratuite_lecture or 0,
            'remb_gratuite_corr': nas.remb_gratuite_corr or 0,
            'remb_client_lecture': nas.remb_client_lecture or 0,
            'remb_client_corr': nas.remb_client_corr or 0,
            'due_back_reception_lecture': nas.dueback_reception_lecture or 0,
            'due_back_reception_corr': nas.dueback_reception_corr or 0,
            'due_back_nb_lecture': nas.dueback_nb_lecture or 0,
            'due_back_nb_corr': nas.dueback_nb_corr or 0,
            'prepare_par': nas.auditor_name or '',
        }
        filler.fill_sheet('Recap', recap_data)

        # ─── 3. Fill TRANSELECT ───
        rest_data = nas.get_json('transelect_restaurant')
        recep_data = nas.get_json('transelect_reception')
        trans_dict = {}
        # Map restaurant terminals to TRANSELECT_MAPPING keys
        term_to_prefix = {
            'Bar 701': 'bar_701', 'Bar 702': 'bar_702', 'Bar 703': 'bar_703',
            'Spesa 704': 'spesa_704', 'Room 705': 'room_705'
        }
        for term_name, prefix in term_to_prefix.items():
            td = rest_data.get(term_name, {}) if rest_data else {}
            for card, card_key in [('debit', 'debit'), ('visa', 'visa'), ('mc', 'master'), ('amex', 'amex')]:
                val = td.get(card, 0)
                if val:
                    trans_dict[f'{prefix}_{card_key}'] = val

        # Map reception data
        for card, card_key in [('debit', 'debit'), ('visa', 'visa'), ('mc', 'master'), ('amex', 'amex')]:
            cd = recep_data.get(card, {}) if recep_data else {}
            if card == 'debit':
                if cd.get('k053'): trans_dict['reception_debit'] = cd['k053']
                if cd.get('term8'): trans_dict['reception_debit_term8'] = cd['term8']
            else:
                if cd.get('fusebox'): trans_dict[f'fusebox_{card_key}'] = cd['fusebox']
                if cd.get('k053'): trans_dict[f'reception_{card_key}_term'] = cd['k053']

        if trans_dict:
            filler.fill_sheet('transelect', trans_dict)

        # ─── 4. Fill GEAC ───
        geac_co = nas.get_json('geac_cashout')
        geac_dr = nas.get_json('geac_daily_rev')
        geac_dict = {}
        for card in ['amex', 'diners', 'master', 'visa', 'discover']:
            if geac_co.get(card): geac_dict[f'{card}_cash_out'] = geac_co[card]
            if geac_dr.get(card): geac_dict[f'{card}_daily_revenue'] = geac_dr[card]
        if geac_dict:
            filler.fill_sheet('geac_ux', geac_dict)

        # ─── 5. Fill JOUR for the new day ───
        jour_dict = nas_jour_to_excel_dict(nas)
        if jour_dict:
            filler.fill_jour_day(vjour, jour_dict)

        # Also run macro equivalents: envoie_dans_jour and calcul_carte
        try:
            filler.envoie_dans_jour(vjour)
        except Exception as e:
            logger.warning(f"envoie_dans_jour failed during export: {e}")

        try:
            filler.calcul_carte(vjour)
        except Exception as e:
            logger.warning(f"calcul_carte failed during export: {e}")

        # ─── 6. Fill DueBack entries ───
        dueback_entries = nas.get_json('dueback_entries')
        if isinstance(dueback_entries, list):
            for entry in dueback_entries:
                name = entry.get('name', '')
                previous = entry.get('previous', 0) or 0
                nouveau = entry.get('nouveau', 0) or 0
                # Try to find the receptionist column
                from utils.rj_mapper import DUEBACK_RECEPTIONIST_COLUMNS
                col_letter = None
                for rname, rletter in DUEBACK_RECEPTIONIST_COLUMNS.items():
                    if rname.lower() in name.lower() or name.lower() in rname.lower():
                        col_letter = rletter
                        break
                if col_letter:
                    try:
                        if previous:
                            filler.fill_dueback_by_col(vjour, col_letter, previous, 'previous')
                        if nouveau:
                            filler.fill_dueback_by_col(vjour, col_letter, nouveau, 'nouveau')
                    except Exception as e:
                        logger.warning(f"DueBack fill error for {name}: {e}")

        # ─── 7. Save — rebuild with VBA from original ───
        modified_bytes = filler.save_to_bytes().getvalue()

        # Rebuild OLE file: combine modified Workbook stream with
        # original VBA macros and metadata streams
        rj_orig = _load_rj(d.isoformat())
        if rj_orig:
            rj_orig.seek(0)
            orig_bytes = rj_orig.getvalue()
            try:
                final_bytes = rebuild_xls_with_vba(orig_bytes, modified_bytes)
                output = io.BytesIO(final_bytes)
            except Exception as e:
                logger.warning(f"VBA rebuild failed, using plain export: {e}")
                output = io.BytesIO(modified_bytes)
        else:
            # No original file → plain export (no VBA to preserve)
            output = io.BytesIO(modified_bytes)

        filename = f"Rj {d.strftime('%d-%m-%Y')}.xls"

        # ★ Auto-archive the exported RJ to DB (output of tonight's work)
        try:
            output.seek(0)
            _archive_rj_to_db(
                output.getvalue(), d,
                source_filename=filename,
                uploaded_by=session.get('user_name', 'export')
            )
            output.seek(0)
        except Exception as arc_err:
            logger.warning(f"Archive export failed (non-blocking): {arc_err}")
            output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.ms-excel'
        )

    except Exception as e:
        logger.error(f"RJ filled export error: {e}", exc_info=True)
        return jsonify({'error': f"Erreur lors de l'export: {str(e)}"}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# PARSE & FILL — Upload Lightspeed reports → auto-fill NAS fields
# ═══════════════════════════════════════════════════════════════════════════════

@rj_native_bp.route('/api/rj/native/parse-and-fill', methods=['POST'])
@auth_required
def native_parse_and_fill():
    """Parse an uploaded Lightspeed report and auto-fill NAS fields.

    Form data:
        doc_type: 'daily_revenue' | 'freedompay' | 'advance_deposit' |
                  'ar_summary' | 'hp_excel' | 'sales_journal' | 'sd_deposit'
        file: uploaded document
        date: audit session date (YYYY-MM-DD)
        day: (optional) day of month for HP/SD parsers

    Returns JSON with filled fields count and details.
    """
    from utils.parsers import ParserFactory

    doc_type = request.form.get('doc_type')
    f = request.files.get('file')
    audit_date_str = request.form.get('date')

    if not doc_type:
        return jsonify({'error': 'Type de document requis'}), 400
    if not f:
        return jsonify({'error': 'Fichier requis'}), 400
    if not audit_date_str:
        return jsonify({'error': 'Date de session requise'}), 400

    # Load NAS
    try:
        d = datetime.strptime(audit_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400

    nas = NightAuditSession.query.filter_by(audit_date=d).first()
    if not nas:
        return jsonify({'error': 'Session non trouvée. Créez ou importez une session d\'abord.'}), 404
    if nas.status == 'locked':
        return jsonify({'error': 'Session verrouillée'}), 403

    try:
        file_bytes = f.read()

        # Extra kwargs per parser type
        extra_kwargs = {}
        day_param = request.form.get('day')
        if day_param:
            try:
                extra_kwargs['day'] = int(day_param)
            except (ValueError, TypeError):
                extra_kwargs['day'] = d.day
        elif doc_type in ('hp_excel', 'sd_deposit'):
            extra_kwargs['day'] = d.day

        # Parse
        parser = ParserFactory.create(doc_type, file_bytes, filename=f.filename, **extra_kwargs)
        result = parser.get_result()

        if not result.get('success', True):
            return jsonify({
                'error': result.get('error', 'Erreur de parsing'),
                'warnings': result.get('warnings', [])
            }), 400

        # ★ Store source file for SD and HP (for write-back on save)
        if doc_type in ('sd_deposit', 'hp_excel'):
            from routes.audit.rj_core import SD_FILES, HP_FILES, get_session_id
            session_id = get_session_id()
            file_buf = io.BytesIO(file_bytes)
            if doc_type == 'sd_deposit':
                SD_FILES[session_id] = file_buf
            elif doc_type == 'hp_excel':
                HP_FILES[session_id] = file_buf

        # Map parsed data to NAS
        fill_result = _apply_parsed_to_nas(doc_type, result, nas, extra_kwargs.get('day', d.day))

        # Recalculate and save
        nas.calculate_all()
        db.session.commit()

        filled_count = fill_result.get('count', 0)
        warnings = result.get('warnings', [])

        # ★ If file was stored but no fields filled (empty day), add helpful message
        if doc_type in ('sd_deposit', 'hp_excel') and filled_count == 0:
            file_label = 'SD' if doc_type == 'sd_deposit' else 'HP'
            warnings.append(
                f'Fichier {file_label} chargé — aucune donnée pour le jour {extra_kwargs.get("day", d.day)}. '
                f'Remplissez manuellement et les données seront écrites dans le fichier.'
            )

        return jsonify({
            'success': True,
            'doc_type': doc_type,
            'confidence': result.get('confidence', 0),
            'fields_filled': filled_count,
            'sections_updated': fill_result.get('sections', []),
            'details': fill_result.get('details', {}),
            'warnings': warnings,
            'session': nas.to_dict(),
            'message': f'{filled_count} champs remplis depuis {f.filename}'
        })

    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        logger.error(f"Parse-and-fill error: {e}", exc_info=True)
        return jsonify({'error': f"Erreur: {str(e)}"}), 500


@rj_native_bp.route('/api/rj/native/parser-types', methods=['GET'])
@auth_required
def native_parser_types():
    """Return available parser types and metadata for the UI."""
    from utils.parsers import ParserFactory
    return jsonify(ParserFactory.get_type_info())


@rj_native_bp.route('/api/rj/native/download/<file_type>', methods=['GET'])
@auth_required
def native_download_file(file_type):
    """Download SD or HP file (with write-back modifications).

    Args:
        file_type: 'sd' or 'hp'
    """
    from routes.audit.rj_core import SD_FILES, HP_FILES, get_session_id

    session_id = get_session_id()

    if file_type == 'sd':
        if session_id not in SD_FILES:
            return jsonify({'error': 'Aucun fichier SD uploadé'}), 404
        file_buf = SD_FILES[session_id]
        file_buf.seek(0)
        return send_file(
            io.BytesIO(file_buf.read()),
            mimetype='application/vnd.ms-excel',
            as_attachment=True,
            download_name=f'SD_modifie_{datetime.now().strftime("%Y-%m-%d")}.xls'
        )
    elif file_type == 'hp':
        if session_id not in HP_FILES:
            return jsonify({'error': 'Aucun fichier HP uploadé'}), 404
        file_buf = HP_FILES[session_id]
        file_buf.seek(0)
        return send_file(
            io.BytesIO(file_buf.read()),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'HP_modifie_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
        )
    else:
        return jsonify({'error': f'Type invalide: {file_type}'}), 400


@rj_native_bp.route('/api/rj/native/print/<report_type>', methods=['GET'])
@auth_required
def native_print_report(report_type):
    """Render a printable report page that reproduces the Excel tab layout.

    Args:
        report_type: 'sd', 'recap', 'geac', or 'hp'
    """
    if report_type not in ('sd', 'recap', 'geac', 'hp'):
        return jsonify({'error': f'Type invalide: {report_type}'}), 400

    date_str = request.args.get('date')
    if not date_str:
        return jsonify({'error': 'Paramètre date requis'}), 400

    try:
        d = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide (YYYY-MM-DD)'}), 400

    nas = NightAuditSession.query.filter_by(audit_date=d).first()
    if not nas:
        return jsonify({'error': f'Aucune session pour {date_str}'}), 404

    def _nf(attr):
        """Get numeric field from NAS."""
        v = getattr(nas, attr, None)
        return float(v) if v is not None else 0.0

    ctx = {
        'report_type': report_type,
        'audit_date_display': d.strftime('%d/%m/%Y'),
        'audit_day': d.day,
        'auditor_name': nas.auditor_name or '',
    }

    # ── SD ──────────────────────────────────────────────────────────────
    if report_type == 'sd':
        ctx['title'] = f'SD — Sommaire des Dépôts {d.strftime("%d/%m/%Y")}'
        sd_raw = nas.get_json('sd_entries') or []
        sd_entries = []
        for e in sd_raw:
            amt = float(e.get('amount', 0) or 0)
            ver = float(e.get('verified', 0) or 0)
            remb = float(e.get('reimbursement', 0) or 0)
            variance = round(amt - ver - remb, 2)
            sd_entries.append({
                'department': e.get('department', ''),
                'name': e.get('name', ''),
                'currency': e.get('currency', 'CDN'),
                'amount': amt,
                'verified': ver,
                'reimbursement': remb,
                'variance': variance,
            })
        ctx['sd_entries'] = sd_entries
        ctx['sd_totals'] = {
            'montant': sum(e['amount'] for e in sd_entries),
            'verifie': sum(e['verified'] for e in sd_entries),
            'remb': sum(e['reimbursement'] for e in sd_entries),
            'variance': sum(e['variance'] for e in sd_entries),
        }

    # ── RECAP (Lecture / Correction / Net) ──────────────────────────────
    elif report_type == 'recap':
        ctx['title'] = f'Recap {d.strftime("%d/%m/%Y")}'

        # Build 8 lines with lecture + correction + net
        lines = [
            ('Comptant LightSpeed', 'cash_ls'),
            ('Comptant POSitouch', 'cash_pos'),
            ('Chèque A/R', 'cheque_ar'),
            ('Chèque Daily Rev', 'cheque_dr'),
            ('Remb. gratuite', 'remb_gratuite'),
            ('Remb. client', 'remb_client'),
            ('DueBack Réception', 'dueback_reception'),
            ('DueBack N/B', 'dueback_nb'),
        ]
        recap = {}
        for _label, prefix in lines:
            lec = _nf(f'{prefix}_lecture')
            cor = _nf(f'{prefix}_corr')
            recap[f'{prefix}_lecture'] = lec
            recap[f'{prefix}_corr'] = cor
            recap[f'{prefix}_net'] = round(lec + cor, 2)

        # Subtotals — cash (first 4 lines)
        recap['sub_cash_lecture'] = round(sum(recap[f'{p}_lecture'] for _, p in lines[:4]), 2)
        recap['sub_cash_corr'] = round(sum(recap[f'{p}_corr'] for _, p in lines[:4]), 2)
        recap['sub_cash_net'] = round(sum(recap[f'{p}_net'] for _, p in lines[:4]), 2)

        # Subtotals — refunds (lines 4-5)
        recap['sub_remb_lecture'] = round(sum(recap[f'{p}_lecture'] for _, p in lines[4:6]), 2)
        recap['sub_remb_corr'] = round(sum(recap[f'{p}_corr'] for _, p in lines[4:6]), 2)
        recap['sub_remb_net'] = round(sum(recap[f'{p}_net'] for _, p in lines[4:6]), 2)

        recap['deposit_cdn'] = _nf('deposit_cdn')
        recap['deposit_us'] = _nf('deposit_us')

        # Balance = cash_net + remb_net + dueback_net − deposits
        total_net = recap['sub_cash_net'] + recap['sub_remb_net'] \
                    + recap['dueback_reception_net'] + recap['dueback_nb_net']
        recap['balance'] = round(total_net - recap['deposit_cdn'] - recap['deposit_us'], 2)

        ctx['recap'] = recap

    # ── GEAC ────────────────────────────────────────────────────────────
    elif report_type == 'geac':
        ctx['title'] = f'GEAC {d.strftime("%d/%m/%Y")}'

        cashout = nas.get_json('geac_cashout') or {}
        daily_rev = nas.get_json('geac_daily_rev') or {}

        # Card types — 2-line cashout + daily revenue
        card_keys = [
            ('AMEX', 'amex'), ('DINERS', 'diners'), ('MASTER', 'mc'), ('VISA', 'visa'),
        ]
        geac_cards = []
        totals_co1 = totals_co2 = totals_co = totals_dr = 0
        for label, key in card_keys:
            co1 = float(cashout.get(f'{key}_line1', 0) or cashout.get(key, 0) or 0)
            co2 = float(cashout.get(f'{key}_line2', 0) or 0)
            co_total = round(co1 + co2, 2)
            dr = float(daily_rev.get(key, 0) or 0)
            geac_cards.append({
                'label': label, 'co1': co1, 'co2': co2,
                'co_total': co_total, 'daily_rev': dr,
            })
            totals_co1 += co1
            totals_co2 += co2
            totals_co += co_total
            totals_dr += dr

        card_variance = round(totals_co - totals_dr, 2)
        ctx['geac_cards'] = geac_cards
        ctx['geac_totals'] = {
            'co1': round(totals_co1, 2), 'co2': round(totals_co2, 2),
            'co_total': round(totals_co, 2), 'daily_rev': round(totals_dr, 2),
            'card_variance': card_variance,
        }

        # Balance Sheet
        bs = nas.get_json('geac_balance_sheet') or {}
        ctx['bs'] = {
            'prev_dr': float(bs.get('prev_dr', 0) or 0),
            'prev_gl': float(bs.get('prev_gl', 0) or 0),
            'today_dr': float(bs.get('today_dr', 0) or 0),
            'today_gl': float(bs.get('today_gl', 0) or 0),
            'facture_dr': float(bs.get('facture_dr', 0) or 0),
            'facture_ar': float(bs.get('facture_ar', 0) or 0),
            'advdep_dr': float(bs.get('advdep_dr', 0) or 0),
            'advdep_ad': float(bs.get('advdep_ad', 0) or 0),
            'newbal_dr': float(bs.get('newbal_dr', 0) or 0),
            'newbal_gl': float(bs.get('newbal_gl', 0) or 0),
        }

    # ── HP ──────────────────────────────────────────────────────────────
    elif report_type == 'hp':
        ctx['title'] = f'HP — Rapport Quotidien {d.strftime("%d/%m/%Y")}'

        hp_raw = nas.get_json('hp_admin_entries') or []

        # ── Page 1: DONNÉES — transactions brutes du jour ──────────────
        hp_donnees = []
        for e in hp_raw:
            entry = {
                'area': e.get('area', ''),
                'nourriture': float(e.get('nourriture', 0) or 0),
                'boisson': float(e.get('boisson', 0) or 0),
                'biere': float(e.get('biere', 0) or 0),
                'vin': float(e.get('vin', 0) or 0),
                'mineraux': float(e.get('mineraux', 0) or 0),
                'tabagie': float(e.get('tabagie', 0) or 0),
                'autres': float(e.get('autre', 0) or 0),
                'pourboire': float(e.get('pourboire', 0) or 0),
                'paiement': e.get('raison', ''),     # raison = "Promotion" or "Administration"
                'raison': e.get('raison', ''),
                'qui': e.get('autorise_par', ''),
            }
            # Map raison → paiement code
            r_lower = entry['raison'].lower()
            if 'promo' in r_lower:
                entry['paiement_code'] = '15 - Promotion'
            elif 'admin' in r_lower:
                entry['paiement_code'] = '14 - Administration'
            else:
                entry['paiement_code'] = entry['raison']
            entry['total'] = round(sum([
                entry['nourriture'], entry['boisson'], entry['biere'],
                entry['vin'], entry['mineraux'], entry['tabagie'],
                entry['autres'], entry['pourboire']
            ]), 2)
            hp_donnees.append(entry)

        # Totals for données
        don_totals = {}
        for k in ['nourriture', 'boisson', 'biere', 'vin', 'mineraux',
                   'tabagie', 'autres', 'pourboire', 'total']:
            don_totals[k] = round(sum(e[k] for e in hp_donnees), 2)
        ctx['hp_donnees'] = hp_donnees
        ctx['hp_don_totals'] = don_totals

        # ── Page 2: JOURNALIER — pivot Area × Catégorie, Promo vs Admin ──
        from collections import defaultdict, OrderedDict
        cats = ['nourriture', 'tabagie', 'boisson', 'biere', 'mineraux',
                'vin', 'autres', 'pourboire', 'total']
        cat_labels = {
            'nourriture': '00',  # "00" is what Excel shows for Nourriture
            'tabagie': 'Sum of Tabagie',
            'boisson': 'Somme de Boisson',
            'biere': 'Somme de Biere',
            'mineraux': 'Somme de Mineraux',
            'vin': 'Somme de Vin',
            'autres': 'Somme de Autres',
            'pourboire': 'Somme de Pourboire',
            'total': 'Somme de Total',
        }

        # Aggregate by area → paiement_type → category
        pivot = defaultdict(lambda: {'promo': defaultdict(float), 'admin': defaultdict(float)})
        for e in hp_donnees:
            area = e['area']
            pmt_key = 'promo' if '15' in e.get('paiement_code', '') or 'promo' in e['raison'].lower() else 'admin'
            for cat in cats:
                if cat == 'total':
                    pivot[area][pmt_key]['total'] += e['total']
                else:
                    pivot[area][pmt_key][cat] += e[cat]

        # Build ordered pivot rows (matching Excel Journalier layout)
        # Areas in order they appear
        area_order = []
        for e in hp_donnees:
            if e['area'] and e['area'] not in area_order:
                area_order.append(e['area'])

        journalier_areas = []
        grand_totals = {'promo': defaultdict(float), 'admin': defaultdict(float)}

        for area in area_order:
            area_data = pivot[area]
            rows = []
            for cat in cats:
                promo = round(area_data['promo'][cat], 2)
                admin = round(area_data['admin'][cat], 2)
                total = round(promo + admin, 2)
                if total != 0 or cat == 'total':
                    rows.append({
                        'label': cat_labels[cat],
                        'promo': promo,
                        'admin': admin,
                        'total': total,
                        'is_total': cat == 'total',
                    })
                grand_totals['promo'][cat] += area_data['promo'][cat]
                grand_totals['admin'][cat] += area_data['admin'][cat]
            journalier_areas.append({'name': area, 'rows': rows})

        # Grand total rows
        grand_rows = []
        for cat in cats:
            promo = round(grand_totals['promo'][cat], 2)
            admin = round(grand_totals['admin'][cat], 2)
            total = round(promo + admin, 2)
            if total != 0 or cat == 'total':
                grand_rows.append({
                    'label': f'Total {cat_labels[cat]}',
                    'promo': promo,
                    'admin': admin,
                    'total': total,
                    'is_total': cat == 'total',
                })

        ctx['journalier_areas'] = journalier_areas
        ctx['journalier_grand'] = grand_rows

    return render_template('audit/rj/print_reports.html', **ctx)


def _apply_parsed_to_nas(doc_type, result, nas, day):
    """Apply parsed document data to a NightAuditSession.

    Routes to the appropriate mapper function based on doc_type.

    Returns:
        dict with 'count', 'sections', 'details'
    """
    data = result.get('data', {})

    mappers = {
        'daily_revenue': _fill_from_daily_revenue,
        'freedompay': _fill_from_freedompay,
        'advance_deposit': _fill_from_advance_deposit,
        'ar_summary': _fill_from_ar_summary,
        'hp_excel': _fill_from_hp_excel,
        'sales_journal': _fill_from_sales_journal,
        'sd_deposit': _fill_from_sd_deposit,
    }

    mapper_fn = mappers.get(doc_type)
    if not mapper_fn:
        return {'count': 0, 'sections': [], 'details': {'error': f'No mapper for {doc_type}'}}

    return mapper_fn(data, nas, day)


def _safe_set(nas, field, value, count_ref):
    """Safely set a NAS field, incrementing count_ref[0] on success."""
    if value is None:
        return
    if not hasattr(nas, field):
        return
    try:
        if isinstance(value, str):
            setattr(nas, field, value)
        elif field.startswith('jour_rooms_') or field in ('jour_nb_clients', 'jour_rooms_hors_usage'):
            setattr(nas, field, int(float(value)))
        else:
            setattr(nas, field, float(value))
        count_ref[0] += 1
    except (ValueError, TypeError):
        pass


# ─── DAILY REVENUE ─── (the big one: fills Recap + GEAC + Jour)
def _fill_from_daily_revenue(data, nas, day):
    count = [0]
    sections = []

    # --- RECAP ---
    settlements = data.get('settlements', {})
    _safe_set(nas, 'cash_ls_lecture', settlements.get('comptant'), count)
    _safe_set(nas, 'cheque_dr_lecture', settlements.get('cheque'), count)
    if count[0] > 0:
        sections.append('recap')

    # --- GEAC ---
    rj = data.get('rj_mapping', {})
    geac_map = rj.get('geac_ux', {})

    # NOTE: Daily Revenue PDF "settlements" = GEAC "Daily Revenue" row (R11),
    # NOT the "Cashout" row (R9). Cashout comes from FreedomPay/bank report.
    # We do NOT fill geac_cashout from Daily Revenue.

    # Daily Revenue (R11) — from settlements in Daily Rev PDF
    geac_dr = nas.get_json('geac_daily_rev')
    settle_card_map = {
        'american_express': 'amex',
        'visa': 'visa',
        'mastercard': 'master',
        'diners': 'diners',
        'discover': 'discover',
    }
    for settle_key, geac_key in settle_card_map.items():
        val = settlements.get(settle_key)
        if val is not None and abs(float(val)) > 0:
            geac_dr[geac_key] = abs(float(val))
            count[0] += 1
    nas.set_json('geac_daily_rev', geac_dr)

    # AR balances — write to old scalar fields for backward compatibility
    _safe_set(nas, 'geac_ar_previous', geac_map.get('balance_prev_day'), count)
    _safe_set(nas, 'geac_ar_charges', geac_map.get('balance_today'), count)
    _safe_set(nas, 'geac_ar_new_balance', geac_map.get('new_balance'), count)

    # ★ Balance Sheet JSON — Daily Revenue column (DR) values
    # These are the first column values in the GEAC Balance Sheet
    balance = data.get('balance', {})
    adv_deps = data.get('advance_deposits', {})
    bs = nas.get_json('geac_balance_sheet')
    if not isinstance(bs, dict):
        bs = {}
    bs_changed = False

    # Balance Previous Day (DR) — from Daily Revenue balance section
    prev_day = balance.get('prev_day') or geac_map.get('balance_prev_day')
    if prev_day is not None:
        bs['prev_dr'] = abs(float(prev_day))
        bs_changed = True
        count[0] += 1

    # Balance Today (DR)
    today_bal = balance.get('today') or geac_map.get('balance_today')
    if today_bal is not None:
        bs['today_dr'] = abs(float(today_bal))
        bs_changed = True
        count[0] += 1

    # Facture Direct (DR) — from settlements
    facture_direct = settlements.get('facture_direct')
    if facture_direct is not None and abs(float(facture_direct)) > 0:
        bs['facture_dr'] = abs(float(facture_direct))
        bs_changed = True
        count[0] += 1

    # Advance Deposit Applied (DR)
    adv_applied = adv_deps.get('applied')
    if adv_applied is not None and abs(float(adv_applied)) > 0:
        bs['advdep_dr'] = abs(float(adv_applied))
        bs_changed = True
        count[0] += 1

    # New Balance (DR)
    new_bal = balance.get('new_balance') or geac_map.get('new_balance')
    if new_bal is not None:
        bs['newbal_dr'] = abs(float(new_bal))
        bs_changed = True
        count[0] += 1

    if bs_changed:
        nas.set_json('geac_balance_sheet', bs)

    if bs_changed or 'geac_ux' in str(geac_map):
        sections.append('geac')

    # --- TRANSELECT RECEPTION — daily_rev column ---
    # FreedomPay = same values as Daily Rev settlements (user confirmed).
    # So we copy settlements directly to transelect_reception.daily_rev column.
    recep = nas.get_json('transelect_reception')
    transelect_card_map = {
        'american_express': 'amex',
        'visa': 'visa',
        'mastercard': 'mc',
        'discover': 'discover',
        'carte_debit': 'debit',
    }
    transelect_changed = False
    for settle_key, card_type in transelect_card_map.items():
        val = settlements.get(settle_key)
        if val is not None and abs(float(val)) > 0:
            if card_type not in recep:
                recep[card_type] = {}
            recep[card_type]['daily_rev'] = abs(float(val))
            count[0] += 1
            transelect_changed = True
    if transelect_changed:
        nas.set_json('transelect_reception', recep)
        sections.append('transelect')

    # --- JOUR ---
    revenue = data.get('revenue', {})
    non_rev = data.get('non_revenue', {})

    # Room revenue
    chambres = revenue.get('chambres', {})
    _safe_set(nas, 'jour_room_revenue', chambres.get('total'), count)

    # Taxes — collect from ALL Daily Rev department sections:
    # chambres_tax + restaurant_piazza + banquet + services_chambres + la_spesa + internet_nonrev
    # NOTE: "autres_revenus_nonrev" taxes are EXCLUDED — they are non-revenue deductions
    # that don't belong in the Jour TPS/TVQ totals.
    # These are set here (not additive) because Daily Rev is always parsed FIRST.
    # Sales Journal taxes are then ADDED on top in _fill_from_sales_journal().
    taxes = non_rev.get('chambres_tax', {})
    _safe_set(nas, 'jour_taxe_hebergement', taxes.get('taxe_hebergement'), count)

    # Sum TPS/TVQ from Daily Rev NON-F&B sections only.
    # F&B department taxes (restaurant_piazza, banquet, services_chambres, la_spesa)
    # are ALREADY included in the Sales Journal total — do NOT double-count them.
    # Only include: chambres_tax + autres_revenus_nonrev + internet_nonrev
    dr_tps = float(taxes.get('tps') or 0)
    dr_tvq = float(taxes.get('tvq') or 0)
    for section_key in ['autres_revenus_nonrev', 'internet_nonrev']:
        section = non_rev.get(section_key, {})
        dr_tps += float(section.get('tps') or section.get('tps_autres') or 0)
        dr_tvq += float(section.get('tvq') or section.get('tvq_autres') or 0)

    if dr_tps != 0:
        _safe_set(nas, 'jour_tps', round(dr_tps, 2), count)
    if dr_tvq != 0:
        _safe_set(nas, 'jour_tvq', round(dr_tvq, 2), count)

    # Telephones
    _safe_set(nas, 'jour_tel_local', revenue.get('telephones', {}).get('total'), count)

    # Internet
    _safe_set(nas, 'jour_internet', revenue.get('internet', {}).get('total'), count)

    # Autres revenus
    autres = revenue.get('autres_revenus', {})
    _safe_set(nas, 'jour_massage', autres.get('massage'), count)
    _safe_set(nas, 'jour_location_salle', autres.get('location_salle_forfait'), count)

    # Comptabilité / GL
    compta = revenue.get('comptabilite', {})
    _safe_set(nas, 'jour_autres_gl', compta.get('autres_grand_livre'), count)

    # Gift cards (Givex)
    _safe_set(nas, 'jour_gift_cards', revenue.get('givex', {}).get('total'), count)

    # Club lounge
    _safe_set(nas, 'jour_club_lounge', non_rev.get('club_lounge', {}).get('total'), count)

    # F&B departments (Piazza, Banquet, Services Chambres, La Spesa, Café):
    # These are NOT filled from Daily Revenue because the Sales Journal is the
    # source of truth for POS F&B data. Daily Revenue includes adjustments and
    # deductions (e.g. banquet alcool=-600) that don't belong in the jour fields.
    # The Sales Journal parser (_fill_from_sales_journal) handles these fields.

    # Pourboires
    _safe_set(nas, 'jour_pourboires', non_rev.get('services_chambres', {}).get('pourboire'), count)

    # Balance (for A/R)
    balance = data.get('balance', {})
    _safe_set(nas, 'jour_ar_misc', balance.get('hotel_moved_out'), count)
    _safe_set(nas, 'jour_deposit_on_hand', data.get('advance_deposits', {}).get('applied'), count)

    # Occupation from Daily Revenue (rooms sold = balance today / ADR if available)
    if count[0] > 0:
        sections.append('jour')

    # --- Débourse → Recap (Remboursement Client) ---
    debourse = non_rev.get('debourse', {})
    debourse_total = debourse.get('total')
    if debourse_total is not None and float(debourse_total) != 0:
        _safe_set(nas, 'remb_client_lecture', -abs(float(debourse_total)), count)
        sections.append('recap')

    # --- DueBack N/B → Recap ---
    comptabilite = non_rev.get('comptabilite', {})
    dueback_nb = comptabilite.get('due_back_nourriture')
    if dueback_nb is not None and float(dueback_nb) != 0:
        _safe_set(nas, 'dueback_nb_lecture', abs(float(dueback_nb)), count)
        sections.append('recap')

    # --- DueBack Réception → Recap (from débourse total) ---
    if debourse_total is not None and float(debourse_total) != 0:
        _safe_set(nas, 'dueback_reception_lecture', abs(float(debourse_total)), count)

    return {'count': count[0], 'sections': list(set(sections)), 'details': {'doc_type': 'daily_revenue'}}


# ─── FREEDOMPAY ─── (GEAC cashout + Transelect fusebox)
def _fill_from_freedompay(data, nas, day):
    count = [0]
    sections = []

    # GEAC cashout
    geac_co = nas.get_json('geac_cashout')
    for card_src, card_dst in [('amex_cash_out', 'amex'), ('master_cash_out', 'master'),
                                ('visa_cash_out', 'visa'), ('diners_cash_out', 'diners'),
                                ('discover_cash_out', 'discover')]:
        val = data.get(card_src)
        if val is not None:
            geac_co[card_dst] = float(val)
            count[0] += 1
    nas.set_json('geac_cashout', geac_co)

    # GEAC daily rev
    geac_dr = nas.get_json('geac_daily_rev')
    for card_src, card_dst in [('amex_daily_revenue', 'amex'), ('mastercard_daily_revenue', 'master'),
                                ('visa_daily_revenue', 'visa'), ('diners_daily_revenue', 'diners'),
                                ('discover_daily_revenue', 'discover')]:
        val = data.get(card_src)
        if val is not None:
            geac_dr[card_dst] = float(val)
            count[0] += 1
    nas.set_json('geac_daily_rev', geac_dr)
    if count[0] > 0:
        sections.append('geac')

    # Transelect fusebox (reception)
    recep = nas.get_json('transelect_reception')
    for card_src, card_type in [('fusebox_visa', 'visa'), ('fusebox_master', 'mc'),
                                 ('fusebox_amex', 'amex')]:
        val = data.get(card_src)
        if val is not None:
            if card_type not in recep:
                recep[card_type] = {}
            recep[card_type]['fusebox'] = float(val)
            count[0] += 1
    nas.set_json('transelect_reception', recep)
    if 'fusebox_visa' in data or 'fusebox_master' in data:
        sections.append('transelect')

    return {'count': count[0], 'sections': sections, 'details': {'doc_type': 'freedompay'}}


# ─── ADVANCE DEPOSIT ─── (GEAC ligne 44)
def _fill_from_advance_deposit(data, nas, day):
    count = [0]
    sections = set()

    # Jour deposit on hand (existing behavior)
    _safe_set(nas, 'jour_deposit_on_hand', data.get('deposit_on_hand'), count)
    if count[0] > 0:
        sections.add('jour')

    # ★ Balance Sheet JSON — Advance Deposit column (advdep_ad)
    adv_deposit = data.get('adv_deposit') or data.get('deposit_on_hand')
    if adv_deposit is not None:
        bs = nas.get_json('geac_balance_sheet')
        if not isinstance(bs, dict):
            bs = {}
        bs['advdep_ad'] = abs(float(adv_deposit))
        nas.set_json('geac_balance_sheet', bs)
        count[0] += 1
        sections.add('geac')

    return {'count': count[0], 'sections': list(sections),
            'details': {'doc_type': 'advance_deposit'}}


# ─── AR SUMMARY ─── (GEAC AR section)
def _fill_from_ar_summary(data, nas, day):
    count = [0]
    rj = data.get('rj_mapping', {}).get('geac_ux', {})

    # Old scalar fields (backward compatibility)
    _safe_set(nas, 'geac_ar_previous', rj.get('ar_previous_balance') or data.get('ar_balance_previous'), count)
    _safe_set(nas, 'geac_ar_new_balance', rj.get('ar_end_of_day') or data.get('ar_balance_end_of_day'), count)

    # Front office transfers → AR charges
    transfers = data.get('front_office_transfers', {})
    _safe_set(nas, 'geac_ar_charges', transfers.get('total'), count)

    # Payments
    _safe_set(nas, 'geac_ar_payments', data.get('payments'), count)

    # ★ Balance Sheet JSON — A/R Summary column (facture_ar)
    # The AR Summary "Front Office Transfers total" maps to the Facture Direct AR column
    bs = nas.get_json('geac_balance_sheet')
    if not isinstance(bs, dict):
        bs = {}
    bs_changed = False

    ar_transfers = rj.get('ar_transfers') or (transfers.get('total') if transfers else None)
    if ar_transfers is not None:
        bs['facture_ar'] = abs(float(ar_transfers))
        bs_changed = True
        count[0] += 1

    if bs_changed:
        nas.set_json('geac_balance_sheet', bs)

    return {'count': count[0], 'sections': ['geac'] if count[0] > 0 else [],
            'details': {'doc_type': 'ar_summary', 'balanced': data.get('balanced', False)}}


# ─── HP EXCEL ─── (Jour HP deductions + HP/Admin tab)
def _fill_from_hp_excel(data, nas, day):
    count = [0]

    # HP deductions go into jour_* fields via column mapping
    jour_deductions = data.get('jour_deductions', {})
    # Also check day-specific deductions
    day_key = f'day_{day}_deductions'
    if day_key in data:
        jour_deductions = data[day_key]

    # Map jour column indices to NAS fields
    # Jour column indices → NAS field names
    # See Excel 'jour' sheet for column layout
    hp_col_to_nas = {
        9: 'jour_piazza_nourriture',
        10: 'jour_piazza_boisson',
        11: 'jour_piazza_bieres',
        12: 'jour_piazza_mineraux',
        13: 'jour_piazza_vins',
        14: 'jour_spesa_nourriture',
        15: 'jour_spesa_boisson',
        16: 'jour_spesa_bieres',
        17: 'jour_spesa_mineraux',
        18: 'jour_spesa_vins',
        19: 'jour_chambres_svc_nourriture',
        20: 'jour_chambres_svc_boisson',
        21: 'jour_chambres_svc_bieres',
        22: 'jour_chambres_svc_mineraux',
        23: 'jour_chambres_svc_vins',
        24: 'jour_banquet_nourriture',
        25: 'jour_banquet_boisson',
        26: 'jour_banquet_bieres',
        27: 'jour_banquet_mineraux',
        28: 'jour_banquet_vins',
        35: 'jour_tabagie',
    }

    # HP deductions are NEGATIVE adjustments subtracted from jour fields.
    # 1) Apply deductions to jour_* via column mapping
    # 2) Store entries in hp_admin_entries for display
    sections = set()

    for col_idx_key, nas_field in hp_col_to_nas.items():
        val = jour_deductions.get(str(col_idx_key)) or jour_deductions.get(col_idx_key)
        if val is not None and float(val) != 0:
            existing = getattr(nas, nas_field, None) or 0
            # HP deductions are subtracted (stored as negative in jour)
            new_val = round(float(existing) - abs(float(val)), 2)
            _safe_set(nas, nas_field, new_val, count)
            sections.add('jour')

    # Build hp_admin_entries for display (with negative values)
    hp_entries = []
    day_departments = data.get(f'day_{day}_departments', {})
    for area, total in day_departments.items():
        if total and float(total) != 0:
            hp_entries.append({
                'area': area,
                'nourriture': -abs(float(total)),
                'boisson': 0, 'biere': 0, 'vin': 0, 'mineraux': 0,
                'autre': 0, 'pourboire': 0,
                'raison': 'HP Import', 'autorise_par': ''
            })
            count[0] += 1

    if hp_entries:
        nas.set_json('hp_admin_entries', hp_entries)
        sections.add('hpadmin')

    return {'count': count[0], 'sections': list(sections),
            'details': {'doc_type': 'hp_excel', 'day': day}}


# ─── SALES JOURNAL ─── (Jour F&B + Recap + Transelect)
def _fill_from_sales_journal(data, nas, day):
    count = [0]
    sections = []

    # --- Departments → Jour F&B ---
    depts = data.get('departments', {})

    dept_mapping = {
        'piazza': {
            'nourriture': 'jour_piazza_nourriture',
            'boisson': 'jour_piazza_boisson',
            'bieres': 'jour_piazza_bieres',
            'mineraux': 'jour_piazza_mineraux',
            'vins': 'jour_piazza_vins',
        },
        'banquet': {
            'nourriture': 'jour_banquet_nourriture',
            'boisson': 'jour_banquet_boisson',
            'bieres': 'jour_banquet_bieres',
            'mineraux': 'jour_banquet_mineraux',
            'vins': 'jour_banquet_vins',
        },
        'chambres': {
            'nourriture': 'jour_chambres_svc_nourriture',
            'boisson': 'jour_chambres_svc_boisson',
            'bieres': 'jour_chambres_svc_bieres',
            'mineraux': 'jour_chambres_svc_mineraux',
            'vins': 'jour_chambres_svc_vins',
        },
        'spesa': {
            'nourriture': 'jour_spesa_nourriture',
            'boisson': 'jour_spesa_boisson',
            'bieres': 'jour_spesa_bieres',
            'mineraux': 'jour_spesa_mineraux',
            'vins': 'jour_spesa_vins',
        },
        'cafe': {
            'nourriture': 'jour_cafe_nourriture',
            'boisson': 'jour_cafe_boisson',
            'bieres': 'jour_cafe_bieres',
            'mineraux': 'jour_cafe_mineraux',
            'vins': 'jour_cafe_vins',
        },
    }

    for dept_name, field_map in dept_mapping.items():
        dept_data = depts.get(dept_name, {})
        for src_key, nas_field in field_map.items():
            _safe_set(nas, nas_field, dept_data.get(src_key), count)

    # Tabagie — can be in its own department OR inside spesa.tabagie
    tabagie_val = depts.get('tabagie', {}).get('total')
    if tabagie_val is None or tabagie_val == 0:
        tabagie_val = depts.get('spesa', {}).get('tabagie')
    _safe_set(nas, 'jour_tabagie', tabagie_val, count)

    # Pourboires
    _safe_set(nas, 'jour_pourboires', data.get('adjustments', {}).get('pourboire_charge'), count)

    if count[0] > 0:
        sections.append('jour')

    # --- Taxes → Jour (ADDITIVE: Sales Journal F&B taxes ADD to Daily Rev chambres taxes) ---
    taxes = data.get('taxes', {})
    sj_tps = taxes.get('tps')
    sj_tvq = taxes.get('tvq')
    if sj_tps is not None:
        existing_tps = getattr(nas, 'jour_tps', None) or 0
        _safe_set(nas, 'jour_tps', float(existing_tps) + float(sj_tps), count)
    if sj_tvq is not None:
        existing_tvq = getattr(nas, 'jour_tvq', None) or 0
        _safe_set(nas, 'jour_tvq', float(existing_tvq) + float(sj_tvq), count)

    # --- Adjustments ---
    adjustments = data.get('adjustments', {})

    # POURBOIRE CHARGE → Remboursement Gratuité (negated: debit → negative in Recap)
    pourboire_charge = adjustments.get('pourboire_charge')
    if pourboire_charge is not None:
        _safe_set(nas, 'remb_gratuite_lecture', -abs(float(pourboire_charge)), count)
        sections.append('recap')

    # FORFAIT → stored in jour_forfait_sj (negative value from Sales Journal)
    # jour_diff_forfait is AUTO-CALCULATED in calculate_all(): = jour_forfait_sj + g4_montant
    forfait_val = adjustments.get('forfait')
    if forfait_val is not None and float(forfait_val) != 0:
        _safe_set(nas, 'jour_forfait_sj', -abs(float(forfait_val)), count)
        sections.append('jour')

    # --- Payments ---
    payments = data.get('payments', {})

    # COMPTANT → Recap "Comptant Positouch"
    comptant_val = payments.get('comptant')
    if comptant_val is not None and float(comptant_val) != 0:
        _safe_set(nas, 'cash_pos_lecture', float(comptant_val), count)
        sections.append('recap')

    # --- Payments → Transelect restaurant ---
    rest = nas.get_json('transelect_restaurant')
    # Sales Journal gives POS terminal totals
    for card_src, card_dst in [('visa', 'visa'), ('mastercard', 'mc'),
                                ('amex', 'amex'), ('interac', 'debit')]:
        val = payments.get(card_src)
        if val is not None:
            # Add to a generic POS terminal entry
            if 'POS' not in rest:
                rest['POS'] = {}
            rest['POS'][card_dst] = float(val)
            count[0] += 1
    if 'POS' in rest:
        nas.set_json('transelect_restaurant', rest)
        sections.append('transelect')

    return {'count': count[0], 'sections': list(set(sections)),
            'details': {'doc_type': 'sales_journal'}}


# ─── SD DEPOSIT ─── (SetD personnel)
def _fill_from_sd_deposit(data, nas, day):
    count = [0]
    sections = set()

    # SD parser gives us setd_fillable: {col_letter: variance_amount}
    fillable = data.get('setd_fillable', {})
    entries = data.get('entries', [])
    totals = data.get('totals', {})

    # --- Totals → Recap (deposit_cdn) ---
    # SD "Montant Vérifié" total = Dépôt Canadien in Recap
    total_verifie = totals.get('total_verifie')
    if total_verifie is not None and float(total_verifie) != 0:
        _safe_set(nas, 'deposit_cdn', abs(float(total_verifie)), count)
        sections.add('recap')

    # SD "Total Variance" = Surplus/Déficit (informational, already calculated by Recap)
    # We store it for cross-validation but don't overwrite the calculated value
    total_variance = totals.get('total_variance')
    if total_variance is not None:
        # Store as metadata for validation in Sommaire
        existing_meta = nas.get_json('sd_validation') if hasattr(nas, 'get_json') else {}
        if not isinstance(existing_meta, dict):
            existing_meta = {}
        existing_meta['sd_total_variance'] = float(total_variance)
        existing_meta['sd_total_montant'] = float(totals.get('total_montant', 0))
        existing_meta['sd_total_verifie'] = float(total_verifie or 0)
        existing_meta['sd_total_remboursement'] = float(totals.get('total_remboursement', 0))
        # Only store if model supports it (graceful fallback)
        try:
            nas.set_json('sd_validation', existing_meta)
        except Exception:
            pass  # Model may not have this column yet

    # Build personnel list for NAS
    personnel = nas.get_json('setd_personnel')
    if not isinstance(personnel, list):
        personnel = []

    # Add from SD entries
    for entry in entries:
        nom = entry.get('nom', '')
        montant = entry.get('montant_verifie', entry.get('montant', 0))
        if nom and montant:
            personnel.append({
                'name': nom,
                'amount': float(montant),
                'column': '',  # Will be resolved by SetD mapper
                'department': entry.get('departement', ''),
            })
            count[0] += 1

    if personnel:
        nas.set_json('setd_personnel', personnel)
        sections.add('setd')

    # Also store SD entries for the SD tab
    sd_entries = []
    for entry in entries:
        sd_entries.append({
            'department': entry.get('departement', ''),
            'name': entry.get('nom', ''),
            'currency': entry.get('devise', 'CDN'),
            'amount': float(entry.get('montant', 0) or 0),
            'verified': float(entry.get('montant_verifie', 0) or 0),
            'reimbursement': float(entry.get('remboursement', 0) or 0),
        })
        count[0] += 1

    if sd_entries:
        nas.set_json('sd_entries', sd_entries)
        sections.add('sd')

    return {'count': count[0], 'sections': list(sections),
            'details': {'doc_type': 'sd_deposit', 'entries': len(entries),
                        'matched': data.get('matched_count', 0),
                        'unmatched': data.get('unmatched_count', 0),
                        'deposit_cdn': float(total_verifie) if total_verifie else 0}}


# ═══════════════════════════════════════
# API — SAVE SECTIONS
# ═══════════════════════════════════════

def _get_session(data):
    """Get or return error for a session by date."""
    audit_date_str = data.get('date')
    if not audit_date_str:
        return None, jsonify({'error': 'Date requise'}), 400
    try:
        d = datetime.strptime(audit_date_str, '%Y-%m-%d').date()
    except ValueError:
        return None, jsonify({'error': 'Format invalide'}), 400
    nas = NightAuditSession.query.filter_by(audit_date=d).first()
    if not nas:
        return None, jsonify({'error': 'Session non trouvée'}), 404
    if nas.status == 'locked':
        return None, jsonify({'error': 'Session verrouillée'}), 403
    return nas, None, None


@rj_native_bp.route('/api/rj/native/save/controle', methods=['POST'])
@auth_required
def save_controle():
    """Save controle (metadata) section."""
    data = request.get_json(force=True)
    result = _get_session(data)
    nas, err, code = result
    if err:
        return err, code

    nas.auditor_name = data.get('auditor_name', nas.auditor_name)
    nas.temperature = data.get('temperature', nas.temperature)
    nas.weather_condition = data.get('weather_condition', nas.weather_condition)
    nas.chambres_refaire = int(data.get('chambres_refaire', nas.chambres_refaire or 0))
    nas.notes = data.get('notes', nas.notes)
    # Auto-calculate jours_dans_mois from audit_date
    if nas.audit_date:
        import calendar as cal_mod
        _, days = cal_mod.monthrange(nas.audit_date.year, nas.audit_date.month)
        nas.jours_dans_mois = days
    if nas.status == 'draft':
        nas.status = 'in_progress'

    db.session.commit()
    return jsonify({'success': True, 'section': 'controle'})


@rj_native_bp.route('/api/rj/native/save/recap', methods=['POST'])
@auth_required
def save_recap():
    """Save recap (cash reconciliation) section."""
    data = request.get_json(force=True)
    result = _get_session(data)
    nas, err, code = result
    if err:
        return err, code

    float_fields = [
        'cash_ls_lecture', 'cash_ls_corr', 'cash_pos_lecture', 'cash_pos_corr',
        'cheque_ar_lecture', 'cheque_ar_corr', 'cheque_dr_lecture', 'cheque_dr_corr',
        'remb_gratuite_lecture', 'remb_gratuite_corr', 'remb_client_lecture', 'remb_client_corr',
        'dueback_reception_lecture', 'dueback_reception_corr',
        'dueback_nb_lecture', 'dueback_nb_corr',
        'deposit_cdn', 'deposit_us'
    ]
    for f in float_fields:
        if f in data:
            try:
                setattr(nas, f, float(data[f] or 0))
            except (ValueError, TypeError):
                pass

    nas.calculate_all()
    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()

    return jsonify({
        'success': True,
        'section': 'recap',
        'recap_balance': nas.recap_balance,
        'is_recap_balanced': nas.is_recap_balanced
    })


@rj_native_bp.route('/api/rj/native/save/transelect', methods=['POST'])
@auth_required
def save_transelect():
    """Save transelect (credit card) section."""
    data = request.get_json(force=True)
    result = _get_session(data)
    nas, err, code = result
    if err:
        return err, code

    if 'restaurant' in data:
        nas.set_json('transelect_restaurant', data['restaurant'])
    if 'reception' in data:
        nas.set_json('transelect_reception', data['reception'])

    nas.calculate_all()
    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()

    return jsonify({
        'success': True,
        'section': 'transelect',
        'transelect_variance': nas.transelect_variance,
        'is_transelect_balanced': nas.is_transelect_balanced,
        'quasimodo': nas.get_json('transelect_quasimodo')
    })


@rj_native_bp.route('/api/rj/native/save/geac', methods=['POST'])
@auth_required
def save_geac():
    """Save GEAC/UX section."""
    data = request.get_json(force=True)
    result = _get_session(data)
    nas, err, code = result
    if err:
        return err, code

    if 'cashout' in data:
        nas.set_json('geac_cashout', data['cashout'])
    if 'daily_rev' in data:
        nas.set_json('geac_daily_rev', data['daily_rev'])
    if 'balance_sheet' in data:
        nas.set_json('geac_balance_sheet', data['balance_sheet'])

    for f in ['geac_ar_previous', 'geac_ar_charges', 'geac_ar_payments', 'geac_ar_new_balance']:
        if f in data:
            try:
                setattr(nas, f, float(data[f] or 0))
            except (ValueError, TypeError):
                pass

    nas.calculate_all()
    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()

    return jsonify({
        'success': True,
        'section': 'geac',
        'geac_ar_variance': nas.geac_ar_variance,
        'is_ar_balanced': nas.is_ar_balanced
    })


@rj_native_bp.route('/api/rj/native/save/dueback', methods=['POST'])
@auth_required
def save_dueback():
    """Save DueBack section."""
    data = request.get_json(force=True)
    result = _get_session(data)
    nas, err, code = result
    if err:
        return err, code

    entries = data.get('entries', [])
    clean = []
    for e in entries:
        if not e.get('name', '').strip():
            continue
        clean.append({
            'name': e['name'].strip(),
            'previous': float(e.get('previous', 0) or 0),
            'nouveau': float(e.get('nouveau', 0) or 0)
        })
    nas.set_json('dueback_entries', clean)
    nas.calculate_all()

    # Also update recap dueback_reception_lecture from total
    nas.dueback_reception_lecture = nas.dueback_total

    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()

    return jsonify({
        'success': True,
        'section': 'dueback',
        'dueback_total': nas.dueback_total,
        'entry_count': len(clean)
    })


# ═══════════════════════════════════════
# API — SAVE: SD, DEPOT, SETD, JOUR
# ═══════════════════════════════════════

@rj_native_bp.route('/api/rj/native/save/sd', methods=['POST'])
@auth_required
def save_sd():
    """Save SD (Sommaire Journalier des Dépôts) section."""
    data = request.get_json(force=True)
    result = _get_session(data)
    nas, err, code = result
    if err:
        return err, code

    entries = data.get('entries', [])
    clean = []
    for e in entries:
        if not e.get('name', '').strip():
            continue
        clean.append({
            'department': e.get('department', '').strip(),
            'name': e['name'].strip(),
            'currency': e.get('currency', 'CDN'),
            'amount': float(e.get('amount', 0) or 0),
            'verified': float(e.get('verified', 0) or 0),
            'reimbursement': float(e.get('reimbursement', 0) or 0),
        })
    nas.set_json('sd_entries', clean)

    # ★ Propagate SD data to dependent sections (same logic as upload)
    sections_updated = ['sd']

    # 1) SetD personnel — SD verified amounts → SetD list
    personnel = []
    for entry in clean:
        if entry['verified'] and float(entry['verified']) != 0:
            personnel.append({
                'name': entry['name'],
                'amount': float(entry['verified']),
                'department': entry.get('department', ''),
            })
    if personnel:
        nas.set_json('setd_personnel', personnel)
        sections_updated.append('setd')

    # 2) Depot — SD total verified → Client 6 first row + Recap deposit_cdn
    total_verified = sum(float(e.get('verified', 0) or 0) for e in clean)
    if total_verified != 0:
        abs_total = abs(total_verified)
        # Update depot_data with the SD total in Client 6
        depot = nas.get_json('depot_data')
        if not isinstance(depot, dict):
            depot = {}
        if 'client6' not in depot:
            depot['client6'] = {'date': '', 'amounts': []}
        depot['client6']['amounts'] = [abs_total]
        nas.set_json('depot_data', depot)
        sections_updated.append('depot')
        # Also directly update Recap deposit_cdn
        nas.deposit_cdn = abs_total
        sections_updated.append('recap')

    # ★ Write-back to SD Excel file (if uploaded)
    try:
        from routes.audit.rj_core import SD_FILES, get_session_id
        from utils.sd_writer import SDWriter
        session_id = get_session_id()
        if session_id in SD_FILES:
            day = nas.audit_date.day if nas.audit_date else None
            if day:
                sd_entries_for_excel = []
                for entry in clean:
                    variance = round(float(entry.get('amount', 0) or 0)
                                     - float(entry.get('verified', 0) or 0)
                                     - float(entry.get('reimbursement', 0) or 0), 2)
                    sd_entries_for_excel.append({
                        'departement': entry.get('department', ''),
                        'nom': entry.get('name', ''),
                        'cdn_us': entry.get('currency', 'CDN'),
                        'montant': float(entry.get('amount', 0) or 0),
                        'montant_verifie': float(entry.get('verified', 0) or 0),
                        'remboursement': float(entry.get('reimbursement', 0) or 0),
                        'variance': variance,
                    })
                SD_FILES[session_id] = SDWriter.write_entries(
                    SD_FILES[session_id], day, sd_entries_for_excel)
    except Exception as e:
        logger.warning(f"SD write-back failed (non-critical): {e}")

    nas.calculate_all()

    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()

    return jsonify({
        'success': True,
        'section': 'sd',
        'sections_updated': sections_updated,
        'sd_total_verified': nas.sd_total_verified,
        'entry_count': len(clean),
        'session': nas.to_dict()
    })


@rj_native_bp.route('/api/rj/native/save/depot', methods=['POST'])
@auth_required
def save_depot():
    """Save Depot (Client 6 + Client 8) section."""
    data = request.get_json(force=True)
    result = _get_session(data)
    nas, err, code = result
    if err:
        return err, code

    depot = {}
    for client in ['client6', 'client8']:
        cdata = data.get(client, {})
        depot[client] = {
            'date': cdata.get('date', ''),
            'amounts': [float(a or 0) for a in cdata.get('amounts', []) if a is not None]
        }
    nas.set_json('depot_data', depot)
    nas.calculate_all()

    # Auto-update Recap deposit_cdn from depot total
    nas.deposit_cdn = nas.depot_total

    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()

    return jsonify({
        'success': True,
        'section': 'depot',
        'depot_total': nas.depot_total
    })


@rj_native_bp.route('/api/rj/native/save/setd', methods=['POST'])
@auth_required
def save_setd():
    """Save SetD (sommaire mensuel) section."""
    data = request.get_json(force=True)
    result = _get_session(data)
    nas, err, code = result
    if err:
        return err, code

    personnel = data.get('personnel', [])
    clean = []
    for p in personnel:
        if not p.get('name', '').strip():
            continue
        clean.append({
            'name': p['name'].strip(),
            'column_letter': p.get('column_letter', ''),
            'amount': float(p.get('amount', 0) or 0),
        })
    nas.set_json('setd_personnel', clean)
    nas.calculate_all()

    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()

    return jsonify({
        'success': True,
        'section': 'setd',
        'setd_rj_balance': nas.setd_rj_balance,
        'personnel_count': len(clean)
    })


@rj_native_bp.route('/api/rj/native/save/jour', methods=['POST'])
@auth_required
def save_jour():
    """Save Jour (daily revenue) section — ~55 fields."""
    data = request.get_json(force=True)
    result = _get_session(data)
    nas, err, code = result
    if err:
        return err, code

    # Float fields
    float_fields = [
        'jour_cafe_nourriture', 'jour_cafe_boisson', 'jour_cafe_bieres', 'jour_cafe_mineraux', 'jour_cafe_vins',
        'jour_piazza_nourriture', 'jour_piazza_boisson', 'jour_piazza_bieres', 'jour_piazza_mineraux', 'jour_piazza_vins',
        'jour_spesa_nourriture', 'jour_spesa_boisson', 'jour_spesa_bieres', 'jour_spesa_mineraux', 'jour_spesa_vins',
        'jour_chambres_svc_nourriture', 'jour_chambres_svc_boisson', 'jour_chambres_svc_bieres', 'jour_chambres_svc_mineraux', 'jour_chambres_svc_vins',
        'jour_banquet_nourriture', 'jour_banquet_boisson', 'jour_banquet_bieres', 'jour_banquet_mineraux', 'jour_banquet_vins',
        'jour_pourboires', 'jour_tabagie', 'jour_location_salle',
        'jour_adj_cafe', 'jour_adj_piazza', 'jour_adj_spesa',
        'jour_adj_chambres_svc', 'jour_adj_banquet', 'jour_adj_tabagie',
        'jour_room_revenue', 'jour_tel_local', 'jour_tel_interurbain', 'jour_tel_publics',
        'jour_nettoyeur', 'jour_machine_distrib', 'jour_autres_gl', 'jour_sonifi',
        'jour_lit_pliant', 'jour_boutique', 'jour_internet', 'jour_massage',
        'jour_diff_forfait', 'jour_forfait_sj', 'g4_montant',
        'jour_tvq', 'jour_tps', 'jour_taxe_hebergement',
        'jour_gift_cards', 'jour_certificats',
        'jour_club_lounge', 'jour_deposit_on_hand', 'jour_ar_misc',
    ]
    for f in float_fields:
        if f in data:
            try:
                setattr(nas, f, float(data[f] or 0))
            except (ValueError, TypeError):
                pass

    # Integer fields
    int_fields = [
        'jour_rooms_simple', 'jour_rooms_double', 'jour_rooms_suite', 'jour_rooms_comp',
        'jour_nb_clients', 'jour_rooms_hors_usage',
    ]
    for f in int_fields:
        if f in data:
            try:
                setattr(nas, f, int(data[f] or 0))
            except (ValueError, TypeError):
                pass

    # JSON: adjustment notes
    if 'jour_adj_notes' in data:
        nas.set_json('jour_adj_notes', data['jour_adj_notes'])

    nas.calculate_all()

    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()

    return jsonify({
        'success': True,
        'section': 'jour',
        'jour_total_fb': nas.jour_total_fb,
        'jour_total_revenue': nas.jour_total_revenue,
        'jour_occupancy_rate': nas.jour_occupancy_rate,
        'jour_adr': nas.jour_adr,
        'jour_revpar': nas.jour_revpar,
    })



@rj_native_bp.route('/api/rj/native/save/hp_admin', methods=['POST'])
@auth_required
def save_hp_admin():
    """Save HP/Admin entries."""
    data = request.get_json(force=True)
    nas, err, code = _get_session(data)
    if err:
        return err, code

    entries = data.get('entries', [])
    nas.set_json('hp_admin_entries', entries)

    # ★ Write-back to HP Excel file (if uploaded)
    try:
        from routes.audit.rj_core import HP_FILES, get_session_id
        from utils.hp_writer import HPWriter
        session_id = get_session_id()
        if session_id in HP_FILES:
            day = nas.audit_date.day if nas.audit_date else None
            if day and entries:
                HP_FILES[session_id] = HPWriter.write_entries(
                    HP_FILES[session_id], day, entries)
    except Exception as e:
        logger.warning(f"HP write-back failed (non-critical): {e}")

    nas.calculate_all()
    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()
    return jsonify({'success': True, 'section': 'hp_admin', 'total': nas.hp_admin_total})


@rj_native_bp.route('/api/rj/native/save/internet', methods=['POST'])
@auth_required
def save_internet():
    """Save Internet section."""
    data = request.get_json(force=True)
    nas, err, code = _get_session(data)
    if err:
        return err, code

    nas.internet_ls_361 = float(data.get('internet_ls_361') or 0)
    nas.internet_ls_365 = float(data.get('internet_ls_365') or 0)
    nas.calculate_all()
    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()
    return jsonify({'success': True, 'section': 'internet', 'variance': nas.internet_variance})


@rj_native_bp.route('/api/rj/native/save/sonifi', methods=['POST'])
@auth_required
def save_sonifi():
    """Save Sonifi section."""
    data = request.get_json(force=True)
    nas, err, code = _get_session(data)
    if err:
        return err, code

    nas.sonifi_cd_352 = float(data.get('sonifi_cd_352') or 0)
    nas.sonifi_email = float(data.get('sonifi_email') or 0)
    nas.calculate_all()
    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()
    return jsonify({'success': True, 'section': 'sonifi', 'variance': nas.sonifi_variance})


@rj_native_bp.route('/api/rj/native/save/quasimodo', methods=['POST'])
@auth_required
def save_quasimodo():
    """Save Quasimodo reconciliation."""
    data = request.get_json(force=True)
    nas, err, code = _get_session(data)
    if err:
        return err, code

    for ct in ['debit', 'visa', 'mc', 'amex', 'discover']:
        setattr(nas, f'quasi_fb_{ct}', float(data.get(f'quasi_fb_{ct}') or 0))
        setattr(nas, f'quasi_rec_{ct}', float(data.get(f'quasi_rec_{ct}') or 0))
    nas.quasi_amex_factor = float(data.get('quasi_amex_factor') or 0.9735)
    nas.quasi_cash_cdn = float(data.get('quasi_cash_cdn') or 0)
    nas.quasi_cash_usd = float(data.get('quasi_cash_usd') or 0)
    nas.calculate_all()
    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()
    return jsonify({'success': True, 'section': 'quasimodo',
                    'total': nas.quasi_total, 'variance': nas.quasi_variance})


@rj_native_bp.route('/api/rj/native/save/dbrs', methods=['POST'])
@auth_required
def save_dbrs():
    """Save DBRS section."""
    data = request.get_json(force=True)
    nas, err, code = _get_session(data)
    if err:
        return err, code

    if 'market_segments' in data:
        nas.set_json('dbrs_market_segments', data['market_segments'])
    if 'otb_data' in data:
        nas.set_json('dbrs_otb_data', data['otb_data'])
    nas.dbrs_noshow_count = int(data.get('dbrs_noshow_count') or 0)
    nas.dbrs_noshow_revenue = float(data.get('dbrs_noshow_revenue') or 0)
    nas.calculate_all()
    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()
    return jsonify({'success': True, 'section': 'dbrs',
                    'adr': nas.dbrs_adr, 'daily_rev': nas.dbrs_daily_rev_today})

# ═══════════════════════════════════════
# API — SAVE — NEW SPECIALIZED SHEETS
# ═══════════════════════════════════════

@rj_native_bp.route('/api/rj/native/save/analyse_gl_101100', methods=['POST'])
@auth_required
def save_analyse_gl_101100():
    """Save Analyse GL 101100 (suspense account)."""
    data = request.get_json(force=True)
    nas, err, code = _get_session(data)
    if err:
        return err, code
    nas.gl_101100_previous = float(data.get('previous') or 0)
    nas.gl_101100_additions = float(data.get('additions') or 0)
    nas.gl_101100_deductions = float(data.get('deductions') or 0)
    nas.gl_101100_new_balance = float(data.get('new_balance') or 0)
    nas.gl_101100_notes = data.get('notes', '')
    nas.calculate_all()
    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()
    return jsonify({'success': True, 'section': 'analyse_gl_101100',
                    'variance': nas.gl_101100_variance})


@rj_native_bp.route('/api/rj/native/save/analyse_gl_100401', methods=['POST'])
@auth_required
def save_analyse_gl_100401():
    """Save Analyse GL 100401 (cash/bank account)."""
    data = request.get_json(force=True)
    nas, err, code = _get_session(data)
    if err:
        return err, code
    nas.gl_100401_previous = float(data.get('previous') or 0)
    nas.gl_100401_additions = float(data.get('additions') or 0)
    nas.gl_100401_deductions = float(data.get('deductions') or 0)
    nas.gl_100401_new_balance = float(data.get('new_balance') or 0)
    nas.gl_100401_notes = data.get('notes', '')
    nas.calculate_all()
    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()
    return jsonify({'success': True, 'section': 'analyse_gl_100401',
                    'variance': nas.gl_100401_variance})


@rj_native_bp.route('/api/rj/native/save/diff_caisse', methods=['POST'])
@auth_required
def save_diff_caisse():
    """Save Diff.Caisse (cash register variances)."""
    data = request.get_json(force=True)
    nas, err, code = _get_session(data)
    if err:
        return err, code
    if 'entries' in data:
        nas.set_json('diff_caisse_entries', data['entries'])
    nas.calculate_all()
    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()
    return jsonify({'success': True, 'section': 'diff_caisse',
                    'total': nas.diff_caisse_total, 'reconciled': nas.diff_caisse_reconciled})


@rj_native_bp.route('/api/rj/native/save/socan', methods=['POST'])
@auth_required
def save_socan():
    """Save SOCAN (music royalties)."""
    data = request.get_json(force=True)
    nas, err, code = _get_session(data)
    if err:
        return err, code
    nas.socan_allocation_resto = float(data.get('allocation_resto') or 0)
    nas.socan_allocation_bar = float(data.get('allocation_bar') or 0)
    nas.socan_allocation_banquet = float(data.get('allocation_banquet') or 0)
    nas.socan_notes = data.get('notes', '')
    nas.calculate_all()
    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()
    return jsonify({'success': True, 'section': 'socan', 'charge': nas.socan_charge})


@rj_native_bp.route('/api/rj/native/save/resonne', methods=['POST'])
@auth_required
def save_resonne():
    """Save Résonne (conference/AV system charges)."""
    data = request.get_json(force=True)
    nas, err, code = _get_session(data)
    if err:
        return err, code
    if 'entries' in data:
        nas.set_json('resonne_entries', data['entries'])
    nas.calculate_all()
    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()
    return jsonify({'success': True, 'section': 'resonne', 'total': nas.resonne_total})


@rj_native_bp.route('/api/rj/native/save/vestiaire', methods=['POST'])
@auth_required
def save_vestiaire():
    """Save Vestiaire (coat check revenue)."""
    data = request.get_json(force=True)
    nas, err, code = _get_session(data)
    if err:
        return err, code
    if 'entries' in data:
        nas.set_json('vestiaire_entries', data['entries'])
    nas.calculate_all()
    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()
    return jsonify({'success': True, 'section': 'vestiaire',
                    'total_revenue': nas.vestiaire_total_revenue,
                    'total_variance': nas.vestiaire_total_variance})


@rj_native_bp.route('/api/rj/native/save/admin', methods=['POST'])
@auth_required
def save_admin():
    """Save AD — Administration charges."""
    data = request.get_json(force=True)
    nas, err, code = _get_session(data)
    if err:
        return err, code
    if 'entries' in data:
        nas.set_json('admin_entries', data['entries'])
    nas.calculate_all()
    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()
    return jsonify({'success': True, 'section': 'admin', 'total': nas.admin_total})


@rj_native_bp.route('/api/rj/native/save/massage', methods=['POST'])
@auth_required
def save_massage():
    """Save Massage (detailed spa breakdown)."""
    data = request.get_json(force=True)
    nas, err, code = _get_session(data)
    if err:
        return err, code
    if 'entries' in data:
        nas.set_json('massage_entries', data['entries'])
    nas.calculate_all()
    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()
    return jsonify({'success': True, 'section': 'massage',
                    'total_revenue': nas.massage_total_revenue,
                    'total_tips': nas.massage_total_tips})


@rj_native_bp.route('/api/rj/native/save/ristourne', methods=['POST'])
@auth_required
def save_ristourne():
    """Save Ristourne (rebates/discounts)."""
    data = request.get_json(force=True)
    nas, err, code = _get_session(data)
    if err:
        return err, code
    if 'entries' in data:
        nas.set_json('ristourne_entries', data['entries'])
    nas.ristourne_analysis_notes = data.get('analysis_notes', '')
    nas.calculate_all()
    if nas.status == 'draft':
        nas.status = 'in_progress'
    db.session.commit()
    return jsonify({'success': True, 'section': 'ristourne',
                    'total': nas.ristourne_total,
                    'by_dept': nas.get_json('ristourne_by_dept')})


# ═══════════════════════════════════════
# API — CALCULATE & VALIDATE
# ═══════════════════════════════════════

@rj_native_bp.route('/api/rj/native/calculate', methods=['POST'])
@auth_required
def calculate():
    """Recalculate all balances and return results."""
    data = request.get_json(force=True)
    result = _get_session(data)
    nas, err, code = result
    if err:
        return err, code

    nas.calculate_all()
    db.session.commit()

    messages = []
    if not nas.is_recap_balanced:
        messages.append({
            'type': 'error',
            'section': 'recap',
            'message': f'Recap non balancé: ${nas.recap_balance:+.2f}'
        })
    if not nas.is_transelect_balanced:
        messages.append({
            'type': 'warning' if abs(nas.transelect_variance) < 5 else 'error',
            'section': 'transelect',
            'message': f'Variance Transelect: ${nas.transelect_variance:+.2f}'
        })
    if not nas.is_ar_balanced:
        messages.append({
            'type': 'error',
            'section': 'geac',
            'message': f'Variance AR: ${nas.geac_ar_variance:+.2f}'
        })
    if nas.is_fully_balanced:
        messages.append({
            'type': 'success',
            'section': 'all',
            'message': 'Toutes les sections sont balancées!'
        })

    return jsonify({
        'success': True,
        'recap_balance': nas.recap_balance,
        'transelect_variance': nas.transelect_variance,
        'geac_ar_variance': nas.geac_ar_variance,
        'is_recap_balanced': nas.is_recap_balanced,
        'is_transelect_balanced': nas.is_transelect_balanced,
        'is_ar_balanced': nas.is_ar_balanced,
        'is_fully_balanced': nas.is_fully_balanced,
        'dueback_total': nas.dueback_total,
        'quasimodo': nas.get_json('transelect_quasimodo'),
        'messages': messages
    })


# ═══════════════════════════════════════
# API — SUBMIT (finalize & lock)
# ═══════════════════════════════════════

def sync_to_dashboard(nas, d):
    """
    Sync NightAuditSession data to DailyReconciliation + DueBack + DailyJourMetrics.
    Called by submit_session() and the manual re-sync endpoint.
    """
    # Write DailyReconciliation snapshot
    recon = DailyReconciliation.query.filter_by(audit_date=d).first()
    if not recon:
        recon = DailyReconciliation(audit_date=d)
        db.session.add(recon)

    recon.auditor_name = nas.auditor_name
    recon.cash_lightspeed = (nas.cash_ls_lecture or 0) + (nas.cash_ls_corr or 0)
    recon.cash_positouch = (nas.cash_pos_lecture or 0) + (nas.cash_pos_corr or 0)
    recon.cheque_ar = (nas.cheque_ar_lecture or 0) + (nas.cheque_ar_corr or 0)
    recon.cheque_daily_rev = (nas.cheque_dr_lecture or 0) + (nas.cheque_dr_corr or 0)
    recon.remb_gratuite = (nas.remb_gratuite_lecture or 0) + (nas.remb_gratuite_corr or 0)
    recon.remb_client = (nas.remb_client_lecture or 0) + (nas.remb_client_corr or 0)
    recon.dueback_reception = (nas.dueback_reception_lecture or 0) + (nas.dueback_reception_corr or 0)
    recon.dueback_nb = (nas.dueback_nb_lecture or 0) + (nas.dueback_nb_corr or 0)
    recon.surplus_deficit = nas.recap_balance
    recon.deposit_cdn = nas.deposit_cdn or 0
    recon.deposit_us = nas.deposit_us or 0

    # Card totals from quasimodo
    quasi = nas.get_json('transelect_quasimodo')
    recon.card_debit_terminal = quasi.get('debit', 0)
    recon.card_visa_terminal = quasi.get('visa', 0)
    recon.card_mc_terminal = quasi.get('mc', 0)
    recon.card_amex_terminal = quasi.get('amex', 0)

    # Reception: keyed by card type → {fusebox, term8, k053, ...}
    recep = nas.get_json('transelect_reception')
    recon.card_visa_bank = recep.get('visa', {}).get('fusebox', 0)
    recon.card_mc_bank = recep.get('mc', {}).get('fusebox', 0)
    recon.card_amex_bank = recep.get('amex', {}).get('fusebox', 0)
    recon.card_debit_bank = recep.get('debit', {}).get('fusebox', 0)

    # AR
    recon.ar_previous = nas.geac_ar_previous or 0
    recon.ar_charges = nas.geac_ar_charges or 0
    recon.ar_payments = nas.geac_ar_payments or 0
    recon.ar_new_balance = nas.geac_ar_new_balance or 0

    recon.calculate_variances()

    # Write DueBack entries
    entries = nas.get_json('dueback_entries')
    if isinstance(entries, list):
        DueBack.query.filter_by(audit_date=d).delete()
        for e in entries:
            if e.get('name'):
                db.session.add(DueBack(
                    audit_date=d,
                    receptionist_name=e['name'],
                    balance=e.get('nouveau', 0),
                    entry_count=1
                ))

    # ── Auto-run macros: envoie_dans_jour + calcul_carte ──
    djm = DailyJourMetrics.query.filter_by(date=d).first()
    if not djm:
        djm = DailyJourMetrics(
            date=d, year=d.year, month=d.month, day_of_month=d.day,
            source='rj_native'
        )
        db.session.add(djm)

    # envoie_dans_jour: Recap → Jour
    djm.opening_balance = (nas.cash_ls_lecture or 0) + (nas.cash_ls_corr or 0)
    djm.cash_difference = nas.recap_balance or 0
    djm.closing_balance = (nas.deposit_cdn or 0) + (nas.deposit_us or 0)
    djm.rooms_ch_refaire = nas.chambres_refaire or 0

    # calcul_carte: Transelect → Jour
    quasi = nas.get_json('transelect_quasimodo')
    djm.debit_total = quasi.get('debit', 0)
    djm.visa_total = quasi.get('visa', 0)
    djm.mastercard_total = quasi.get('mc', 0)
    djm.amex_elavon_total = quasi.get('amex', 0)
    djm.discover_total = quasi.get('discover', 0)
    djm.total_cards = round(sum(quasi.values()), 2)

    # Jour F&B revenue (from native form)
    djm.cafe_link_total = sum(getattr(nas, f'jour_cafe_{c}', 0) or 0 for c in ['nourriture','boisson','bieres','mineraux','vins'])
    djm.piazza_total = sum(getattr(nas, f'jour_piazza_{c}', 0) or 0 for c in ['nourriture','boisson','bieres','mineraux','vins'])
    djm.spesa_total = sum(getattr(nas, f'jour_spesa_{c}', 0) or 0 for c in ['nourriture','boisson','bieres','mineraux','vins'])
    djm.room_svc_total = sum(getattr(nas, f'jour_chambres_svc_{c}', 0) or 0 for c in ['nourriture','boisson','bieres','mineraux','vins'])
    djm.banquet_total = sum(getattr(nas, f'jour_banquet_{c}', 0) or 0 for c in ['nourriture','boisson','bieres','mineraux','vins'])
    djm.tips_total = nas.jour_pourboires or 0
    djm.tabagie_total = nas.jour_tabagie or 0
    djm.fb_revenue = nas.jour_total_fb or 0

    # F&B by category (cross-department totals)
    djm.total_nourriture = sum(getattr(nas, f'jour_{d}_nourriture', 0) or 0 for d in ['cafe','piazza','spesa','chambres_svc','banquet'])
    djm.total_boisson = sum(getattr(nas, f'jour_{d}_boisson', 0) or 0 for d in ['cafe','piazza','spesa','chambres_svc','banquet'])
    djm.total_bieres = sum(getattr(nas, f'jour_{d}_bieres', 0) or 0 for d in ['cafe','piazza','spesa','chambres_svc','banquet'])
    djm.total_mineraux = sum(getattr(nas, f'jour_{d}_mineraux', 0) or 0 for d in ['cafe','piazza','spesa','chambres_svc','banquet'])
    djm.total_vins = sum(getattr(nas, f'jour_{d}_vins', 0) or 0 for d in ['cafe','piazza','spesa','chambres_svc','banquet'])

    # Hébergement + Autres
    djm.room_revenue = nas.jour_room_revenue or 0
    djm.other_revenue = sum(getattr(nas, f'jour_{f}', 0) or 0 for f in
        ['nettoyeur','machine_distrib','autres_gl','sonifi','lit_pliant','boutique','internet','massage',
         'tel_local','tel_interurbain','tel_publics'])
    djm.total_revenue = nas.jour_total_revenue or 0

    # Taxes
    djm.tvq_total = nas.jour_tvq or 0
    djm.tps_total = nas.jour_tps or 0
    djm.tvh_total = nas.jour_taxe_hebergement or 0

    # Occupation
    djm.rooms_simple = nas.jour_rooms_simple or 0
    djm.rooms_double = nas.jour_rooms_double or 0
    djm.rooms_suite = nas.jour_rooms_suite or 0
    djm.rooms_comp = nas.jour_rooms_comp or 0
    djm.total_rooms_sold = djm.rooms_simple + djm.rooms_double + djm.rooms_suite + djm.rooms_comp
    djm.nb_clients = nas.jour_nb_clients or 0
    djm.rooms_hors_usage = nas.jour_rooms_hors_usage or 0
    djm.occupancy_rate = nas.jour_occupancy_rate or 0
    djm.adr = nas.jour_adr or 0
    djm.revpar = nas.jour_revpar or 0

    djm.source = 'rj_native'


# ═══════════════════════════════════════
# API — SUBMIT (finalize & lock)
# ═══════════════════════════════════════

@rj_native_bp.route('/api/rj/native/submit/<audit_date>', methods=['POST'])
@auth_required
def submit_session(audit_date):
    """Validate, write snapshots to DailyReconciliation + DueBack, and lock."""
    try:
        d = datetime.strptime(audit_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400

    nas = NightAuditSession.query.filter_by(audit_date=d).first()
    if not nas:
        return jsonify({'error': 'Session non trouvée'}), 404
    if nas.status == 'locked':
        return jsonify({'error': 'Déjà soumise et verrouillée'}), 403

    # Recalculate
    nas.calculate_all()

    # Sync to dashboard tables
    sync_to_dashboard(nas, d)

    # Lock session
    nas.status = 'locked'
    nas.completed_at = datetime.utcnow()
    db.session.commit()

    return jsonify({
        'success': True,
        'message': 'Session soumise et verrouillée (macros exécutées automatiquement)',
        'is_fully_balanced': nas.is_fully_balanced,
        'recap_balance': nas.recap_balance,
        'transelect_variance': nas.transelect_variance,
        'geac_ar_variance': nas.geac_ar_variance,
        'macros': ['envoie_jour', 'calcul_carte', 'sync_dueback']
    })


# ═══════════════════════════════════════
# API — MANUAL RE-SYNC
# ═══════════════════════════════════════

@rj_native_bp.route('/api/rj/native/sync/<audit_date>', methods=['POST'])
@auth_required
def manual_sync_dashboard(audit_date):
    """Manually re-sync an existing session to dashboard tables without locking."""
    try:
        d = datetime.strptime(audit_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400

    nas = NightAuditSession.query.filter_by(audit_date=d).first()
    if not nas:
        return jsonify({'error': 'Session non trouvée'}), 404

    try:
        # Recalculate first
        nas.calculate_all()

        # Sync to dashboard tables
        sync_to_dashboard(nas, d)

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Dashboard synchronisé avec succès',
            'recap_balance': nas.recap_balance,
            'transelect_variance': nas.transelect_variance,
            'geac_ar_variance': nas.geac_ar_variance
        })
    except Exception as e:
        logger.error(f"Manual sync error: {e}", exc_info=True)
        db.session.rollback()
        return jsonify({'error': f'Erreur de synchronisation: {str(e)}'}), 500


# ═══════════════════════════════════════
# API — MACROS (envoie_dans_jour, calcul_carte, sync_setd)
# ═══════════════════════════════════════

@rj_native_bp.route('/api/rj/native/macro/<macro_name>', methods=['POST'])
@auth_required
def run_macro(macro_name):
    """Execute a macro operation (native version — writes to DailyJourMetrics)."""
    data = request.get_json(force=True)
    audit_date_str = data.get('date')
    if not audit_date_str:
        return jsonify({'error': 'Date requise'}), 400
    try:
        d = datetime.strptime(audit_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400

    nas = NightAuditSession.query.filter_by(audit_date=d).first()
    if not nas:
        return jsonify({'error': 'Session non trouvée'}), 404

    # Recalculate first
    nas.calculate_all()

    # Get or create DailyJourMetrics for this date
    djm = DailyJourMetrics.query.filter_by(date=d).first()
    if not djm:
        djm = DailyJourMetrics(
            date=d, year=d.year, month=d.month, day_of_month=d.day,
            source='rj_native'
        )
        db.session.add(djm)

    try:
        if macro_name == 'envoie_jour':
            # Recap → Jour: cash reconciliation data into jour columns
            # Recap row H19:N19 = [cash_ls_net, cash_pos_net, cheque_ar_net, cheque_dr_net,
            #                      remb_gratuite_net, remb_client_net, dueback_net]
            djm.opening_balance = (nas.cash_ls_lecture or 0) + (nas.cash_ls_corr or 0)
            djm.cash_difference = nas.recap_balance or 0
            djm.closing_balance = (nas.deposit_cdn or 0) + (nas.deposit_us or 0)
            djm.rooms_ch_refaire = nas.chambres_refaire or 0
            djm.source = 'rj_native'
            db.session.commit()
            return jsonify({
                'success': True,
                'message': f'Recap envoyé dans Jour pour {d.isoformat()}'
            })

        elif macro_name == 'calcul_carte':
            # Transelect → Jour: card totals (Quasimodo) into jour payment columns
            quasi = nas.get_json('transelect_quasimodo')
            recep = nas.get_json('transelect_reception')

            # Card totals from Quasimodo (restaurant + reception terminal)
            djm.debit_total = quasi.get('debit', 0)
            djm.visa_total = quasi.get('visa', 0)
            djm.mastercard_total = quasi.get('mc', 0)
            djm.amex_elavon_total = quasi.get('amex', 0)
            djm.discover_total = quasi.get('discover', 0)
            djm.total_cards = round(sum(quasi.values()), 2)

            djm.source = 'rj_native'
            db.session.commit()
            return jsonify({
                'success': True,
                'message': f'Cartes calculées pour {d.isoformat()}',
                'quasimodo': quasi,
                'total_cards': djm.total_cards
            })

        elif macro_name == 'sync_setd':
            # DueBack → write individual DueBack entries for the date
            entries = nas.get_json('dueback_entries')
            if isinstance(entries, list):
                DueBack.query.filter_by(audit_date=d).delete()
                for e in entries:
                    if e.get('name'):
                        db.session.add(DueBack(
                            audit_date=d,
                            receptionist_name=e['name'],
                            balance=e.get('nouveau', 0),
                            entry_count=1
                        ))
                db.session.commit()
                return jsonify({
                    'success': True,
                    'message': f'DueBack synchronisé: {len(entries)} réceptionnistes',
                    'count': len(entries)
                })
            return jsonify({'success': True, 'message': 'Aucune entrée DueBack'})

        else:
            return jsonify({'error': f'Macro inconnue: {macro_name}'}), 400

    except Exception as e:
        logger.error(f"Macro {macro_name} error: {e}", exc_info=True)
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ═══════════════════════════════════════
# API — EXPORT
# ═══════════════════════════════════════

@rj_native_bp.route('/api/rj/native/export/night/<audit_date>')
@auth_required
def export_night(audit_date):
    """Export a complete night audit session as a professional Excel file.

    Generates an .xlsx file with 14 sheets matching the RJ Natif tabs:
    1. Contrôle, 2. DueBack, 3. Recap, 4. Transelect, 5. GEAC, 6. SD-Dépôt,
    7. SetD, 8. HP-Admin, 9. Internet, 10. Sonifi, 11. Jour, 12. Quasimodo,
    13. DBRS, 14. Sommaire
    """
    try:
        d = datetime.strptime(audit_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400

    nas = NightAuditSession.query.filter_by(audit_date=d).first()
    if not nas:
        return jsonify({'error': 'Session non trouvée'}), 404

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Formatting helpers
        header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        currency_fmt = '$#,##0.00'
        percent_fmt = '0.0%'
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        def apply_header_row(ws, row_num, headers):
            """Apply header formatting to a row."""
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

        def apply_currency_format(ws, row, col):
            """Apply currency format to a cell."""
            ws.cell(row=row, column=col).number_format = currency_fmt

        def format_variance_cell(ws, row, col, value):
            """Format a variance cell with green if balanced, red if not."""
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.number_format = currency_fmt
            if value is not None and abs(value) < 0.02:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif value is not None and abs(value) >= 0.02:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        def auto_width(ws, min_width=12):
            """Auto-size columns."""
            for column in ws.columns:
                max_length = min_width
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Sheet 1: Contrôle
        ws = wb.create_sheet('Contrôle')
        ws['A1'] = 'CONTRÔLE'
        ws['A1'].font = Font(bold=True, size=12)

        row = 3
        ws[f'A{row}'] = 'Date'
        ws[f'B{row}'] = d.isoformat()

        row += 1
        ws[f'A{row}'] = 'Auditeur'
        ws[f'B{row}'] = nas.auditor_name or ''

        row += 1
        ws[f'A{row}'] = 'Température'
        ws[f'B{row}'] = nas.temperature or ''

        row += 1
        ws[f'A{row}'] = 'Condition météo'
        ws[f'B{row}'] = nas.weather_condition or ''

        row += 1
        ws[f'A{row}'] = 'Chambres à refaire'
        ws[f'B{row}'] = nas.chambres_refaire or 0

        auto_width(ws)

        # Sheet 2: DueBack
        ws = wb.create_sheet('DueBack')
        ws['A1'] = 'DUEBACK'
        ws['A1'].font = Font(bold=True, size=12)

        apply_header_row(ws, 3, ['Réceptionniste', 'Solde précédent', 'Nouveau', 'Variation'])

        entries = nas.get_json('dueback_entries')
        row = 4
        if isinstance(entries, list):
            for e in entries:
                ws.cell(row=row, column=1, value=e.get('name', ''))
                ws.cell(row=row, column=2, value=e.get('previous', 0))
                apply_currency_format(ws, row, 2)
                ws.cell(row=row, column=3, value=e.get('nouveau', 0))
                apply_currency_format(ws, row, 3)
                ws.cell(row=row, column=4, value=f"=C{row}-B{row}")
                apply_currency_format(ws, row, 4)
                row += 1

        row += 1
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=3, value=nas.dueback_total or 0)
        ws.cell(row=row, column=3).font = Font(bold=True)
        apply_currency_format(ws, row, 3)

        auto_width(ws)

        # Sheet 3: Recap
        ws = wb.create_sheet('Recap')
        ws['A1'] = 'RECAP'
        ws['A1'].font = Font(bold=True, size=12)

        apply_header_row(ws, 3, ['Description', 'Lecture', 'Correction', 'Net'])

        recap_rows = [
            ('Comptant LightSpeed', nas.cash_ls_lecture, nas.cash_ls_corr),
            ('Comptant POSitouch', nas.cash_pos_lecture, nas.cash_pos_corr),
            ('Chèques A/R', nas.cheque_ar_lecture, nas.cheque_ar_corr),
            ('Chèques Daily Rev', nas.cheque_dr_lecture, nas.cheque_dr_corr),
            ('Remboursement gratuit', nas.remb_gratuite_lecture, nas.remb_gratuite_corr),
            ('Remboursement client', nas.remb_client_lecture, nas.remb_client_corr),
            ('DueBack Réception', nas.dueback_reception_lecture, nas.dueback_reception_corr),
            ('DueBack NB', nas.dueback_nb_lecture, nas.dueback_nb_corr),
        ]

        row = 4
        for label, lect, corr in recap_rows:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=lect or 0)
            apply_currency_format(ws, row, 2)
            ws.cell(row=row, column=3, value=corr or 0)
            apply_currency_format(ws, row, 3)
            ws.cell(row=row, column=4, value=f"=B{row}+C{row}")
            apply_currency_format(ws, row, 4)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value='Dépôt CDN')
        ws.cell(row=row, column=4, value=nas.deposit_cdn or 0)
        apply_currency_format(ws, row, 4)

        row += 1
        ws.cell(row=row, column=1, value='Dépôt USD')
        ws.cell(row=row, column=4, value=nas.deposit_us or 0)
        apply_currency_format(ws, row, 4)

        row += 2
        ws.cell(row=row, column=1, value='SOLDE')
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=4, value=nas.recap_balance or 0)
        ws.cell(row=row, column=4).font = Font(bold=True)
        apply_currency_format(ws, row, 4)
        format_variance_cell(ws, row, 4, nas.recap_balance)

        auto_width(ws)

        # Sheet 4: Transelect
        ws = wb.create_sheet('Transelect')
        ws['A1'] = 'TRANSELECT'
        ws['A1'].font = Font(bold=True, size=12)

        rest = nas.get_json('transelect_restaurant')
        recep = nas.get_json('transelect_reception')

        row = 3
        ws.cell(row=row, column=1, value='RESTAURANT').font = Font(bold=True)
        row += 1
        apply_header_row(ws, row, ['Terminal', 'Débit', 'Visa', 'MC', 'Amex', 'Discover', 'Total'])
        row += 1

        card_types = ['debit', 'visa', 'mc', 'amex', 'discover']
        for term, vals in rest.items():
            ws.cell(row=row, column=1, value=term)
            col = 2
            total_col = col + len(card_types)
            for ct in card_types:
                val = vals.get(ct, 0) if isinstance(vals, dict) else 0
                ws.cell(row=row, column=col, value=val)
                apply_currency_format(ws, row, col)
                col += 1
            ws.cell(row=row, column=total_col, value=f"=SUM(B{row}:F{row})")
            apply_currency_format(ws, row, total_col)
            row += 1

        row += 2
        ws.cell(row=row, column=1, value='RÉCEPTION').font = Font(bold=True)
        row += 1
        apply_header_row(ws, row, ['Type Carte', 'Fusebox', 'Terminal 8', 'K053', 'Total Rec', 'Daily Rev', 'Variance'])
        row += 1

        for ct in card_types:
            ct_data = recep.get(ct, {}) if isinstance(recep, dict) else {}
            ws.cell(row=row, column=1, value=ct.upper())

            fusebox = ct_data.get('fusebox', 0) if isinstance(ct_data, dict) else 0
            term8 = ct_data.get('term8', 0) if isinstance(ct_data, dict) else 0
            k053 = ct_data.get('k053', 0) if isinstance(ct_data, dict) else 0
            daily_rev = ct_data.get('daily_rev', 0) if isinstance(ct_data, dict) else 0

            ws.cell(row=row, column=2, value=fusebox)
            apply_currency_format(ws, row, 2)
            ws.cell(row=row, column=3, value=term8)
            apply_currency_format(ws, row, 3)
            ws.cell(row=row, column=4, value=k053)
            apply_currency_format(ws, row, 4)

            ws.cell(row=row, column=5, value=f"=B{row}+C{row}+D{row}")
            apply_currency_format(ws, row, 5)

            ws.cell(row=row, column=6, value=daily_rev)
            apply_currency_format(ws, row, 6)

            ws.cell(row=row, column=7, value=f"=E{row}-F{row}")
            apply_currency_format(ws, row, 7)

            row += 1

        auto_width(ws)

        # Sheet 5: GEAC
        ws = wb.create_sheet('GEAC')
        ws['A1'] = 'GEAC'
        ws['A1'].font = Font(bold=True, size=12)

        geac_cashout = nas.get_json('geac_cashout')
        geac_daily_rev = nas.get_json('geac_daily_rev')

        row = 3
        ws.cell(row=row, column=1, value='CASHOUT vs DAILY REV').font = Font(bold=True)
        row += 1
        apply_header_row(ws, row, ['Type Carte', 'Cashout', 'Daily Rev', 'Variance'])
        row += 1

        for ct in card_types:
            ws.cell(row=row, column=1, value=ct.upper())

            co_val = geac_cashout.get(ct, 0) if isinstance(geac_cashout, dict) else 0
            dr_val = geac_daily_rev.get(ct, 0) if isinstance(geac_daily_rev, dict) else 0

            ws.cell(row=row, column=2, value=co_val)
            apply_currency_format(ws, row, 2)
            ws.cell(row=row, column=3, value=dr_val)
            apply_currency_format(ws, row, 3)
            ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
            apply_currency_format(ws, row, 4)

            row += 1

        row += 2
        ws.cell(row=row, column=1, value='AR BALANCE').font = Font(bold=True)
        row += 1
        apply_header_row(ws, row, ['Description', 'Montant'])
        row += 1

        ar_rows = [
            ('Solde précédent', nas.geac_ar_previous),
            ('Nouvelles charges', nas.geac_ar_charges),
            ('Paiements', nas.geac_ar_payments),
            ('Nouveau solde', nas.geac_ar_new_balance),
            ('Variance', nas.geac_ar_variance),
        ]
        for label, val in ar_rows:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=val or 0)
            apply_currency_format(ws, row, 2)
            row += 1

        auto_width(ws)

        # Sheet 6: SD-Dépôt
        ws = wb.create_sheet('SD-Dépôt')
        ws['A1'] = 'SD - SOMMAIRE JOURNALIER DES DÉPÔTS'
        ws['A1'].font = Font(bold=True, size=12)

        row = 3
        ws.cell(row=row, column=1, value='SD ENTRIES').font = Font(bold=True)
        row += 1
        apply_header_row(ws, row, ['Département', 'Nom', 'Devise', 'Montant', 'Vérifié', 'Remboursement'])
        row += 1

        sd_entries = nas.get_json('sd_entries')
        if isinstance(sd_entries, list):
            for e in sd_entries:
                ws.cell(row=row, column=1, value=e.get('department', ''))
                ws.cell(row=row, column=2, value=e.get('name', ''))
                ws.cell(row=row, column=3, value=e.get('currency', ''))
                ws.cell(row=row, column=4, value=e.get('amount', 0))
                apply_currency_format(ws, row, 4)
                ws.cell(row=row, column=5, value=e.get('verified', 0))
                apply_currency_format(ws, row, 5)
                ws.cell(row=row, column=6, value=e.get('reimbursement', 0))
                apply_currency_format(ws, row, 6)
                row += 1

        row += 2
        ws.cell(row=row, column=1, value='DÉPÔT ENVELOPES').font = Font(bold=True)
        row += 1
        apply_header_row(ws, row, ['Client', 'Date', 'Montants'])
        row += 1

        depot_data = nas.get_json('depot_data')
        if isinstance(depot_data, dict):
            for client, data in depot_data.items():
                if isinstance(data, dict):
                    ws.cell(row=row, column=1, value=client)
                    ws.cell(row=row, column=2, value=data.get('date', ''))
                    amounts = data.get('amounts', [])
                    ws.cell(row=row, column=3, value=sum(a for a in amounts if isinstance(a, (int, float))))
                    apply_currency_format(ws, row, 3)
                    row += 1

        auto_width(ws)

        # Sheet 7: SetD
        ws = wb.create_sheet('SetD')
        ws['A1'] = 'SETD - SET-DÉJEUNER'
        ws['A1'].font = Font(bold=True, size=12)

        row = 3
        ws.cell(row=row, column=1, value='RJ Balance')
        ws.cell(row=row, column=2, value=nas.setd_rj_balance or 0)
        apply_currency_format(ws, row, 2)

        row += 2
        ws.cell(row=row, column=1, value='PERSONNEL').font = Font(bold=True)
        row += 1
        apply_header_row(ws, row, ['Nom', 'Colonne', 'Montant'])
        row += 1

        setd_personnel = nas.get_json('setd_personnel')
        if isinstance(setd_personnel, list):
            for e in setd_personnel:
                ws.cell(row=row, column=1, value=e.get('name', ''))
                ws.cell(row=row, column=2, value=e.get('column_letter', ''))
                ws.cell(row=row, column=3, value=e.get('amount', 0))
                apply_currency_format(ws, row, 3)
                row += 1

        auto_width(ws)

        # Sheet 8: HP-Admin
        ws = wb.create_sheet('HP-Admin')
        ws['A1'] = 'HP/ADMIN - HÔTEL PROMOTION & ADMINISTRATION'
        ws['A1'].font = Font(bold=True, size=12)

        row = 3
        apply_header_row(ws, row, ['Secteur', 'Nourriture', 'Boisson', 'Bière', 'Vin', 'Minéraux', 'Autre', 'Pourboire', 'Raison', 'Autorisé par'])
        row += 1

        hp_entries = nas.get_json('hp_admin_entries')
        if isinstance(hp_entries, list):
            for e in hp_entries:
                ws.cell(row=row, column=1, value=e.get('area', ''))
                col = 2
                for field in ['nourriture', 'boisson', 'biere', 'vin', 'mineraux', 'autre', 'pourboire']:
                    val = e.get(field, 0)
                    ws.cell(row=row, column=col, value=val)
                    apply_currency_format(ws, row, col)
                    col += 1
                ws.cell(row=row, column=col, value=e.get('raison', ''))
                col += 1
                ws.cell(row=row, column=col, value=e.get('autorise_par', ''))
                row += 1

        auto_width(ws)

        # Sheet 9: Internet
        ws = wb.create_sheet('Internet')
        ws['A1'] = 'INTERNET'
        ws['A1'].font = Font(bold=True, size=12)

        row = 3
        apply_header_row(ws, row, ['Description', 'Montant'])
        row += 1

        ws.cell(row=row, column=1, value='CD 36.1 (Lightspeed)')
        ws.cell(row=row, column=2, value=nas.internet_ls_361 or 0)
        apply_currency_format(ws, row, 2)
        row += 1

        ws.cell(row=row, column=1, value='CD 36.5 (Lightspeed)')
        ws.cell(row=row, column=2, value=nas.internet_ls_365 or 0)
        apply_currency_format(ws, row, 2)
        row += 1

        ws.cell(row=row, column=1, value='VARIANCE')
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=nas.internet_variance or 0)
        ws.cell(row=row, column=2).font = Font(bold=True)
        apply_currency_format(ws, row, 2)
        format_variance_cell(ws, row, 2, nas.internet_variance)

        auto_width(ws)

        # Sheet 10: Sonifi
        ws = wb.create_sheet('Sonifi')
        ws['A1'] = 'SONIFI'
        ws['A1'].font = Font(bold=True, size=12)

        row = 3
        apply_header_row(ws, row, ['Description', 'Montant'])
        row += 1

        ws.cell(row=row, column=1, value='CD 35.2 (Lightspeed)')
        ws.cell(row=row, column=2, value=nas.sonifi_cd_352 or 0)
        apply_currency_format(ws, row, 2)
        row += 1

        ws.cell(row=row, column=1, value='Email Courriel (03h00)')
        ws.cell(row=row, column=2, value=nas.sonifi_email or 0)
        apply_currency_format(ws, row, 2)
        row += 1

        ws.cell(row=row, column=1, value='VARIANCE')
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=nas.sonifi_variance or 0)
        ws.cell(row=row, column=2).font = Font(bold=True)
        apply_currency_format(ws, row, 2)
        format_variance_cell(ws, row, 2, nas.sonifi_variance)

        auto_width(ws)

        # Sheet 11: Jour
        ws = wb.create_sheet('Jour')
        ws['A1'] = 'JOUR - REVENU JOURNALIER'
        ws['A1'].font = Font(bold=True, size=12)

        row = 3
        ws.cell(row=row, column=1, value='F&B RESTAURATION').font = Font(bold=True)
        row += 1

        fb_depts = ['Café Link', 'Piazza/Cupola', 'Marché Spesa', 'Service Chambres', 'Banquet']
        fb_cats = ['nourriture', 'boisson', 'bieres', 'mineraux', 'vins']
        dept_keys = ['cafe', 'piazza', 'spesa', 'chambres_svc', 'banquet']

        apply_header_row(ws, row, ['Département'] + [c.capitalize() for c in fb_cats] + ['Total'])
        row += 1

        for dept_label, dept_key in zip(fb_depts, dept_keys):
            ws.cell(row=row, column=1, value=dept_label)
            col = 2
            for cat in fb_cats:
                val = getattr(nas, f'jour_{dept_key}_{cat}', 0) or 0
                ws.cell(row=row, column=col, value=val)
                apply_currency_format(ws, row, col)
                col += 1
            ws.cell(row=row, column=col, value=f"=SUM(B{row}:F{row})")
            apply_currency_format(ws, row, col)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value='F&B EXTRA').font = Font(bold=True)
        row += 1

        extra_fields = [
            ('Pourboires', 'jour_pourboires'),
            ('Tabagie', 'jour_tabagie'),
            ('Location Salle', 'jour_location_salle'),
        ]
        for label, field in extra_fields:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=getattr(nas, field, 0) or 0)
            apply_currency_format(ws, row, 2)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value='HÉBERGEMENT').font = Font(bold=True)
        row += 1

        hebergement_fields = [
            ('Revenu Chambres', 'jour_room_revenue'),
            ('Téléphone Local', 'jour_tel_local'),
            ('Téléphone Interurbain', 'jour_tel_interurbain'),
            ('Téléphones Publics', 'jour_tel_publics'),
        ]
        for label, field in hebergement_fields:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=getattr(nas, field, 0) or 0)
            apply_currency_format(ws, row, 2)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value='AUTRES REVENUS').font = Font(bold=True)
        row += 1

        autres_fields = [
            ('Nettoyeur', 'jour_nettoyeur'),
            ('Machine Distribution', 'jour_machine_distrib'),
            ('Autres GL', 'jour_autres_gl'),
            ('Sonifi', 'jour_sonifi'),
            ('Lit Pliant', 'jour_lit_pliant'),
            ('Boutique', 'jour_boutique'),
            ('Internet', 'jour_internet'),
            ('Massage', 'jour_massage'),
        ]
        for label, field in autres_fields:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=getattr(nas, field, 0) or 0)
            apply_currency_format(ws, row, 2)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value='TAXES').font = Font(bold=True)
        row += 1

        tax_fields = [
            ('TVQ', 'jour_tvq'),
            ('TPS', 'jour_tps'),
            ('Taxe Hébergement', 'jour_taxe_hebergement'),
        ]
        for label, field in tax_fields:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=getattr(nas, field, 0) or 0)
            apply_currency_format(ws, row, 2)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value='OCCUPATION').font = Font(bold=True)
        row += 1

        occ_fields = [
            ('Simples', 'jour_rooms_simple'),
            ('Doubles', 'jour_rooms_double'),
            ('Suites', 'jour_rooms_suite'),
            ('Comp', 'jour_rooms_comp'),
            ('Clients', 'jour_nb_clients'),
            ('Hors Usage', 'jour_rooms_hors_usage'),
            ('Taux Occupancy %', 'jour_occupancy_rate'),
        ]
        for label, field in occ_fields:
            ws.cell(row=row, column=1, value=label)
            val = getattr(nas, field, 0) or 0
            ws.cell(row=row, column=2, value=val)
            if 'occupancy' in field:
                ws.cell(row=row, column=2).number_format = percent_fmt
            row += 1

        row += 1
        ws.cell(row=row, column=1, value='TOTAL REVENU').font = Font(bold=True)
        ws.cell(row=row, column=2, value=nas.jour_total_revenue or 0)
        ws.cell(row=row, column=2).font = Font(bold=True)
        apply_currency_format(ws, row, 2)

        row += 2
        ws.cell(row=row, column=1, value='KPIs').font = Font(bold=True)
        row += 1

        kpi_fields = [
            ('ADR', 'jour_adr'),
            ('RevPAR', 'jour_revpar'),
        ]
        for label, field in kpi_fields:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=getattr(nas, field, 0) or 0)
            apply_currency_format(ws, row, 2)
            row += 1

        auto_width(ws)

        # Sheet 12: Quasimodo
        ws = wb.create_sheet('Quasimodo')
        ws['A1'] = 'QUASIMODO - RÉCONCILIATION GLOBALE'
        ws['A1'].font = Font(bold=True, size=12)

        row = 3
        ws.cell(row=row, column=1, value='CARTES').font = Font(bold=True)
        row += 1
        apply_header_row(ws, row, ['Type Carte', 'F&B', 'Réception', 'Total'])
        row += 1

        for ct in card_types:
            ws.cell(row=row, column=1, value=ct.upper())

            fb_val = getattr(nas, f'quasi_fb_{ct}', 0) or 0
            rec_val = getattr(nas, f'quasi_rec_{ct}', 0) or 0
            total = fb_val + rec_val

            ws.cell(row=row, column=2, value=fb_val)
            apply_currency_format(ws, row, 2)
            ws.cell(row=row, column=3, value=rec_val)
            apply_currency_format(ws, row, 3)
            ws.cell(row=row, column=4, value=total)
            apply_currency_format(ws, row, 4)

            row += 1

        row += 2
        ws.cell(row=row, column=1, value='AMEX NET (factor 0.9735)').font = Font(bold=True)
        amex_gross = (nas.quasi_fb_amex or 0) + (nas.quasi_rec_amex or 0)
        amex_net = round(amex_gross * (nas.quasi_amex_factor or 0.9735), 2)
        ws.cell(row=row, column=2, value=amex_net)
        apply_currency_format(ws, row, 2)

        row += 1
        ws.cell(row=row, column=1, value='CASH CDN').font = Font(bold=True)
        ws.cell(row=row, column=2, value=nas.quasi_cash_cdn or 0)
        apply_currency_format(ws, row, 2)

        row += 1
        ws.cell(row=row, column=1, value='CASH USD').font = Font(bold=True)
        ws.cell(row=row, column=2, value=nas.quasi_cash_usd or 0)
        apply_currency_format(ws, row, 2)

        row += 2
        ws.cell(row=row, column=1, value='TOTAL QUASIMODO').font = Font(bold=True, size=11)
        ws.cell(row=row, column=2, value=nas.quasi_total or 0)
        ws.cell(row=row, column=2).font = Font(bold=True, size=11)
        apply_currency_format(ws, row, 2)

        row += 1
        ws.cell(row=row, column=1, value='TOTAL RJ (Jour)').font = Font(bold=True, size=11)
        ws.cell(row=row, column=2, value=nas.quasi_rj_total or 0)
        ws.cell(row=row, column=2).font = Font(bold=True, size=11)
        apply_currency_format(ws, row, 2)

        row += 1
        ws.cell(row=row, column=1, value='VARIANCE').font = Font(bold=True, size=11)
        ws.cell(row=row, column=2, value=nas.quasi_variance or 0)
        ws.cell(row=row, column=2).font = Font(bold=True, size=11)
        apply_currency_format(ws, row, 2)
        format_variance_cell(ws, row, 2, nas.quasi_variance)

        auto_width(ws)

        # Sheet 13: DBRS
        ws = wb.create_sheet('DBRS')
        ws['A1'] = 'DBRS - DAILY BUSINESS REVIEW SUMMARY'
        ws['A1'].font = Font(bold=True, size=12)

        row = 3
        apply_header_row(ws, row, ['Segment', 'Montant'])
        row += 1

        dbrs_segments = nas.get_json('dbrs_market_segments')
        if isinstance(dbrs_segments, dict):
            for segment, amount in dbrs_segments.items():
                ws.cell(row=row, column=1, value=segment.capitalize())
                ws.cell(row=row, column=2, value=amount or 0)
                apply_currency_format(ws, row, 2)
                row += 1

        row += 2
        apply_header_row(ws, row, ['Métrique', 'Valeur'])
        row += 1

        dbrs_fields = [
            ('Revenu Chambre Aujourd\'hui', nas.dbrs_daily_rev_today),
            ('ADR', nas.dbrs_adr),
            ('Nombre Chambres Vendues', nas.dbrs_house_count),
            ('Nombre No-show', nas.dbrs_noshow_count),
            ('Revenu No-show', nas.dbrs_noshow_revenue),
        ]
        for label, val in dbrs_fields:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=val or 0)
            if 'ADR' in label or 'Revenu' in label:
                apply_currency_format(ws, row, 2)
            row += 1

        auto_width(ws)

        # Sheet 14: Sommaire
        ws = wb.create_sheet('Sommaire')
        ws['A1'] = 'SOMMAIRE - VALIDATION & STATUT'
        ws['A1'].font = Font(bold=True, size=12)

        row = 3
        apply_header_row(ws, row, ['Vérification', 'Statut', 'Notes'])
        row += 1

        checks = [
            ('Recap Équilibré', nas.is_recap_balanced, f'Solde: {nas.recap_balance:.2f}'),
            ('Transelect Équilibré', nas.is_transelect_balanced, f'Variance: {nas.transelect_variance:.2f}'),
            ('AR Équilibré', nas.is_ar_balanced, f'Variance: {nas.geac_ar_variance:.2f}'),
            ('Internet Équilibré', abs(nas.internet_variance or 0) < 0.02, f'Variance: {nas.internet_variance:.2f}'),
            ('Sonifi Équilibré', abs(nas.sonifi_variance or 0) < 0.02, f'Variance: {nas.sonifi_variance:.2f}'),
            ('Quasimodo Équilibré', abs(nas.quasi_variance or 0) < 0.02, f'Variance: {nas.quasi_variance:.2f}'),
            ('Entièrement Équilibré', nas.is_fully_balanced, 'Tous les contrôles réussis' if nas.is_fully_balanced else 'Vérifier les variances'),
        ]

        for check_name, is_balanced, note in checks:
            ws.cell(row=row, column=1, value=check_name)
            status = 'ÉQUILIBRÉ' if is_balanced else 'VARIANCE'
            ws.cell(row=row, column=2, value=status)
            if is_balanced:
                ws.cell(row=row, column=2).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            else:
                ws.cell(row=row, column=2).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            ws.cell(row=row, column=3, value=note)
            row += 1

        row += 2
        ws.cell(row=row, column=1, value='INFORMATION GÉNÉRALE').font = Font(bold=True)
        row += 1

        info_fields = [
            ('Date Audit', d.isoformat()),
            ('Auditeur', nas.auditor_name or ''),
            ('Statut Session', nas.status),
            ('Date Création', nas.created_at.strftime('%Y-%m-%d %H:%M:%S') if nas.created_at else ''),
            ('Date Complétion', nas.completed_at.strftime('%Y-%m-%d %H:%M:%S') if nas.completed_at else ''),
            ('Notes', nas.notes or ''),
        ]
        for label, val in info_fields:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=str(val))
            row += 1

        auto_width(ws)

        # Save to bytes buffer
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f'RJ_{d.strftime("%Y-%m-%d")}.xlsx'
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"Excel export error: {e}", exc_info=True)
        return jsonify({'error': f'Erreur lors de l\'export: {str(e)}'}), 500


@rj_native_bp.route('/api/rj/native/export/excel/<audit_date>')
@auth_required
def export_excel(audit_date):
    """Generate an Excel file from the native session data."""
    try:
        d = datetime.strptime(audit_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400

    nas = NightAuditSession.query.filter_by(audit_date=d).first()
    if not nas:
        return jsonify({'error': 'Session non trouvée'}), 404

    try:
        import openpyxl
        wb = openpyxl.Workbook()

        # Recap sheet
        ws = wb.active
        ws.title = 'Recap'
        ws['A1'] = 'RECAP - Audit de nuit'
        ws['A2'] = f'Date: {d.isoformat()}'
        ws['A3'] = f'Auditeur: {nas.auditor_name}'

        headers = ['', 'Lecture', 'Correction', 'Net']
        for i, h in enumerate(headers, 1):
            ws.cell(row=5, column=i, value=h)

        recap_rows = [
            ('Comptant LightSpeed', nas.cash_ls_lecture, nas.cash_ls_corr),
            ('Comptant POSitouch', nas.cash_pos_lecture, nas.cash_pos_corr),
            ('Chq A/R', nas.cheque_ar_lecture, nas.cheque_ar_corr),
            ('Chq Daily Rev', nas.cheque_dr_lecture, nas.cheque_dr_corr),
            ('Remb. gratuite', nas.remb_gratuite_lecture, nas.remb_gratuite_corr),
            ('Remb. client', nas.remb_client_lecture, nas.remb_client_corr),
            ('DueBack Réception', nas.dueback_reception_lecture, nas.dueback_reception_corr),
            ('DueBack NB', nas.dueback_nb_lecture, nas.dueback_nb_corr),
        ]
        for i, (label, lect, corr) in enumerate(recap_rows, 6):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=lect or 0)
            ws.cell(row=i, column=3, value=corr or 0)
            ws.cell(row=i, column=4, value=(lect or 0) + (corr or 0))

        row = 6 + len(recap_rows) + 1
        ws.cell(row=row, column=1, value='Dépôt CDN')
        ws.cell(row=row, column=4, value=nas.deposit_cdn or 0)
        ws.cell(row=row + 1, column=1, value='Dépôt US')
        ws.cell(row=row + 1, column=4, value=nas.deposit_us or 0)
        ws.cell(row=row + 3, column=1, value='BALANCE')
        ws.cell(row=row + 3, column=4, value=nas.recap_balance)

        # Transelect sheet
        ws2 = wb.create_sheet('Transelect')
        ws2['A1'] = 'TRANSELECT - Cartes de crédit'
        ws2['A3'] = 'Restaurant'
        rest = nas.get_json('transelect_restaurant')
        row = 4
        for term, vals in rest.items():
            ws2.cell(row=row, column=1, value=term)
            if isinstance(vals, dict):
                ws2.cell(row=row, column=2, value=vals.get('debit', 0))
                ws2.cell(row=row, column=3, value=vals.get('visa', 0))
                ws2.cell(row=row, column=4, value=vals.get('mc', 0))
                ws2.cell(row=row, column=5, value=vals.get('amex', 0))
            row += 1

        row += 1
        ws2.cell(row=row, column=1, value='Variance')
        ws2.cell(row=row, column=2, value=nas.transelect_variance)

        # DueBack sheet
        ws3 = wb.create_sheet('DueBack')
        ws3['A1'] = 'DUEBACK'
        ws3.cell(row=3, column=1, value='Réceptionniste')
        ws3.cell(row=3, column=2, value='Solde préc.')
        ws3.cell(row=3, column=3, value='Nouveau')
        entries = nas.get_json('dueback_entries')
        if isinstance(entries, list):
            for i, e in enumerate(entries, 4):
                ws3.cell(row=i, column=1, value=e.get('name', ''))
                ws3.cell(row=i, column=2, value=e.get('previous', 0))
                ws3.cell(row=i, column=3, value=e.get('nouveau', 0))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f'RJ_Native_{d.isoformat()}.xlsx'
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        logger.error(f"Excel export error: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@rj_native_bp.route('/api/rj/native/export/month/<int:year>/<int:month>')
@auth_required
def export_month(year, month):
    """Generate a monthly summary Excel with one row per day."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from calendar import monthrange
        from datetime import date as date_type

        # Validate month/year
        if not (1 <= month <= 12):
            return jsonify({'error': 'Mois invalide (1-12)'}), 400

        # Query all sessions for this month
        first_day = date_type(year, month, 1)
        last_day = date_type(year, month, monthrange(year, month)[1])

        sessions = NightAuditSession.query.filter(
            NightAuditSession.audit_date >= first_day,
            NightAuditSession.audit_date <= last_day
        ).order_by(NightAuditSession.audit_date).all()

        # Create workbook with 4 sheets
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Define styles
        font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        fill_header = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        fill_alt = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
        fill_warning = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        font_bold = Font(name='Arial', size=10, bold=True)
        font_regular = Font(name='Arial', size=10)
        alignment_center = Alignment(horizontal='center', vertical='center')
        alignment_right = Alignment(horizontal='right', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet 1: Sommaire
        ws1 = wb.create_sheet('Sommaire', 0)
        ws1.sheet_properties.tabColor = '0070C0'  # Blue

        headers_sommaire = [
            'Date', 'Auditeur', 'Statut',
            'Revenu Chambres', 'Revenu F&B', 'Revenu Total',
            'ADR', 'Occ%', 'RevPAR',
            'Chambres vendues', 'Nb Clients',
            'Deposit CDN', 'Deposit US',
            'Recap Balance', 'Quasimodo Variance',
            'Internet Var', 'Sonifi Var'
        ]

        for col_idx, header in enumerate(headers_sommaire, 1):
            cell = ws1.cell(row=1, column=col_idx, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = alignment_center
            cell.border = thin_border

        # Freeze panes
        ws1.freeze_panes = 'A2'

        # Data rows
        row_num = 2
        session_dict = {s.audit_date: s for s in sessions}

        for day in range(1, monthrange(year, month)[1] + 1):
            current_date = date_type(year, month, day)
            session = session_dict.get(current_date)

            # Alternate row colors
            fill = fill_alt if day % 2 == 0 else None

            col = 1
            # Date
            cell = ws1.cell(row=row_num, column=col, value=current_date)
            cell.font = font_regular
            cell.border = thin_border
            if fill:
                cell.fill = fill
            col += 1

            if session:
                # Check if status is not 'locked' -> apply yellow warning
                if session.status != 'locked':
                    fill_row = fill_warning
                else:
                    fill_row = fill

                # Auditeur
                cell = ws1.cell(row=row_num, column=col, value=session.auditor_name or '')
                cell.font = font_regular
                cell.border = thin_border
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Statut
                cell = ws1.cell(row=row_num, column=col, value=session.status or '')
                cell.font = font_regular
                cell.border = thin_border
                cell.alignment = alignment_center
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Revenu Chambres
                cell = ws1.cell(row=row_num, column=col, value=session.jour_room_revenue or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Revenu F&B
                cell = ws1.cell(row=row_num, column=col, value=session.jour_total_fb or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Revenu Total
                cell = ws1.cell(row=row_num, column=col, value=session.jour_total_revenue or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # ADR
                cell = ws1.cell(row=row_num, column=col, value=session.jour_adr or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Occ%
                cell = ws1.cell(row=row_num, column=col, value=(session.jour_occupancy_rate or 0) / 100)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '0.0%'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # RevPAR
                cell = ws1.cell(row=row_num, column=col, value=session.jour_revpar or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Chambres vendues
                rooms_sold = (session.jour_rooms_simple or 0) + (session.jour_rooms_double or 0) + \
                             (session.jour_rooms_suite or 0) + (session.jour_rooms_comp or 0)
                cell = ws1.cell(row=row_num, column=col, value=rooms_sold)
                cell.font = font_regular
                cell.border = thin_border
                cell.alignment = alignment_center
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Nb Clients
                cell = ws1.cell(row=row_num, column=col, value=session.jour_nb_clients or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.alignment = alignment_center
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Deposit CDN
                cell = ws1.cell(row=row_num, column=col, value=session.deposit_cdn or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Deposit US
                cell = ws1.cell(row=row_num, column=col, value=session.deposit_us or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Recap Balance
                cell = ws1.cell(row=row_num, column=col, value=session.recap_balance or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Quasimodo Variance
                cell = ws1.cell(row=row_num, column=col, value=session.quasi_variance or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Internet Variance
                cell = ws1.cell(row=row_num, column=col, value=session.internet_variance or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Sonifi Variance
                cell = ws1.cell(row=row_num, column=col, value=session.sonifi_variance or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row

            else:
                # Empty row for missing day
                for c in range(col, len(headers_sommaire) + 1):
                    cell = ws1.cell(row=row_num, column=c, value='—')
                    cell.font = font_regular
                    cell.border = thin_border
                    if fill:
                        cell.fill = fill

            row_num += 1

        # Add total and average row
        total_row = row_num
        total_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        total_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='double'),
            bottom=Side(style='thin')
        )

        cell = ws1.cell(row=total_row, column=1, value='TOTAL / MOYENNE')
        cell.font = font_bold
        cell.fill = total_fill
        cell.border = total_border

        # SUM for numeric columns (starting from col 4 through end)
        for col_idx in range(4, len(headers_sommaire) + 1):
            cell = ws1.cell(row=total_row, column=col_idx)
            cell.font = font_bold
            cell.fill = total_fill
            cell.border = total_border
            col_letter = openpyxl.utils.get_column_letter(col_idx)

            # Columns 7, 8 (ADR, Occ%), 9 (RevPAR) should be AVERAGE
            if col_idx in [7, 8, 9]:
                cell.value = f'=AVERAGE({col_letter}2:{col_letter}{total_row-1})'
            else:
                cell.value = f'=SUM({col_letter}2:{col_letter}{total_row-1})'

            if col_idx == 8:  # Occ%
                cell.number_format = '0.0%'
            elif col_idx >= 4:
                cell.number_format = '$#,##0.00'

        # Auto-fit columns
        for col_idx, header in enumerate(headers_sommaire, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_len = len(str(header)) + 2
            ws1.column_dimensions[col_letter].width = max(max_len, 12)

        # Sheet 2: F&B Détail
        ws2 = wb.create_sheet('F&B Détail', 1)
        ws2.sheet_properties.tabColor = '00B050'  # Green

        headers_fb = [
            'Date', 'Café', 'Piazza', 'Spesa', 'Chambres Svc', 'Banquet',
            'Pourboires', 'Tabagie', 'Location salle', 'Total F&B'
        ]

        for col_idx, header in enumerate(headers_fb, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = alignment_center
            cell.border = thin_border

        ws2.freeze_panes = 'A2'

        row_num = 2
        for day in range(1, monthrange(year, month)[1] + 1):
            current_date = date_type(year, month, day)
            session = session_dict.get(current_date)

            fill = fill_alt if day % 2 == 0 else None

            # Date
            cell = ws2.cell(row=row_num, column=1, value=current_date)
            cell.font = font_regular
            cell.border = thin_border
            if fill:
                cell.fill = fill

            if session:
                col = 2
                # Café (sum of 5 categories)
                cafe_val = (session.jour_cafe_nourriture or 0) + (session.jour_cafe_boisson or 0) + \
                          (session.jour_cafe_bieres or 0) + (session.jour_cafe_mineraux or 0) + \
                          (session.jour_cafe_vins or 0)
                cell = ws2.cell(row=row_num, column=col, value=cafe_val)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill:
                    cell.fill = fill
                col += 1

                # Piazza
                piazza_val = (session.jour_piazza_nourriture or 0) + (session.jour_piazza_boisson or 0) + \
                            (session.jour_piazza_bieres or 0) + (session.jour_piazza_mineraux or 0) + \
                            (session.jour_piazza_vins or 0)
                cell = ws2.cell(row=row_num, column=col, value=piazza_val)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill:
                    cell.fill = fill
                col += 1

                # Spesa
                spesa_val = (session.jour_spesa_nourriture or 0) + (session.jour_spesa_boisson or 0) + \
                           (session.jour_spesa_bieres or 0) + (session.jour_spesa_mineraux or 0) + \
                           (session.jour_spesa_vins or 0)
                cell = ws2.cell(row=row_num, column=col, value=spesa_val)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill:
                    cell.fill = fill
                col += 1

                # Chambres Svc
                svc_val = (session.jour_chambres_svc_nourriture or 0) + (session.jour_chambres_svc_boisson or 0) + \
                         (session.jour_chambres_svc_bieres or 0) + (session.jour_chambres_svc_mineraux or 0) + \
                         (session.jour_chambres_svc_vins or 0)
                cell = ws2.cell(row=row_num, column=col, value=svc_val)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill:
                    cell.fill = fill
                col += 1

                # Banquet
                banquet_val = (session.jour_banquet_nourriture or 0) + (session.jour_banquet_boisson or 0) + \
                             (session.jour_banquet_bieres or 0) + (session.jour_banquet_mineraux or 0) + \
                             (session.jour_banquet_vins or 0)
                cell = ws2.cell(row=row_num, column=col, value=banquet_val)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill:
                    cell.fill = fill
                col += 1

                # Pourboires
                cell = ws2.cell(row=row_num, column=col, value=session.jour_pourboires or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill:
                    cell.fill = fill
                col += 1

                # Tabagie
                cell = ws2.cell(row=row_num, column=col, value=session.jour_tabagie or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill:
                    cell.fill = fill
                col += 1

                # Location salle
                cell = ws2.cell(row=row_num, column=col, value=session.jour_location_salle or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill:
                    cell.fill = fill
                col += 1

                # Total F&B
                cell = ws2.cell(row=row_num, column=col, value=session.jour_total_fb or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill:
                    cell.fill = fill
            else:
                for c in range(2, len(headers_fb) + 1):
                    cell = ws2.cell(row=row_num, column=c, value='—')
                    cell.font = font_regular
                    cell.border = thin_border
                    if fill:
                        cell.fill = fill

            row_num += 1

        # Total row for F&B sheet
        total_row = row_num
        cell = ws2.cell(row=total_row, column=1, value='TOTAL')
        cell.font = font_bold
        cell.fill = total_fill
        cell.border = total_border

        for col_idx in range(2, len(headers_fb) + 1):
            cell = ws2.cell(row=total_row, column=col_idx)
            cell.font = font_bold
            cell.fill = total_fill
            cell.border = total_border
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            cell.value = f'=SUM({col_letter}2:{col_letter}{total_row-1})'
            cell.number_format = '$#,##0.00'

        # Auto-fit columns
        for col_idx, header in enumerate(headers_fb, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws2.column_dimensions[col_letter].width = max(len(str(header)) + 2, 12)

        # Sheet 3: Réconciliation
        ws3 = wb.create_sheet('Réconciliation', 2)
        ws3.sheet_properties.tabColor = 'C00000'  # Red

        headers_recon = [
            'Date', 'Auditeur',
            'Cash LS (Net)', 'Cash POS (Net)', 'Chèque AR', 'Chèque DR',
            'Remb Gratuite', 'Remb Client',
            'DueBack Rec', 'DueBack NB',
            'Balance Recap',
            'Quasi Total', 'Quasi RJ Total', 'Quasi Variance'
        ]

        for col_idx, header in enumerate(headers_recon, 1):
            cell = ws3.cell(row=1, column=col_idx, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = alignment_center
            cell.border = thin_border

        ws3.freeze_panes = 'A2'

        row_num = 2
        for day in range(1, monthrange(year, month)[1] + 1):
            current_date = date_type(year, month, day)
            session = session_dict.get(current_date)

            fill = fill_alt if day % 2 == 0 else None

            col = 1
            # Date
            cell = ws3.cell(row=row_num, column=col, value=current_date)
            cell.font = font_regular
            cell.border = thin_border
            if fill:
                cell.fill = fill
            col += 1

            if session:
                if session.status != 'locked':
                    fill_row = fill_warning
                else:
                    fill_row = fill

                # Auditeur
                cell = ws3.cell(row=row_num, column=col, value=session.auditor_name or '')
                cell.font = font_regular
                cell.border = thin_border
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Cash LS (Net) = lecture + corr
                cash_ls_net = (session.cash_ls_lecture or 0) + (session.cash_ls_corr or 0)
                cell = ws3.cell(row=row_num, column=col, value=cash_ls_net)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Cash POS (Net)
                cash_pos_net = (session.cash_pos_lecture or 0) + (session.cash_pos_corr or 0)
                cell = ws3.cell(row=row_num, column=col, value=cash_pos_net)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Chèque AR
                cheque_ar_net = (session.cheque_ar_lecture or 0) + (session.cheque_ar_corr or 0)
                cell = ws3.cell(row=row_num, column=col, value=cheque_ar_net)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Chèque DR
                cheque_dr_net = (session.cheque_dr_lecture or 0) + (session.cheque_dr_corr or 0)
                cell = ws3.cell(row=row_num, column=col, value=cheque_dr_net)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Remb Gratuite
                remb_grat_net = (session.remb_gratuite_lecture or 0) + (session.remb_gratuite_corr or 0)
                cell = ws3.cell(row=row_num, column=col, value=remb_grat_net)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Remb Client
                remb_cli_net = (session.remb_client_lecture or 0) + (session.remb_client_corr or 0)
                cell = ws3.cell(row=row_num, column=col, value=remb_cli_net)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # DueBack Rec
                db_rec_net = (session.dueback_reception_lecture or 0) + (session.dueback_reception_corr or 0)
                cell = ws3.cell(row=row_num, column=col, value=db_rec_net)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # DueBack NB
                db_nb_net = (session.dueback_nb_lecture or 0) + (session.dueback_nb_corr or 0)
                cell = ws3.cell(row=row_num, column=col, value=db_nb_net)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Balance Recap
                cell = ws3.cell(row=row_num, column=col, value=session.recap_balance or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Quasi Total
                cell = ws3.cell(row=row_num, column=col, value=session.quasi_total or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Quasi RJ Total
                cell = ws3.cell(row=row_num, column=col, value=session.quasi_rj_total or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
                col += 1

                # Quasi Variance
                cell = ws3.cell(row=row_num, column=col, value=session.quasi_variance or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill_row:
                    cell.fill = fill_row
            else:
                for c in range(col, len(headers_recon) + 1):
                    cell = ws3.cell(row=row_num, column=c, value='—')
                    cell.font = font_regular
                    cell.border = thin_border
                    if fill:
                        cell.fill = fill

            row_num += 1

        # Total row for Reconciliation sheet
        total_row = row_num
        cell = ws3.cell(row=total_row, column=1, value='TOTAL / MOYENNE')
        cell.font = font_bold
        cell.fill = total_fill
        cell.border = total_border
        cell = ws3.cell(row=total_row, column=2, value='')
        cell.font = font_bold
        cell.fill = total_fill
        cell.border = total_border

        for col_idx in range(3, len(headers_recon) + 1):
            cell = ws3.cell(row=total_row, column=col_idx)
            cell.font = font_bold
            cell.fill = total_fill
            cell.border = total_border
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            cell.value = f'=SUM({col_letter}2:{col_letter}{total_row-1})'
            cell.number_format = '$#,##0.00'

        # Auto-fit columns
        for col_idx, header in enumerate(headers_recon, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws3.column_dimensions[col_letter].width = max(len(str(header)) + 2, 12)

        # Sheet 4: Occupation
        ws4 = wb.create_sheet('Occupation', 3)
        ws4.sheet_properties.tabColor = '7030A0'  # Purple

        headers_occ = [
            'Date', 'Simple', 'Double', 'Suite', 'Comp', 'Total Vendues',
            'Hors Usage', 'Disponibles',
            'Occ%', 'ADR', 'RevPAR', 'Nb Clients'
        ]

        for col_idx, header in enumerate(headers_occ, 1):
            cell = ws4.cell(row=1, column=col_idx, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = alignment_center
            cell.border = thin_border

        ws4.freeze_panes = 'A2'

        TOTAL_ROOMS = 252  # From models.py

        row_num = 2
        for day in range(1, monthrange(year, month)[1] + 1):
            current_date = date_type(year, month, day)
            session = session_dict.get(current_date)

            fill = fill_alt if day % 2 == 0 else None

            col = 1
            # Date
            cell = ws4.cell(row=row_num, column=col, value=current_date)
            cell.font = font_regular
            cell.border = thin_border
            if fill:
                cell.fill = fill
            col += 1

            if session:
                # Simple
                cell = ws4.cell(row=row_num, column=col, value=session.jour_rooms_simple or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.alignment = alignment_center
                if fill:
                    cell.fill = fill
                col += 1

                # Double
                cell = ws4.cell(row=row_num, column=col, value=session.jour_rooms_double or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.alignment = alignment_center
                if fill:
                    cell.fill = fill
                col += 1

                # Suite
                cell = ws4.cell(row=row_num, column=col, value=session.jour_rooms_suite or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.alignment = alignment_center
                if fill:
                    cell.fill = fill
                col += 1

                # Comp
                cell = ws4.cell(row=row_num, column=col, value=session.jour_rooms_comp or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.alignment = alignment_center
                if fill:
                    cell.fill = fill
                col += 1

                # Total Vendues
                total_vendues = (session.jour_rooms_simple or 0) + (session.jour_rooms_double or 0) + \
                               (session.jour_rooms_suite or 0) + (session.jour_rooms_comp or 0)
                cell = ws4.cell(row=row_num, column=col, value=total_vendues)
                cell.font = font_regular
                cell.border = thin_border
                cell.alignment = alignment_center
                if fill:
                    cell.fill = fill
                col += 1

                # Hors Usage
                cell = ws4.cell(row=row_num, column=col, value=session.jour_rooms_hors_usage or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.alignment = alignment_center
                if fill:
                    cell.fill = fill
                col += 1

                # Disponibles
                disponibles = TOTAL_ROOMS - (session.jour_rooms_hors_usage or 0)
                cell = ws4.cell(row=row_num, column=col, value=disponibles)
                cell.font = font_regular
                cell.border = thin_border
                cell.alignment = alignment_center
                if fill:
                    cell.fill = fill
                col += 1

                # Occ%
                cell = ws4.cell(row=row_num, column=col, value=(session.jour_occupancy_rate or 0) / 100)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '0.0%'
                if fill:
                    cell.fill = fill
                col += 1

                # ADR
                cell = ws4.cell(row=row_num, column=col, value=session.jour_adr or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill:
                    cell.fill = fill
                col += 1

                # RevPAR
                cell = ws4.cell(row=row_num, column=col, value=session.jour_revpar or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.number_format = '$#,##0.00'
                if fill:
                    cell.fill = fill
                col += 1

                # Nb Clients
                cell = ws4.cell(row=row_num, column=col, value=session.jour_nb_clients or 0)
                cell.font = font_regular
                cell.border = thin_border
                cell.alignment = alignment_center
                if fill:
                    cell.fill = fill
            else:
                for c in range(col, len(headers_occ) + 1):
                    cell = ws4.cell(row=row_num, column=c, value='—')
                    cell.font = font_regular
                    cell.border = thin_border
                    if fill:
                        cell.fill = fill

            row_num += 1

        # Total row for Occupation sheet
        total_row = row_num
        cell = ws4.cell(row=total_row, column=1, value='TOTAL / MOYENNE')
        cell.font = font_bold
        cell.fill = total_fill
        cell.border = total_border

        for col_idx in range(2, len(headers_occ) + 1):
            cell = ws4.cell(row=total_row, column=col_idx)
            cell.font = font_bold
            cell.fill = total_fill
            cell.border = total_border
            col_letter = openpyxl.utils.get_column_letter(col_idx)

            # Columns for rates should be AVERAGE, others SUM
            if col_idx in [9, 10, 11]:  # Occ%, ADR, RevPAR
                cell.value = f'=AVERAGE({col_letter}2:{col_letter}{total_row-1})'
                if col_idx == 9:  # Occ%
                    cell.number_format = '0.0%'
                else:
                    cell.number_format = '$#,##0.00'
            else:
                cell.value = f'=SUM({col_letter}2:{col_letter}{total_row-1})'
                if col_idx <= 8:  # Room counts
                    cell.alignment = alignment_center

        # Auto-fit columns
        for col_idx, header in enumerate(headers_occ, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws4.column_dimensions[col_letter].width = max(len(str(header)) + 2, 12)

        # Save and return
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        month_name_fr = ['', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                        'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre'][month]
        filename = f'Sommaire_RJ_{year}-{month:02d}_{month_name_fr}.xlsx'

        return send_file(buf, as_attachment=True, download_name=filename,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        logger.error(f"Monthly export error: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# RJ EXCEL EXPORT — Fill template with session data
# ═══════════════════════════════════════════════════════════════════════════════

def excel_date_serial(d):
    """Convert Python date to Excel serial number (days since 1899-12-30)."""
    if d is None:
        return 0
    base = date(1899, 12, 30)
    delta = d - base
    return delta.days


@rj_native_bp.route('/api/rj/native/export/rj/<audit_date>', methods=['GET'])
@auth_required
def export_rj_excel(audit_date):
    """Export NightAuditSession as filled Rj Excel file.

    Uses the template Rj Vierge.xls and injects session data into mapped cells.
    Preserves all formatting and formulas. Returns .xls file download.
    """
    try:
        # Parse date
        try:
            d = datetime.strptime(audit_date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Format de date invalide (YYYY-MM-DD)'}), 400

        # Load session
        nas = NightAuditSession.query.filter_by(audit_date=d).first()
        if not nas:
            return jsonify({'error': 'Session non trouvée'}), 404

        # Find template
        template_path = 'RJ 2024-2025/RJ 2025-2026/12-Février 2026/Rj Vierge.xls'
        if not os.path.exists(template_path):
            return jsonify({'error': 'Template RJ not found'}), 500

        # Load template workbook with xlrd, copy with xlutils
        rb = xlrd.open_workbook(template_path, formatting_info=True)
        wb = xlutils_copy(rb)

        # Build sheet name to index mapping
        sheet_name_to_idx = {name: idx for idx, name in enumerate(rb.sheet_names())}

        # Helper to write value to sheet/cell
        def write_cell(sheet_name, row, col, value):
            """Write value to cell (0-indexed row/col)."""
            try:
                if sheet_name in sheet_name_to_idx:
                    idx = sheet_name_to_idx[sheet_name]
                    ws = wb.get_sheet(idx)
                    ws.write(row, col, value)
            except Exception as e:
                logger.warning(f"Could not write {sheet_name}!R{row+1}C{col+1}: {e}")

        # ───────────────────────────────────────────────────────────────
        # 1. DATE cells — Set date serial in all key sheets
        # ───────────────────────────────────────────────────────────────
        date_serial = excel_date_serial(d)

        write_cell('Recap', 0, 4, date_serial)  # E1
        write_cell('transelect', 4, 1, date_serial)  # B5
        write_cell('geac_ux', 21, 4, date_serial)  # E22 (0-indexed row 21 = Excel row 22)
        write_cell('controle', 27, 1, date_serial)  # B28
        write_cell('DUBACK#', 2, 1, date_serial)  # B3 (0-indexed row 2 = Excel row 3)
        write_cell('depot', 3, 1, date_serial)  # B4
        write_cell('SetD', 0, 0, date_serial)  # A1
        write_cell('Sonifi', 3, 1, date_serial)  # B4
        write_cell('Internet', 3, 0, date_serial)  # A4

        # ───────────────────────────────────────────────────────────────
        # 2. RECAP sheet — Cash reconciliation
        # ───────────────────────────────────────────────────────────────
        write_cell('Recap', 5, 1, nas.cash_ls_lecture or 0)  # B6
        write_cell('Recap', 5, 2, nas.cash_ls_corr or 0)    # C6
        write_cell('Recap', 6, 1, nas.cash_pos_lecture or 0)  # B7
        write_cell('Recap', 6, 2, nas.cash_pos_corr or 0)    # C7
        write_cell('Recap', 7, 1, nas.cheque_ar_lecture or 0)  # B8
        write_cell('Recap', 7, 2, nas.cheque_ar_corr or 0)    # C8
        write_cell('Recap', 8, 1, nas.cheque_dr_lecture or 0)  # B9
        write_cell('Recap', 8, 2, nas.cheque_dr_corr or 0)    # C9
        write_cell('Recap', 10, 1, nas.remb_gratuite_lecture or 0)  # B11
        write_cell('Recap', 10, 2, nas.remb_gratuite_corr or 0)    # C11
        write_cell('Recap', 11, 1, nas.remb_client_lecture or 0)  # B12
        write_cell('Recap', 11, 2, nas.remb_client_corr or 0)    # C12
        write_cell('Recap', 15, 1, nas.dueback_reception_lecture or 0)  # B16
        write_cell('Recap', 15, 2, nas.dueback_reception_corr or 0)    # C16
        write_cell('Recap', 16, 1, nas.dueback_nb_lecture or 0)  # B17
        write_cell('Recap', 16, 2, nas.dueback_nb_corr or 0)    # C17
        write_cell('Recap', 21, 1, nas.deposit_cdn or 0)  # B22
        write_cell('Recap', 23, 1, nas.deposit_us or 0)  # B24
        write_cell('Recap', 25, 1, nas.auditor_name or '')  # B26 (0-indexed row 25 = Excel row 26)

        # ───────────────────────────────────────────────────────────────
        # 3. TRANSELECT sheet — Restaurant section
        # ───────────────────────────────────────────────────────────────
        rest_data = nas.get_json('transelect_restaurant')
        card_rows = {
            'debit': 8, 'visa': 9, 'mc': 10, 'amex': 11, 'discover': 12
        }  # 0-indexed

        # Terminals should be in order; map to columns B-U (1-20, 0-19 in 0-index)
        terminal_cols = list(range(1, 21))  # B=1, C=2, ..., U=20
        terminal_names = sorted(rest_data.keys()) if rest_data else []

        for term_idx, term_name in enumerate(terminal_names[:20]):  # Max 20 terminals
            term_data = rest_data.get(term_name, {})
            col = terminal_cols[term_idx]

            for card_type, row in card_rows.items():
                value = term_data.get(card_type, 0) or 0
                write_cell('transelect', row, col, value)

        # ───────────────────────────────────────────────────────────────
        # 4. TRANSELECT sheet — Reception section
        # ───────────────────────────────────────────────────────────────
        recep_data = nas.get_json('transelect_reception')
        recep_card_rows = {
            'debit': 19, 'visa': 20, 'mc': 21, 'amex': 23, 'discover': 22
        }  # 0-indexed (rows 20-24 in Excel = 19-23 in 0-index)
        # Note: DISCOVER and AMEX appear to be swapped in some templates; adjust if needed

        # Reception has fixed columns: B=fusebox, C=term8, D=k053
        recep_cols = {'fusebox': 1, 'term8': 2, 'k053': 3}

        for card_type, row in recep_card_rows.items():
            ct_data = recep_data.get(card_type, {})
            for terminal, col in recep_cols.items():
                value = ct_data.get(terminal, 0) or 0
                write_cell('transelect', row, col, value)

        # ───────────────────────────────────────────────────────────────
        # 5. DUBACK# sheet — Receptionist due-back entries
        # ───────────────────────────────────────────────────────────────
        dueback_entries = nas.get_json('dueback_entries')
        if isinstance(dueback_entries, list):
            # Data starts at row 5 (0-indexed=4), odd rows only: 4, 6, 8, 10, ...
            data_rows = [4 + i*2 for i in range(len(dueback_entries))]  # 4, 6, 8, ...

            for idx, entry in enumerate(dueback_entries[:30]):  # Reasonable limit
                if idx >= len(data_rows):
                    break
                row = data_rows[idx]
                name = entry.get('name', '')
                previous = entry.get('previous', 0) or 0
                nouveau = entry.get('nouveau', 0) or 0

                write_cell('DUBACK#', row, 0, name)  # Col A: name
                write_cell('DUBACK#', row, 1, previous)  # Col B: previous
                write_cell('DUBACK#', row + 1, 1, nouveau)  # Col B (next row): nouveau

        # ───────────────────────────────────────────────────────────────
        # 6. GEAC_UX sheet — Card cashout amounts
        # ───────────────────────────────────────────────────────────────
        geac_cashout = nas.get_json('geac_cashout')
        geac_daily_rev = nas.get_json('geac_daily_rev')

        # Standard card types: amex, master, visa
        # Row 5 (0-indexed) = Excel row 6: Daily Cash Out
        write_cell('geac_ux', 5, 1, geac_cashout.get('amex', 0) or 0)  # B6
        write_cell('geac_ux', 5, 6, geac_cashout.get('master', 0) or 0)  # G6
        write_cell('geac_ux', 5, 9, geac_cashout.get('visa', 0) or 0)  # J6

        # Row 7 (0-indexed) = Excel row 8: Difference
        write_cell('geac_ux', 7, 1, geac_cashout.get('amex_variance', 0) or 0)  # B8
        write_cell('geac_ux', 7, 6, geac_cashout.get('master_variance', 0) or 0)  # G8
        write_cell('geac_ux', 7, 9, geac_cashout.get('visa_variance', 0) or 0)  # J8

        # ───────────────────────────────────────────────────────────────
        # 7. DEPOT sheet — Bank deposit amounts (client 6 & 8)
        # ───────────────────────────────────────────────────────────────
        depot_data = nas.get_json('depot_data')

        # Client 6: column B, starting row 10 (0-indexed=9)
        c6_amounts = depot_data.get('client6', {}).get('amounts', [])
        for idx, amt in enumerate(c6_amounts[:20]):  # Reasonable limit
            row = 9 + idx
            write_cell('depot', row, 1, amt or 0)  # Col B

        # Client 8: column F, starting row 10 (0-indexed=9)
        c8_amounts = depot_data.get('client8', {}).get('amounts', [])
        for idx, amt in enumerate(c8_amounts[:20]):
            row = 9 + idx
            write_cell('depot', row, 5, amt or 0)  # Col F

        # ───────────────────────────────────────────────────────────────
        # 8. SetD sheet — Personnel set-déjeuner entries
        # ───────────────────────────────────────────────────────────────
        setd_personnel = nas.get_json('setd_personnel')
        if isinstance(setd_personnel, list):
            for entry in setd_personnel:
                # entry format: {name, column_letter, amount}
                # For now, we'll skip individual personnel entries
                # as they require dynamic column mapping
                pass

        # ───────────────────────────────────────────────────────────────
        # 9. JOUR sheet — Daily F&B revenue and occupation data
        # ───────────────────────────────────────────────────────────────
        # The JOUR sheet is complex with 117 columns and 233 rows.
        # For MVP, fill only the summary/total row or a specific day.
        # This would require detailed column mapping from the template.
        # For now, skip individual jour entries.

        # ───────────────────────────────────────────────────────────────
        # 10. CONTROLE sheet — Metadata
        # ───────────────────────────────────────────────────────────────
        write_cell('controle', 25, 1, nas.auditor_name or '')  # B26

        # ───────────────────────────────────────────────────────────────
        # Save and return
        # ───────────────────────────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Format filename: Rj DD-MM-YYYY.xls
        filename = f"Rj {d.strftime('%d-%m-%Y')}.xls"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.ms-excel'
        )

    except Exception as e:
        logger.error(f"RJ Excel export error: {e}", exc_info=True)
        return jsonify({'error': f"Erreur lors de l'export: {str(e)}"}), 500
