"""
Blueprint for HP (Hotel Promotion / Administration) management.
Upload, view, edit HP Excel files — write entries directly back to the Excel.
Dual storage: Excel file (for compatibility) + SQLite DB (for persistence/ETL).
"""

from flask import Blueprint, request, jsonify, send_file, render_template
from routes.checklist import login_required
from database import db, HPPeriod, HPEntry
from openpyxl import load_workbook
from datetime import datetime
from io import BytesIO
import json, os, re

hp_bp = Blueprint('hp', __name__)

# ── Storage dir for uploaded HP files ──────────────────────────
HP_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'database', 'hp_files')
os.makedirs(HP_DIR, exist_ok=True)

# ── Constants ───────────────────────────────────────────────────
DONNEES_HEADERS_ROW = 12
DONNEES_DATA_START = 13

# Column mapping (1-based for openpyxl)
COL = {
    'day': 1, 'area': 2, 'nourriture': 3, 'boisson': 4, 'biere': 5,
    'vin': 6, 'mineraux': 7, 'tabagie': 8, 'autres': 9, 'pourboire': 10,
    'paiement': 11, 'total': 12, 'raison': 13, 'qui': 14, 'autoriser': 15,
}

AREAS = ['Piazza', 'Tabagie', 'Banquet', 'Link', 'Cupola', 'Serv Ch.']
PAYMENT_METHODS = [
    '14 - Administration',
    '15 - Hotel promotion',
    '17 - Promesse service',
    '500 - 50% Hot rate',
]

PRODUCT_FIELDS = ['nourriture', 'boisson', 'biere', 'vin', 'mineraux', 'autre', 'pourboire']


def _get_active_hp_path():
    meta_file = os.path.join(HP_DIR, 'active.json')
    if os.path.exists(meta_file):
        with open(meta_file) as f:
            meta = json.load(f)
        path = os.path.join(HP_DIR, meta.get('filename', ''))
        if os.path.exists(path):
            return path
    return None


def _set_active_hp(filename):
    meta_file = os.path.join(HP_DIR, 'active.json')
    with open(meta_file, 'w') as f:
        json.dump({'filename': filename, 'uploaded_at': datetime.now().isoformat()}, f)


def _safe_float(val):
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _parse_hp(filepath):
    """Parse HP xlsx and return structured data for the frontend."""
    wb = load_workbook(filepath, data_only=True)

    result = {
        'period': '',
        'available_days': [],
        'entries_by_day': {},
        'monthly_totals': {},
        'areas': AREAS,
        'payment_methods': PAYMENT_METHODS,
    }

    # Get period from mensuel sheet
    if 'mensuel' in wb.sheetnames:
        ws_m = wb['mensuel']
        period_cell = ws_m['D1'].value
        if period_cell:
            result['period'] = str(period_cell)

    # Parse données sheet
    ws = None
    for name in wb.sheetnames:
        if name.lower().strip() in ('données', 'donnees'):
            ws = wb[name]
            break

    if not ws:
        wb.close()
        return result

    entries_by_day = {}
    for r in range(DONNEES_DATA_START, ws.max_row + 1):
        day_val = ws.cell(r, COL['day']).value
        if day_val is None:
            continue
        try:
            day = int(day_val)
        except (ValueError, TypeError):
            continue

        area = str(ws.cell(r, COL['area']).value or '').strip()
        if not area:
            continue

        entry = {
            'row': r,
            'day': day,
            'area': area,
            'nourriture': _safe_float(ws.cell(r, COL['nourriture']).value),
            'boisson': _safe_float(ws.cell(r, COL['boisson']).value),
            'biere': _safe_float(ws.cell(r, COL['biere']).value),
            'vin': _safe_float(ws.cell(r, COL['vin']).value),
            'mineraux': _safe_float(ws.cell(r, COL['mineraux']).value),
            'autre': _safe_float(ws.cell(r, COL['tabagie']).value),
            'pourboire': _safe_float(ws.cell(r, COL['pourboire']).value),
            'paiement': str(ws.cell(r, COL['paiement']).value or ''),
            'total': _safe_float(ws.cell(r, COL['total']).value),
            'raison': str(ws.cell(r, COL['raison']).value or ''),
            'qui': str(ws.cell(r, COL['qui']).value or ''),
            'autorise_par': str(ws.cell(r, COL['autoriser']).value or ''),
        }

        day_key = str(day)
        if day_key not in entries_by_day:
            entries_by_day[day_key] = []
        entries_by_day[day_key].append(entry)

    result['entries_by_day'] = entries_by_day
    result['available_days'] = sorted([int(d) for d in entries_by_day.keys()])

    # Monthly totals from mensuel if available
    if 'mensuel' in wb.sheetnames:
        ws_m = wb['mensuel']
        result['monthly_totals'] = {
            'grand_total': _safe_float(ws_m['I11'].value),
        }

    wb.close()
    return result


def _extract_month_year(period_label):
    """Try to extract month/year from period label like 'Février 2026' or '2026-02'."""
    if not period_label:
        return None, None

    # French month names
    months_fr = {
        'janvier': 1, 'février': 2, 'mars': 3, 'avril': 4, 'mai': 5, 'juin': 6,
        'juillet': 7, 'août': 8, 'septembre': 9, 'octobre': 10, 'novembre': 11, 'décembre': 12
    }
    label_lower = period_label.lower().strip()
    for name, num in months_fr.items():
        if name in label_lower:
            year_match = re.search(r'20\d{2}', label_lower)
            year = int(year_match.group()) if year_match else datetime.now().year
            return num, year

    # Try YYYY-MM format
    m = re.match(r'(\d{4})-(\d{1,2})', period_label)
    if m:
        return int(m.group(2)), int(m.group(1))

    return None, None


def _sync_hp_to_db(parsed_data, source_filename=None):
    """Sync parsed HP data into SQLite. Creates or updates HPPeriod + HPEntry rows."""
    period_label = parsed_data.get('period', '')
    month, year = _extract_month_year(period_label)

    # Find or create period
    if month and year:
        period = HPPeriod.query.filter_by(month=month, year=year).first()
    else:
        # Fallback: match by filename
        period = HPPeriod.query.filter_by(source_filename=source_filename).first() if source_filename else None

    if not period:
        period = HPPeriod(
            period_label=period_label,
            month=month,
            year=year,
            source_filename=source_filename
        )
        db.session.add(period)
        db.session.flush()
    else:
        period.updated_at = datetime.utcnow()
        if source_filename:
            period.source_filename = source_filename
        if period_label:
            period.period_label = period_label

    # Delete existing entries for this period and re-insert
    HPEntry.query.filter_by(period_id=period.id).delete()

    entries_by_day = parsed_data.get('entries_by_day', {})
    for day_key, entries in entries_by_day.items():
        for e in entries:
            nourr = e.get('nourriture', 0)
            boisson = e.get('boisson', 0)
            biere = e.get('biere', 0)
            vin = e.get('vin', 0)
            mineraux = e.get('mineraux', 0)
            autre = e.get('autre', 0)
            pourboire = e.get('pourboire', 0)
            total = e.get('total', 0) or round(nourr + boisson + biere + vin + mineraux + autre + pourboire, 2)

            entry = HPEntry(
                period_id=period.id,
                day=int(day_key),
                area=e.get('area', ''),
                nourriture=nourr,
                boisson=boisson,
                biere=biere,
                vin=vin,
                mineraux=mineraux,
                autre=autre,
                pourboire=pourboire,
                paiement=e.get('paiement', ''),
                total=total,
                raison=e.get('raison', ''),
                qui=e.get('qui', ''),
                autorise_par=e.get('autorise_par', ''),
            )
            db.session.add(entry)

    db.session.commit()


# ── Routes ──────────────────────────────────────────────────────

@hp_bp.route('/hp')
@login_required
def hp_page():
    return render_template('hp.html')


@hp_bp.route('/api/hp/upload', methods=['POST'])
@login_required
def upload_hp():
    """Upload HP xlsx. Saves to disk + syncs to DB."""
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier fourni'}), 400

    f = request.files['file']
    if not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Format invalide. Utilisez un fichier .xlsx'}), 400

    filename = f'HP_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(HP_DIR, filename)
    f.save(filepath)
    _set_active_hp(filename)

    data = _parse_hp(filepath)

    # ★ Auto-sync to database
    _sync_hp_to_db(data, source_filename=filename)

    return jsonify({'success': True, 'data': data, 'filename': filename})


@hp_bp.route('/api/hp/load')
@login_required
def load_hp():
    path = _get_active_hp_path()
    if not path:
        return jsonify({'loaded': False, 'message': 'Aucun fichier HP actif'})

    data = _parse_hp(path)
    return jsonify({'loaded': True, 'data': data})


@hp_bp.route('/api/hp/save-day', methods=['POST'])
@login_required
def save_day():
    """
    Save HP entries for a specific day. Replaces all entries for that day.
    Writes to Excel + syncs to DB.
    """
    path = _get_active_hp_path()
    if not path:
        return jsonify({'error': 'Aucun fichier HP actif'}), 400

    data = request.get_json()
    day = data.get('day')
    entries = data.get('entries', [])

    if not day or not (1 <= int(day) <= 31):
        return jsonify({'error': 'Jour invalide'}), 400

    day = int(day)

    wb = load_workbook(path)

    # Find données sheet
    ws = None
    for name in wb.sheetnames:
        if name.lower().strip() in ('données', 'donnees'):
            ws = wb[name]
            break

    if not ws:
        wb.close()
        return jsonify({'error': "Feuille 'données' non trouvée"}), 400

    # 1) Remove existing rows for this day (bottom to top)
    rows_to_delete = []
    for r in range(DONNEES_DATA_START, ws.max_row + 1):
        cell_val = ws.cell(r, COL['day']).value
        if cell_val is not None:
            try:
                if int(cell_val) == day:
                    rows_to_delete.append(r)
            except (ValueError, TypeError):
                pass

    for r in reversed(rows_to_delete):
        ws.delete_rows(r)

    # 2) Find insertion point (after last row with day < target)
    insert_row = DONNEES_DATA_START
    for r in range(DONNEES_DATA_START, ws.max_row + 1):
        cell_val = ws.cell(r, COL['day']).value
        if cell_val is not None:
            try:
                if int(cell_val) < day:
                    insert_row = r + 1
                elif int(cell_val) > day:
                    break
            except (ValueError, TypeError):
                pass

    # 3) Insert new entries
    for i, entry in enumerate(entries):
        r = insert_row + i
        ws.insert_rows(r)

        nourr = float(entry.get('nourriture', 0) or 0)
        boisson = float(entry.get('boisson', 0) or 0)
        biere = float(entry.get('biere', 0) or 0)
        vin = float(entry.get('vin', 0) or 0)
        mineraux = float(entry.get('mineraux', 0) or 0)
        autre = float(entry.get('autre', 0) or 0)
        pourboire = float(entry.get('pourboire', 0) or 0)
        total = round(nourr + boisson + biere + vin + mineraux + autre + pourboire, 2)

        ws.cell(r, COL['day'], day)
        ws.cell(r, COL['area'], entry.get('area', ''))
        ws.cell(r, COL['nourriture'], nourr)
        ws.cell(r, COL['boisson'], boisson)
        ws.cell(r, COL['biere'], biere)
        ws.cell(r, COL['vin'], vin)
        ws.cell(r, COL['mineraux'], mineraux)
        ws.cell(r, COL['tabagie'], autre)
        ws.cell(r, COL['pourboire'], pourboire)
        ws.cell(r, COL['paiement'], entry.get('paiement', ''))
        ws.cell(r, COL['total'], total)
        ws.cell(r, COL['raison'], entry.get('raison', ''))
        ws.cell(r, COL['qui'], entry.get('qui', ''))
        ws.cell(r, COL['autoriser'], entry.get('autorise_par', ''))

    wb.save(path)
    wb.close()

    # Re-read and sync to DB
    updated = _parse_hp(path)
    _sync_hp_to_db(updated)

    return jsonify({'success': True, 'data': updated})


@hp_bp.route('/api/hp/download')
@login_required
def download_hp():
    path = _get_active_hp_path()
    if not path:
        return jsonify({'error': 'Aucun fichier HP actif'}), 404

    return send_file(
        path,
        as_attachment=True,
        download_name=os.path.basename(path),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ── ETL: Historique HP ─────────────────────────────────────────
@hp_bp.route('/api/hp/history')
@login_required
def hp_history():
    """Return all HP periods and their entry counts for ETL/reporting."""
    periods = HPPeriod.query.order_by(HPPeriod.year.desc(), HPPeriod.month.desc()).all()
    return jsonify({
        'periods': [p.to_dict() for p in periods],
        'total_entries': HPEntry.query.count(),
    })


@hp_bp.route('/api/hp/history/<int:period_id>')
@login_required
def hp_period_detail(period_id):
    """Return all entries for a specific HP period."""
    period = HPPeriod.query.get_or_404(period_id)
    entries = HPEntry.query.filter_by(period_id=period.id).order_by(HPEntry.day, HPEntry.area).all()
    return jsonify({
        'period': period.to_dict(),
        'entries': [e.to_dict() for e in entries],
    })
