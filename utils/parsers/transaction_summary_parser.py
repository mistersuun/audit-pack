"""
FreedomPay/POS TransactionSummarybyCardType Parser.

Parses the FreedomPay "TransactionSummarybyCardType" Excel report (.xlsx).

File structure:
- Sheet: TransactionSummarybyCardType
- Row 4 headers: Store | Card Type | Business Date | Account Number |
                 Card Holder Name | Rid | Invoice Number | Trans Type |
                 Trans Amount | Currency Code
- Data rows: Individual card transactions
- Subtotal rows: "Total:Amex - Credit", "Total:Visa - Credit", etc.
- Grand total row: "YULLS PMS" with final grand total amount

Extracts:
- Per card type: transaction count, total amount
- Card types: Amex, Visa, Mastercard, Debit/Interac, Discover
- Grand total
- Business date
"""

import io
from openpyxl import load_workbook
from utils.parsers.base_parser import BaseParser


class TransactionSummaryParser(BaseParser):
    """
    Parse FreedomPay TransactionSummarybyCardType Excel report.

    Maps card type totals to transelect reception fields:
    - amex_total -> transelect_rec_amex
    - visa_total -> transelect_rec_visa
    - mc_total -> transelect_rec_mc
    - debit_total -> transelect_rec_debit
    - discover_total -> transelect_rec_discover
    - grand_total -> transelect_grand_total
    - business_date -> transaction_date
    """

    FIELD_MAPPINGS = {
        'amex_total': 'transelect_rec_amex',
        'visa_total': 'transelect_rec_visa',
        'mc_total': 'transelect_rec_mc',
        'debit_total': 'transelect_rec_debit',
        'discover_total': 'transelect_rec_discover',
        'grand_total': 'transelect_grand_total',
        'business_date': 'transaction_date',
    }

    # Card type keywords to look for in subtotal rows
    CARD_TYPE_KEYWORDS = {
        'amex': ['amex', 'american express'],
        'visa': ['visa'],
        'mc': ['mastercard', 'master card'],
        'debit': ['debit', 'interac'],
        'discover': ['discover', 'diners'],
    }

    def __init__(self, file_bytes, filename=None, **kwargs):
        """
        Initialize TransactionSummarybyCardType parser.

        Args:
            file_bytes: Excel file bytes
            filename: Optional filename
        """
        super().__init__(file_bytes, filename, **kwargs)
        self.sheet_name = 'TransactionSummarybyCardType'
        self.header_row = None
        self.col_indices = {}  # {column_name: index}

    def parse(self):
        """Parse the TransactionSummarybyCardType Excel file."""
        try:
            self._parse_file()
        except Exception as e:
            self.validation_errors.append(f"Error parsing TransactionSummarybyCardType: {str(e)}")
            self.confidence = 0.0

        self._parsed = True

    def _parse_file(self):
        """Read and parse the Excel workbook."""
        file_stream = io.BytesIO(self.file_bytes)
        wb = load_workbook(file_stream, data_only=True)

        # Try to find the correct sheet
        if self.sheet_name in wb.sheetnames:
            ws = wb[self.sheet_name]
        else:
            ws = wb.active
            if ws.title != self.sheet_name:
                self.validation_warnings.append(
                    f"Expected sheet '{self.sheet_name}' but found '{ws.title}'. Parsing anyway."
                )

        # Find header row (row 4 typically, but search for "Card Type")
        self._find_header_row(ws)

        if self.header_row is None:
            self.validation_errors.append(
                "Could not find header row with 'Card Type' column. "
                "Expected format: Store | Card Type | Business Date | ..."
            )
            self.confidence = 0.0
            return

        # Parse data rows and extract subtotals + grand total
        self._extract_data(ws)

        if self.extracted_data:
            self.confidence = 0.9
        else:
            self.validation_errors.append("No transaction data extracted from file")
            self.confidence = 0.0

    def _find_header_row(self, ws):
        """Find the header row containing actual column headers (Store, Card Type, Trans Amount, etc.)."""
        for row_idx in range(1, min(10, ws.max_row + 1)):
            row_values = [cell.value for cell in ws[row_idx]]
            row_text = ' '.join(str(v or '') for v in row_values).lower()

            # Look for a row that has multiple expected headers, not just a title containing "Card Type"
            header_count = 0
            for col_val in row_values:
                if col_val and isinstance(col_val, str):
                    col_lower = col_val.lower()
                    if any(h in col_lower for h in ['store', 'card type', 'trans amount', 'business date', 'account number']):
                        header_count += 1

            # If this row has at least 3 expected header columns, it's likely the header row
            if header_count >= 3:
                self.header_row = row_idx
                # Map column names to their indices
                self._map_columns(ws, row_idx)
                break

    def _map_columns(self, ws, header_row):
        """Map column names to their indices based on header row."""
        expected_columns = [
            'Store', 'Card Type', 'Business Date', 'Account Number',
            'Card Holder Name', 'Rid', 'Invoice Number', 'Trans Type',
            'Trans Amount', 'Currency Code'
        ]

        row_values = [cell.value for cell in ws[header_row]]

        for col_name in expected_columns:
            for idx, val in enumerate(row_values):
                if val and col_name.lower() in str(val).lower():
                    self.col_indices[col_name] = idx
                    break

    def _extract_data(self, ws):
        """Extract transaction data, subtotals, and grand total."""
        if not self.header_row or 'Card Type' not in self.col_indices:
            return

        card_type_idx = self.col_indices.get('Card Type', 1)
        trans_amount_idx = self.col_indices.get('Trans Amount', 8)
        business_date_idx = self.col_indices.get('Business Date', 2)

        # Initialize totals dict
        totals = {
            'amex': 0.0,
            'visa': 0.0,
            'mc': 0.0,
            'debit': 0.0,
            'discover': 0.0,
        }
        transaction_counts = {k: 0 for k in totals.keys()}
        grand_total = 0.0
        first_business_date = None

        # Iterate through data rows
        for row_idx in range(self.header_row + 1, ws.max_row + 1):
            row = ws[row_idx]
            store_cell = row[0].value
            card_type_cell = row[card_type_idx].value if card_type_idx < len(row) else None
            trans_amount_cell = row[trans_amount_idx].value if trans_amount_idx < len(row) else None
            business_date_cell = row[business_date_idx].value if business_date_idx < len(row) else None

            # Extract business date from first data row
            if first_business_date is None and business_date_cell:
                first_business_date = business_date_cell

            # Check if this is a subtotal row
            if card_type_cell and isinstance(card_type_cell, str):
                subtotal_match = self._match_subtotal_row(card_type_cell)
                if subtotal_match:
                    # Extract amount from this subtotal row
                    amount = self._safe_float(trans_amount_cell, 0.0)
                    totals[subtotal_match] = amount
                    continue

            # Check if this is the grand total row ("YULLS PMS")
            if store_cell and isinstance(store_cell, str) and 'YULLS PMS' in store_cell.upper():
                amount = self._safe_float(trans_amount_cell, 0.0)
                if amount > 0:
                    grand_total = amount
                continue

            # Regular data row: accumulate by card type
            if card_type_cell and isinstance(card_type_cell, str):
                card_key = self._match_card_type(card_type_cell)
                if card_key:
                    amount = self._safe_float(trans_amount_cell, 0.0)
                    if amount > 0:
                        totals[card_key] += amount
                        transaction_counts[card_key] += 1

        # Store extracted data
        self.extracted_data['amex_total'] = totals.get('amex', 0.0)
        self.extracted_data['visa_total'] = totals.get('visa', 0.0)
        self.extracted_data['mc_total'] = totals.get('mc', 0.0)
        self.extracted_data['debit_total'] = totals.get('debit', 0.0)
        self.extracted_data['discover_total'] = totals.get('discover', 0.0)

        # Use grand_total if found, otherwise sum all card types
        if grand_total > 0:
            self.extracted_data['grand_total'] = grand_total
        else:
            self.extracted_data['grand_total'] = sum(totals.values())

        if first_business_date:
            self.extracted_data['business_date'] = first_business_date

        # Store metadata
        self.extracted_data['transaction_counts'] = transaction_counts
        self.extracted_data['source'] = 'TransactionSummarybyCardType'
        self.extracted_data['sheet_name'] = self.sheet_name

    def _match_subtotal_row(self, text):
        """
        Check if text is a subtotal row (e.g., "Total:Amex - Credit").

        Returns:
            str: Card type key ('amex', 'visa', 'mc', 'debit', 'discover') or None
        """
        text_lower = text.lower()

        if 'total:' in text_lower:
            for card_key, keywords in self.CARD_TYPE_KEYWORDS.items():
                for keyword in keywords:
                    if keyword in text_lower:
                        return card_key
        return None

    def _match_card_type(self, text):
        """
        Match card type keywords in text.

        Returns:
            str: Card type key ('amex', 'visa', 'mc', 'debit', 'discover') or None
        """
        text_lower = text.lower()

        for card_key, keywords in self.CARD_TYPE_KEYWORDS.items():
            for keyword in keywords:
                if keyword in text_lower:
                    return card_key
        return None

    def validate(self):
        """Validate extracted data."""
        if not self.extracted_data:
            self.validation_errors.append("No data extracted from TransactionSummarybyCardType")
            return False

        grand_total = self.extracted_data.get('grand_total', 0)
        if grand_total <= 0:
            self.validation_errors.append(f"Grand total is invalid: {grand_total}")
            return False

        # Verify that at least some card types have amounts
        card_totals = [
            self.extracted_data.get('amex_total', 0),
            self.extracted_data.get('visa_total', 0),
            self.extracted_data.get('mc_total', 0),
            self.extracted_data.get('debit_total', 0),
            self.extracted_data.get('discover_total', 0),
        ]

        if sum(card_totals) == 0:
            self.validation_errors.append("No card transaction totals found")
            return False

        # Verify grand total matches sum of card types (within tolerance)
        sum_cards = sum(card_totals)
        variance = abs(sum_cards - grand_total)
        tolerance = 0.01  # Allow $0.01 rounding difference

        if variance > tolerance:
            self.validation_warnings.append(
                f"Grand total variance: ${sum_cards:.2f} (cards) vs "
                f"${grand_total:.2f} (grand) = ${variance:.2f}"
            )

        return len(self.validation_errors) == 0

    def get_fillable_data(self):
        """
        Return data ready to fill transelect reception fields.

        Returns:
            dict: {field_name: value} for transelect reception section
        """
        if not self._parsed:
            self.parse()

        fillable = {}
        for field_key, cell_ref in self.FIELD_MAPPINGS.items():
            val = self.extracted_data.get(field_key)
            if val is not None and (isinstance(val, str) or val != 0):
                fillable[cell_ref] = val

        return fillable
