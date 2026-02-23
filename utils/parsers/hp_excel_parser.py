"""
HP Excel Parser.
Extracts Honor Paid (HP) deductions from HP Excel file.
Supports both monthly aggregation (mensuel sheet) and daily extraction (données sheet).

The HP file tracks comped meals / promotional charges by department and product category.
Night auditors need the daily HP totals per department to deduct from Jour sheet columns.
"""

import io
from openpyxl import load_workbook
from utils.parsers.base_parser import BaseParser


class HPExcelParser(BaseParser):
    """
    Parse HP Excel file for Honor Paid deductions.

    Target: jour sheet (HP deduction columns J-S)

    Sheets used:
    - mensuel: Monthly aggregation (department × product matrix + payment methods + GL)
    - données: Raw per-transaction data with day, department, product breakdowns
    - Journalier: Daily summary for a specific day

    Daily HP data maps to Jour columns:
    - J (9):  Piazza Nourriture
    - K (10): Piazza Alcool (Boisson)
    - L (11): Piazza Bières
    - M (12): Piazza Non Alcool (Minéraux)
    - N (13): Piazza Vins
    - O (14): Banquet Nourriture
    - P (15): Banquet Boissons
    - S (18): Tabagie
    """

    # Area names in the données sheet → Jour column category mapping
    # Jour sheet column layout (0-indexed):
    #  9: Piazza Nourriture   14: Spesa Nourriture   19: Chambres Svc Nourr   24: Banquet Nourr
    # 10: Piazza Boisson      15: Spesa Boisson      20: Chambres Svc Boisson 25: Banquet Boisson
    # 11: Piazza Bières       16: Spesa Bières       21: Chambres Svc Bières  26: Banquet Bières
    # 12: Piazza Minéraux     17: Spesa Minéraux     22: Chambres Svc Min.    27: Banquet Minéraux
    # 13: Piazza Vins         18: Spesa Vins         23: Chambres Svc Vins    28: Banquet Vins
    # 35: Tabagie
    AREA_PRODUCT_TO_JOUR = {
        # (area, product_column) → jour_column_index
        ('Piazza', 'nourriture'): 9,
        ('Piazza', 'boisson'): 10,
        ('Piazza', 'biere'): 11,
        ('Piazza', 'mineraux'): 12,
        ('Piazza', 'vin'): 13,
        ('Banquet', 'nourriture'): 24,
        ('Banquet', 'boisson'): 25,
        ('Banquet', 'biere'): 26,
        ('Banquet', 'mineraux'): 27,
        ('Banquet', 'vin'): 28,
        # Tabagie: ONLY the 'tabagie' product column maps to jour col 35.
        # Food items bought at the Tabagie (nourriture) are NOT deducted from jour_tabagie.
        ('Tabagie', 'tabagie'): 35,
        ('Tabagie', 'nourriture'): 14,  # Tabagie food → deduct from Spesa Nourriture
        ('Link', 'nourriture'): 9,      # Link food grouped with Piazza Nourr
        ('Cupola', 'nourriture'): 9,    # Cupola food grouped with Piazza Nourr
        ('Serv Ch.', 'nourriture'): 19, # Service Chambres food
    }

    # Données sheet column mapping (0-indexed from A)
    DONNEES_COLS = {
        'day': 0,           # A: Day number (1-31)
        'area': 1,          # B: Department (Piazza, Tabagie, etc.)
        'nourriture': 2,    # C: Food
        'boisson': 3,       # D: Alcohol/Beverages
        'biere': 4,         # E: Beer
        'vin': 5,           # F: Wine
        'mineraux': 6,      # G: Non-alcoholic/Minerals
        'tabagie': 7,       # H: Tobacco (product)
        'autres': 8,        # I: Other
        'pourboire': 9,     # J: Tips
        'paiement': 10,     # K: Payment method
        'total': 11,        # L: Transaction total
        'raison': 12,       # M: Reason
        'qui': 13,          # N: Who
        'autoriser': 14,    # O: Authorized by
    }

    PRODUCT_CATEGORIES = ['nourriture', 'boisson', 'biere', 'vin', 'mineraux', 'tabagie', 'autres']

    FIELD_MAPPINGS = {
        # Monthly aggregation fields (from mensuel sheet)
        'grand_total': 'I11',
        'pmt_total': 'I20',
        # Daily fields are dynamic (set during parse_daily)
    }

    def __init__(self, file_bytes, filename=None, day=None):
        super().__init__(file_bytes, filename)
        self.workbook = None
        self.period = None
        self.target_day = day
        self.daily_data = {}       # {day: {area: {product: amount}}}
        self.daily_totals = {}     # {day: {area: total}}
        self.jour_deductions = {}  # {jour_col_index: hp_deduction_amount}
        self.available_days = []

    def parse(self):
        """Parse HP Excel file — extracts both monthly and daily data."""
        try:
            file_stream = io.BytesIO(self.file_bytes)
            wb = load_workbook(file_stream, data_only=True)
            self.workbook = wb

            # --- Parse mensuel sheet (monthly totals) ---
            if 'mensuel' in wb.sheetnames:
                self._parse_mensuel(wb['mensuel'])

            # --- Parse données sheet (per-day transactions) ---
            if 'données' in wb.sheetnames:
                self._parse_donnees(wb['données'])
            elif 'donnees' in wb.sheetnames:
                self._parse_donnees(wb['donnees'])

            # --- If a target day was specified, compute jour deductions ---
            if self.target_day is not None:
                self._compute_jour_deductions(self.target_day)

            # Calculate confidence
            self.confidence = self._calculate_confidence()
            self._parsed = True

        except Exception as e:
            self.validation_errors.append(f"Erreur parsing HP Excel: {str(e)}")
            self.confidence = 0.0
            self._parsed = True

    def _parse_mensuel(self, ws):
        """Parse mensuel sheet for monthly aggregation."""
        # Extract period
        period_cell = ws['D1'].value
        if period_cell:
            self.period = str(period_cell)
            self.extracted_data['period'] = self.period

        # Department × product matrix (rows 5-10)
        departments = {
            'Link': 5, 'Cupola': 6, 'Piazza': 7,
            'Banquet': 8, 'Serv Ch.': 9, 'Tabagie': 10
        }
        col_map = {
            'nourr': 'B', 'boisson': 'C', 'biere': 'D',
            'vin': 'E', 'min': 'F', 'pourb': 'G', 'tabagie': 'H'
        }

        for dept, row in departments.items():
            dept_key = dept.lower().replace(' ', '').replace('.', '')
            for cat, col in col_map.items():
                value = self._safe_float(ws[f"{col}{row}"].value)
                self.extracted_data[f"{dept_key}_{cat}"] = value

        # Department totals (row 11)
        for cat, col in col_map.items():
            value = self._safe_float(ws[f"{col}11"].value)
            self.extracted_data[f"total_{cat}"] = value

        # Grand total (I11) and control total (I12)
        self.extracted_data['grand_total'] = self._safe_float(ws['I11'].value)
        self.extracted_data['control_total'] = self._safe_float(ws['I12'].value)

        # Payment method totals (rows 16-20)
        payment_methods = {
            'admin14': 16, 'promo15': 17, 'promesse17': 18, 'hotrate500': 19
        }
        for pmt, row in payment_methods.items():
            total = self._safe_float(ws[f"I{row}"].value)
            self.extracted_data[f"{pmt}_total"] = total
        self.extracted_data['pmt_total'] = self._safe_float(ws['I20'].value)

        # GL accounts (rows 34-43)
        gl_accounts = {
            944000: 34, 937150: 35, 406830: 36, 422000: 37,
            502650: 38, 432050: 39, 432051: 40, 432052: 41, 432053: 42
        }
        for gl, row in gl_accounts.items():
            value = self._safe_float(ws[f"G{row}"].value)
            self.extracted_data[f"gl_{gl}"] = value
        self.extracted_data['gl_total'] = self._safe_float(ws['G43'].value)

    def _parse_donnees(self, ws):
        """Parse données sheet for per-day transaction data."""
        data_start_row = 13  # Row 12 is headers, data starts at 13
        cols = self.DONNEES_COLS

        for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, values_only=False):
            day_cell = row[cols['day']].value
            if day_cell is None:
                continue

            try:
                day = int(day_cell)
            except (ValueError, TypeError):
                continue

            area = str(row[cols['area']].value or '').strip()
            if not area:
                continue

            # Initialize nested dicts
            if day not in self.daily_data:
                self.daily_data[day] = {}
                self.daily_totals[day] = {}
            if area not in self.daily_data[day]:
                self.daily_data[day][area] = {cat: 0.0 for cat in self.PRODUCT_CATEGORIES}
                self.daily_totals[day][area] = 0.0

            # Accumulate product amounts
            for cat in self.PRODUCT_CATEGORIES:
                col_idx = cols.get(cat)
                if col_idx is not None and col_idx < len(row):
                    value = self._safe_float(row[col_idx].value)
                    self.daily_data[day][area][cat] += value

            # Accumulate line total
            total_val = self._safe_float(row[cols['total']].value)
            self.daily_totals[day][area] += total_val

        self.available_days = sorted(self.daily_data.keys())
        self.extracted_data['available_days'] = self.available_days
        self.extracted_data['total_transactions'] = sum(
            len(areas) for areas in self.daily_data.values()
        )

    def _compute_jour_deductions(self, day):
        """Compute Jour sheet HP deductions for a specific day.

        Aggregates per-department, per-product-category amounts from the
        données sheet and maps them to Jour sheet column indices.

        Args:
            day: int, the day of month (1-31)

        Sets self.jour_deductions = {col_index: deduction_amount}
        """
        self.jour_deductions = {}

        if day not in self.daily_data:
            self.extracted_data[f'day_{day}_total'] = 0.0
            self.extracted_data[f'day_{day}_deductions'] = {}
            return

        day_data = self.daily_data[day]

        for area, products in day_data.items():
            for product, amount in products.items():
                if amount == 0:
                    continue
                key = (area, product)
                if key in self.AREA_PRODUCT_TO_JOUR:
                    col_idx = self.AREA_PRODUCT_TO_JOUR[key]
                    self.jour_deductions[col_idx] = self.jour_deductions.get(col_idx, 0.0) + amount

        # Also add tips (pourboire) to the appropriate columns
        # Tips follow the same department mapping as nourriture
        for area, products in day_data.items():
            tip_key = (area, 'nourriture')  # Tips follow food department
            if tip_key in self.AREA_PRODUCT_TO_JOUR:
                # Tips are in the pourboire field, already parsed as part of the total
                pass  # Tips are included in the line total, not separately deducted

        # Store in extracted_data for API responses
        day_total = sum(self.daily_totals.get(day, {}).values())
        self.extracted_data[f'day_{day}_total'] = day_total
        self.extracted_data[f'day_{day}_departments'] = self.daily_totals.get(day, {})
        self.extracted_data[f'day_{day}_deductions'] = {
            str(k): round(v, 2) for k, v in self.jour_deductions.items()
        }
        self.extracted_data['jour_deductions'] = {
            str(k): round(v, 2) for k, v in self.jour_deductions.items()
        }

    def get_daily_deductions(self, day):
        """Get HP deductions for a specific day, mapped to Jour column indices.

        Args:
            day: int, the day of month (1-31)

        Returns:
            dict: {jour_column_index: deduction_amount}
            Example: {9: 36.25, 18: 96.80} meaning
                     J (Piazza Nourr) = 36.25, S (Tabagie) = 96.80
        """
        if not self._parsed:
            self.parse()

        if day not in self.daily_data:
            return {}

        # Recompute for the requested day
        self._compute_jour_deductions(day)
        return dict(self.jour_deductions)

    def get_daily_summary(self, day):
        """Get a human-readable summary of HP deductions for a specific day.

        Returns:
            dict with area totals and product breakdowns
        """
        if not self._parsed:
            self.parse()

        if day not in self.daily_data:
            return {'day': day, 'found': False, 'areas': {}, 'total': 0.0}

        areas = {}
        for area, products in self.daily_data[day].items():
            area_total = self.daily_totals[day].get(area, 0.0)
            non_zero_products = {k: v for k, v in products.items() if v != 0}
            areas[area] = {
                'products': non_zero_products,
                'total': round(area_total, 2)
            }

        return {
            'day': day,
            'found': True,
            'areas': areas,
            'total': round(sum(self.daily_totals[day].values()), 2)
        }

    def _calculate_confidence(self):
        """Calculate confidence based on data completeness."""
        confidence = 0.0

        # Check if we found the données sheet with data
        if self.daily_data:
            confidence += 0.4  # Main data source found
            if len(self.available_days) >= 1:
                confidence += 0.1  # Has at least one day of data

        # Check mensuel sheet data
        grand_total = self.extracted_data.get('grand_total', 0)
        pmt_total = self.extracted_data.get('pmt_total', 0)
        control_total = self.extracted_data.get('control_total', 0)

        # Use the best available total for validation
        best_total = max(grand_total, pmt_total, control_total)
        if best_total > 0:
            confidence += 0.2  # Monthly totals found

        # Check if period was found
        if self.period:
            confidence += 0.1

        # Check if target day has data
        if self.target_day and self.target_day in self.daily_data:
            confidence += 0.2  # Target day found in data

        # Minimum confidence if we at least found the mensuel sheet
        if 'period' in self.extracted_data and confidence < 0.3:
            confidence = 0.3

        return min(confidence, 1.0)

    def validate(self):
        """Validate extracted HP data."""
        if not self.extracted_data or not self._parsed:
            self.validation_errors.append("Aucune donnée extraite du fichier HP")
            return False

        # Check for données sheet data
        if not self.daily_data:
            self.validation_warnings.append(
                "Feuille 'données' absente ou vide — impossible d'extraire les HP quotidiens"
            )

        # If a target day was set, check if it exists
        if self.target_day is not None and self.target_day not in self.daily_data:
            available = ', '.join(str(d) for d in self.available_days[:10])
            self.validation_warnings.append(
                f"Jour {self.target_day} non trouvé dans le fichier HP. "
                f"Jours disponibles: {available}"
            )

        # Cross-validate: sum of daily totals should be close to monthly total
        if self.daily_data:
            daily_grand_total = sum(
                sum(areas.values())
                for areas in self.daily_totals.values()
            )
            monthly_total = max(
                self.extracted_data.get('pmt_total', 0),
                self.extracted_data.get('control_total', 0),
                self.extracted_data.get('grand_total', 0),
            )
            if monthly_total > 0 and daily_grand_total > 0:
                variance = abs(daily_grand_total - monthly_total) / monthly_total
                if variance > 0.05:  # More than 5% variance
                    self.validation_warnings.append(
                        f"Écart entre total quotidien ({daily_grand_total:.2f}) "
                        f"et total mensuel ({monthly_total:.2f}): {variance:.1%}"
                    )

        return len(self.validation_errors) == 0
