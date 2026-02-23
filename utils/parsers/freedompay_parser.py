"""
FreedomPay Auto-Fill Parser.

FreedomPay bank-side card settlement amounts should equal Daily Revenue
system totals (variance = $0.00). Rather than parsing a separate file,
this parser can work in two modes:

1. Auto-fill mode (no file): Uses Daily Revenue card totals to fill
   GEAC/UX Row 6 (Daily Cash Out) — since FreedomPay = Daily Revenue.

2. File parse mode (with file): When user provides an actual FreedomPay
   export, parse it and compare with Daily Revenue for variance checking.

Data flow:
    FreedomPay bank settlement → geac_ux Row 6 (Daily Cash Out)
    Daily Revenue system totals → geac_ux Row 12 (Daily Revenue)
    Variance = Row 6 - Row 12 = should be $0.00

    geac_ux Row 6 also feeds → transelect fusebox rows (B21-B24)
    → Quasimodo REC computation
"""

import io
from utils.parsers.base_parser import BaseParser


class FreedomPayParser(BaseParser):
    """
    Parse FreedomPay data or auto-fill from Daily Revenue.

    Target sheet: geac_ux (Row 6 — Daily Cash Out)

    Fields extracted:
    - amex_cash_out: Amex daily settlement (B6)
    - diners_cash_out: Diners daily settlement (E6)
    - master_cash_out: MasterCard daily settlement (G6)
    - visa_cash_out: Visa daily settlement (J6)
    - discover_cash_out: Discover daily settlement (K6)

    Also fills transelect fusebox rows:
    - fusebox_visa: B21
    - fusebox_master: B22
    - fusebox_amex: B24
    """

    # GEAC/UX Row 6 (Daily Cash Out) cell references
    FIELD_MAPPINGS = {
        'amex_cash_out': 'B6',
        'diners_cash_out': 'E6',
        'master_cash_out': 'G6',
        'visa_cash_out': 'J6',
        'discover_cash_out': 'K6',
    }

    # GEAC/UX Row 12 (Daily Revenue) cell references
    DAILY_REV_MAPPINGS = {
        'amex_daily_revenue': 'B12',
        'diners_daily_revenue': 'E12',
        'master_daily_revenue': 'G12',
        'visa_daily_revenue': 'J12',
        'discover_daily_revenue': 'K12',
    }

    # Transelect fusebox rows (bank report)
    TRANSELECT_MAPPINGS = {
        'fusebox_visa': 'B21',
        'fusebox_master': 'B22',
        'fusebox_amex': 'B24',
    }

    # Map from Daily Revenue field names to our field names
    DAILY_REV_TO_CASHOUT = {
        # AR Summary / Daily Revenue card type → our field
        'visa': 'visa_cash_out',
        'mastercard': 'master_cash_out',
        'amex': 'amex_cash_out',
        'diners': 'diners_cash_out',
        'discover': 'discover_cash_out',
    }

    def __init__(self, file_bytes=None, filename=None, daily_revenue_cards=None, **kwargs):
        """
        Initialize FreedomPay parser.

        Args:
            file_bytes: Optional FreedomPay export file bytes
            filename: Optional filename
            daily_revenue_cards: Optional dict of {card_type: amount} from
                Daily Revenue for auto-fill mode. Example:
                {'visa': 5000, 'mastercard': 3000, 'amex': 2000}
        """
        super().__init__(file_bytes or b'', filename, **kwargs)
        self.daily_revenue_cards = daily_revenue_cards or {}
        self.mode = 'file' if file_bytes and len(file_bytes) > 0 else 'autofill'

    def parse(self):
        """Parse FreedomPay data — either from file or auto-fill from Daily Revenue."""
        if self.mode == 'file' and self.file_bytes and len(self.file_bytes) > 10:
            self._parse_file()
        elif self.daily_revenue_cards:
            self._parse_from_daily_revenue()
        else:
            self.validation_warnings.append(
                "Aucun fichier FreedomPay et aucune donnée Daily Revenue. "
                "Fournir un fichier Excel ou les montants par carte du Daily Revenue."
            )
            self.confidence = 0.0

        self._parsed = True

    def _parse_file(self):
        """
        Parse an actual FreedomPay/Fusebox export file.

        Supports both .xls (xlrd) and .xlsx (openpyxl) formats.
        The FreedomPay export typically has a single sheet with
        card type amounts in a summary section.
        """
        try:
            # Try openpyxl first (.xlsx)
            from openpyxl import load_workbook
            file_stream = io.BytesIO(self.file_bytes)
            try:
                wb = load_workbook(file_stream, data_only=True)
                ws = wb.active
                self._extract_from_xlsx(ws)
                return
            except Exception:
                pass

            # Fall back to xlrd (.xls)
            try:
                import xlrd
                wb = xlrd.open_workbook(file_contents=self.file_bytes)
                ws = wb.sheet_by_index(0)
                self._extract_from_xls(ws)
                return
            except Exception:
                pass

            # Neither worked
            self.validation_warnings.append(
                "Format de fichier FreedomPay non reconnu. "
                "Utilisation du mode auto-fill depuis Daily Revenue."
            )
            if self.daily_revenue_cards:
                self._parse_from_daily_revenue()
            else:
                self.confidence = 0.0

        except Exception as e:
            self.validation_errors.append(f"Erreur parsing FreedomPay: {str(e)}")
            self.confidence = 0.0

    def _extract_from_xlsx(self, ws):
        """Extract card settlement amounts from xlsx worksheet.

        Scans for card type keywords and their associated amounts.
        """
        found = {}

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            row_text = ' '.join(str(c.value or '') for c in row).lower()

            for card_type, field_name in self.DAILY_REV_TO_CASHOUT.items():
                if card_type.lower() in row_text:
                    # Find the first numeric value in this row
                    for cell in row:
                        if isinstance(cell.value, (int, float)) and cell.value != 0:
                            found[field_name] = abs(float(cell.value))
                            break

        if found:
            self.extracted_data.update(found)
            self.extracted_data['source'] = 'file'
            self.confidence = min(0.5 + 0.1 * len(found), 0.9)
        else:
            self.validation_warnings.append("Aucun montant par carte trouvé dans le fichier")
            self.confidence = 0.2

    def _extract_from_xls(self, ws):
        """Extract card settlement amounts from xls worksheet."""
        found = {}

        for row_idx in range(ws.nrows):
            row_text = ' '.join(str(ws.cell_value(row_idx, c) or '') for c in range(ws.ncols)).lower()

            for card_type, field_name in self.DAILY_REV_TO_CASHOUT.items():
                if card_type.lower() in row_text:
                    for col_idx in range(ws.ncols):
                        val = ws.cell_value(row_idx, col_idx)
                        if isinstance(val, (int, float)) and val != 0:
                            found[field_name] = abs(float(val))
                            break

        if found:
            self.extracted_data.update(found)
            self.extracted_data['source'] = 'file'
            self.confidence = min(0.5 + 0.1 * len(found), 0.9)
        else:
            self.validation_warnings.append("Aucun montant par carte trouvé dans le fichier xls")
            self.confidence = 0.2

    def _parse_from_daily_revenue(self):
        """
        Auto-fill mode: Use Daily Revenue card totals as FreedomPay values.

        Since FreedomPay bank settlements should equal Daily Revenue system
        totals (variance = 0), we can safely use DR values for both:
        - Row 6 (Daily Cash Out)
        - Row 12 (Daily Revenue)
        """
        for card_type, field_name in self.DAILY_REV_TO_CASHOUT.items():
            amount = self.daily_revenue_cards.get(card_type, 0)
            if amount:
                abs_amount = abs(float(amount))
                self.extracted_data[field_name] = abs_amount

        # Also prepare daily_revenue fields (Row 12) with same values
        # Card type mapping: visa, mastercard, amex, diners, discover
        daily_rev_mappings = {
            'visa': 'visa_daily_revenue',
            'mastercard': 'master_daily_revenue',
            'amex': 'amex_daily_revenue',
            'diners': 'diners_daily_revenue',
            'discover': 'discover_daily_revenue',
        }
        for card_type, field_name in daily_rev_mappings.items():
            amount = self.daily_revenue_cards.get(card_type, 0)
            if amount:
                self.extracted_data[field_name] = abs(float(amount))

        # Also prepare fusebox fields for transelect
        self.extracted_data['fusebox_visa'] = self.extracted_data.get('visa_cash_out', 0)
        self.extracted_data['fusebox_master'] = self.extracted_data.get('master_cash_out', 0)
        self.extracted_data['fusebox_amex'] = self.extracted_data.get('amex_cash_out', 0)

        self.extracted_data['source'] = 'daily_revenue_autofill'
        self.extracted_data['note'] = 'Valeurs copiées depuis Daily Revenue (variance attendue: $0.00)'

        if any(v for k, v in self.extracted_data.items() if k.endswith('_cash_out') and v):
            self.confidence = 0.95  # High confidence since DR is already parsed
        else:
            self.confidence = 0.3
            self.validation_warnings.append("Aucun montant par carte dans Daily Revenue")

    def get_geac_fillable(self):
        """Return data ready to fill GEAC/UX Row 6 (Daily Cash Out).

        Returns:
            dict: {cell_ref: value} for GEAC/UX sheet
        """
        if not self._parsed:
            self.parse()

        fillable = {}
        for field_key, cell_ref in self.FIELD_MAPPINGS.items():
            val = self.extracted_data.get(field_key)
            if val is not None and val != 0:
                fillable[cell_ref] = val

        return fillable

    def get_transelect_fillable(self):
        """Return data ready to fill Transelect fusebox rows.

        Returns:
            dict: {cell_ref: value} for transelect sheet
        """
        if not self._parsed:
            self.parse()

        fillable = {}
        for field_key, cell_ref in self.TRANSELECT_MAPPINGS.items():
            val = self.extracted_data.get(field_key)
            if val is not None and val != 0:
                fillable[cell_ref] = val

        return fillable

    def get_daily_revenue_fillable(self):
        """Return data ready to fill GEAC/UX Row 12 (Daily Revenue).

        Returns:
            dict: {cell_ref: value} for GEAC/UX sheet Row 12
        """
        if not self._parsed:
            self.parse()

        fillable = {}
        for field_key, cell_ref in self.DAILY_REV_MAPPINGS.items():
            val = self.extracted_data.get(field_key)
            if val is not None and val != 0:
                fillable[cell_ref] = val

        return fillable

    def validate(self):
        """Validate extracted data."""
        if not self.extracted_data:
            self.validation_errors.append("Aucune donnée FreedomPay extraite")
            return False

        # Check for negative amounts
        for key, val in self.extracted_data.items():
            if isinstance(val, (int, float)) and val < 0:
                self.validation_warnings.append(f"Montant négatif: {key} = {val}")

        # Warn if not all card types found
        card_fields = [f for f in self.FIELD_MAPPINGS if self.extracted_data.get(f)]
        if len(card_fields) < 3:
            self.validation_warnings.append(
                f"Seulement {len(card_fields)} type(s) de carte trouvé(s) sur 5"
            )

        return len(self.validation_errors) == 0
