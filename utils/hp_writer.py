"""
HP (Hotel Promotion) Writer
Writes HP/Admin entries back to the HP Excel file (données sheet).

HP files use .xlsx format (openpyxl), unlike SD files which use .xls (xlrd/xlutils).
"""

from io import BytesIO
from openpyxl import load_workbook


class HPWriter:
    """
    Writer for HP Excel files.
    Writes daily HP/Admin entries to the 'données' sheet.
    """

    # Column mapping for données sheet (0-indexed)
    DONNEES_COLS = {
        'day': 0,           # A: Day number
        'area': 1,          # B: Department/Area
        'nourriture': 2,    # C: Food
        'boisson': 3,       # D: Alcohol/Beverages
        'biere': 4,         # E: Beer
        'vin': 5,           # F: Wine
        'mineraux': 6,      # G: Non-alcoholic
        'tabagie': 7,       # H: Tobacco
        'autres': 8,        # I: Other
        'pourboire': 9,     # J: Tips
        'paiement': 10,     # K: Payment method
        'total': 11,        # L: Transaction total
        'raison': 12,       # M: Reason
        'qui': 13,          # N: Who
        'autoriser': 14,    # O: Authorized by
    }

    @staticmethod
    def write_entries(hp_bytes, day, entries):
        """
        Write HP/Admin entries to the données sheet for a specific day.

        First removes existing rows for that day, then writes new ones.

        Args:
            hp_bytes: BytesIO containing the HP Excel file
            day: Day number (1-31)
            entries: List of entry dicts:
                [
                    {
                        'area': str ('HP' or 'ADMIN' or dept name),
                        'nourriture': float,
                        'boisson': float,
                        'biere': float,
                        'vin': float,
                        'mineraux': float,
                        'autre': float,
                        'pourboire': float,
                        'raison': str,
                        'autorise_par': str,
                    },
                    ...
                ]

        Returns:
            BytesIO: Updated HP file
        """
        if not 1 <= day <= 31:
            raise ValueError(f"Day must be between 1-31, got {day}")

        hp_bytes.seek(0)
        wb = load_workbook(hp_bytes)

        # Find données sheet
        sheet_name = None
        for name in wb.sheetnames:
            if name.lower() in ('données', 'donnees', 'données '):
                sheet_name = name
                break

        if not sheet_name:
            raise Exception("Sheet 'données' not found in HP file")

        ws = wb[sheet_name]
        cols = HPWriter.DONNEES_COLS
        data_start_row = 13  # Row 12 is headers, data starts at 13

        # 1) Find and remove existing rows for this day
        rows_to_delete = []
        for row_idx in range(data_start_row, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=cols['day'] + 1).value
            if cell_val is not None:
                try:
                    if int(cell_val) == day:
                        rows_to_delete.append(row_idx)
                except (ValueError, TypeError):
                    pass

        # Delete from bottom to top to preserve row indices
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)

        # 2) Find insertion point (after the last row with day < target day)
        insert_row = data_start_row
        for row_idx in range(data_start_row, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=cols['day'] + 1).value
            if cell_val is not None:
                try:
                    if int(cell_val) < day:
                        insert_row = row_idx + 1
                    elif int(cell_val) > day:
                        break
                except (ValueError, TypeError):
                    pass

        # 3) Insert new rows and write entries
        for i, entry in enumerate(entries):
            row_idx = insert_row + i
            ws.insert_rows(row_idx)

            # Calculate total
            nourr = float(entry.get('nourriture', 0) or 0)
            boisson = float(entry.get('boisson', 0) or 0)
            biere = float(entry.get('biere', 0) or 0)
            vin = float(entry.get('vin', 0) or 0)
            mineraux = float(entry.get('mineraux', 0) or 0)
            autre = float(entry.get('autre', 0) or 0)
            pourboire = float(entry.get('pourboire', 0) or 0)
            total = round(nourr + boisson + biere + vin + mineraux + autre + pourboire, 2)

            ws.cell(row=row_idx, column=cols['day'] + 1, value=day)
            ws.cell(row=row_idx, column=cols['area'] + 1, value=entry.get('area', ''))
            ws.cell(row=row_idx, column=cols['nourriture'] + 1, value=nourr)
            ws.cell(row=row_idx, column=cols['boisson'] + 1, value=boisson)
            ws.cell(row=row_idx, column=cols['biere'] + 1, value=biere)
            ws.cell(row=row_idx, column=cols['vin'] + 1, value=vin)
            ws.cell(row=row_idx, column=cols['mineraux'] + 1, value=mineraux)
            ws.cell(row=row_idx, column=cols['tabagie'] + 1, value=autre)  # 'autre' maps to tabagie col
            ws.cell(row=row_idx, column=cols['pourboire'] + 1, value=pourboire)
            ws.cell(row=row_idx, column=cols['total'] + 1, value=total)
            ws.cell(row=row_idx, column=cols['raison'] + 1, value=entry.get('raison', ''))
            ws.cell(row=row_idx, column=cols['autoriser'] + 1, value=entry.get('autorise_par', ''))

        # Save
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
