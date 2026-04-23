import sys, json
sys.stdout.reconfigure(encoding='utf-8')

# 1. Sales Journal
print('=== SJ ===')
with open(r'K:\Audition\04 - April\15-04-2026\SALES_JOURNAL.txt', 'r', encoding='latin-1') as f:
    print(f.read())

# 2. PDFs
import pdfplumber
for name in ['DAILY_REV', 'AR_SUMMARY', 'MARKET_SEGMENT']:
    path = rf'K:\Audition\04 - April\15-04-2026\{name}.pdf'
    print(f'\n{"="*60}\n{name}\n{"="*60}')
    with pdfplumber.open(path) as pdf:
        for pi, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text:
                print(f'--- PAGE {pi+1} ---')
                print(text)

# 3. HP day 15
print(f'\n{"="*60}\nHP DAY 15\n{"="*60}')
from openpyxl import load_workbook
wb = load_workbook(r'K:\HP 2026-2027\04-April 2026\HP 04 2026.xlsx', data_only=True)
ws = wb['données']
print('HP donnees entries for April 15:')
header = [ws.cell(1, c).value for c in range(1, 16)]
print(f'  HEADER: {header}')
for r in range(2, ws.max_row+1):
    d = ws.cell(r, 1).value
    if d and hasattr(d, 'day') and d.day == 15 and d.month == 4:
        row = {header[c-1]: ws.cell(r, c).value for c in range(1, 16) if ws.cell(r, c).value is not None}
        print(f'  {row}')

# 4. RJ current state
print(f'\n{"="*60}\nRJ DAY 15 CURRENT\n{"="*60}')
import win32com.client as win32
excel = win32.Dispatch('Excel.Application')
excel.Visible = False
excel.DisplayAlerts = False
excel.AskToUpdateLinks = False
try:
    wb2 = excel.Workbooks.Open(r'K:\RJ 2026-2027\02-AVRIL 2026\Rj 15-04-2026.xls', ReadOnly=True)
    ws2 = wb2.Sheets('jour')
    # Day 15 = row 17
    print('Jour Day 15 (row 17):')
    for c in range(1, 117):
        cell = ws2.Cells(17, c)
        v = cell.Value
        if v not in (None, '', 0):
            tag = 'F' if cell.HasFormula else 'V'
            f = cell.Formula if cell.HasFormula else ''
            from functools import reduce
            def col_letter(n):
                s = ''
                while n > 0:
                    n, r = divmod(n-1, 26)
                    s = chr(65+r) + s
                return s
            print(f'  {col_letter(c)}17 [{tag}]: {f if tag=="F" else v}')

    # Check Recap, Transelect, GEAC state
    for sn in ['Recap', 'transelect', 'geac_ux']:
        ws3 = wb2.Sheets(sn)
        print(f'\n{sn} non-empty:')
        for r in range(1, ws3.UsedRange.Rows.Count+1):
            for c2 in range(1, min(ws3.UsedRange.Columns.Count+1, 35)):
                cell = ws3.Cells(r, c2)
                v = cell.Value
                if v not in (None, '', 0, 0.0):
                    tag = 'F' if cell.HasFormula else 'V'
                    print(f'  {col_letter(c2)}{r}[{tag}]: {cell.Formula if tag=="F" else v}')

    wb2.Close(SaveChanges=False)
finally:
    excel.Quit()
