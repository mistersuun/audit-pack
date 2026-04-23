"""Inspect DBRS template structure."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

path = r"C:\Users\Auditeur\Documents\Projects\audit-pack\.playwright-mcp\test-files\DBRS_2026-04-21.xlsm"
wb = openpyxl.load_workbook(path, keep_vba=True)

print("=" * 80)
print("SHEETS:", wb.sheetnames)
print("=" * 80)

print("\n--- DailyRev ---")
ws = wb["DailyRev"]
print(f"Dimensions: {ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}")
for row in range(1, min(ws.max_row + 1, 60)):
    a = ws.cell(row=row, column=1).value
    b = ws.cell(row=row, column=2).value
    c = ws.cell(row=row, column=3).value
    d = ws.cell(row=row, column=4).value
    if a is None and b is None and c is None and d is None:
        continue
    b_is_formula = isinstance(b, str) and b.startswith("=")
    flag = "F" if b_is_formula else " "
    print(f"  R{row:>3}[{flag}]: A={a!r:50}  B={b!r:30}  C={c!r:30}  D={d!r}")

print("\n--- Market Segment ---")
ws = wb["Market Segment"]
print(f"Dimensions: {ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}")
for row in range(1, min(ws.max_row + 1, 100)):
    a = ws.cell(row=row, column=1).value
    b = ws.cell(row=row, column=2).value
    c = ws.cell(row=row, column=3).value
    d = ws.cell(row=row, column=4).value
    if a is None and b is None and c is None and d is None:
        continue
    b_is_formula = isinstance(b, str) and b.startswith("=")
    flag = "F" if b_is_formula else " "
    print(f"  R{row:>3}[{flag}]: A={a!r:55}  B={b!r:40}  C={c!r:30}")
