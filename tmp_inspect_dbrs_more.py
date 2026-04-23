"""Inspect DBRS & DBRS Insertion sheets + look at Market Segment row labels in related sheets."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

path = r"C:\Users\Auditeur\Documents\Projects\audit-pack\.playwright-mcp\test-files\DBRS_2026-04-21.xlsm"
wb = openpyxl.load_workbook(path, keep_vba=True)

# DBRS sheet often references Market Segment rows; labels likely live in the DBRS report itself
for sheet_name in ["DBRS", "DBRS Insertion", "Explications"]:
    print(f"\n\n===== Sheet: {sheet_name} =====")
    ws = wb[sheet_name]
    print(f"Dimensions: {ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}")
    for row in range(1, min(ws.max_row + 1, 120)):
        for col in range(1, min(ws.max_column + 1, 12)):
            v = ws.cell(row=row, column=col).value
            if v is None or v == "":
                continue
            is_formula = isinstance(v, str) and v.startswith("=")
            flag = "F" if is_formula else " "
            col_letter = openpyxl.utils.get_column_letter(col)
            # Truncate long formulas
            vs = repr(v)
            if len(vs) > 80:
                vs = vs[:77] + "..."
            print(f"  {sheet_name[:6]}!{col_letter}{row}[{flag}]: {vs}")

# Also print full A column of Market Segment (maybe it has label hints)
print("\n\n===== Market Segment — FULL row scan (all columns) =====")
ws = wb["Market Segment"]
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=col).value
        if v is None or v == "":
            continue
        col_letter = openpyxl.utils.get_column_letter(col)
        is_formula = isinstance(v, str) and v.startswith("=")
        flag = "F" if is_formula else " "
        print(f"  MS!{col_letter}{row}[{flag}]: {v!r}")
